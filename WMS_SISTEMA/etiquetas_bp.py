from __future__ import annotations

import json
import os
import re
import sys
from functools import wraps
from io import BytesIO
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path
from threading import Lock
from typing import Any

from flask import (
    Blueprint,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from reportlab.lib.units import mm
from etiquetas_100x150 import draw_label_100x150_pdf

import db_mdb


# ---------------------------------------------------------------------------
# Blueprint
# ---------------------------------------------------------------------------

etq_bp = Blueprint("etiquetas", __name__, url_prefix="/etiq")


# ---------------------------------------------------------------------------
# Data paths
# ---------------------------------------------------------------------------

def _get_etiq_data_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


_DATA_DIR = _get_etiq_data_dir()

LEGACY_XLS_FILENAME = "CONTROLE CLIENTES COM ESTOJOS.xlsx"
CODE128_LAYOUT_CONFIG_PATH = _DATA_DIR / "etiq_code128_layout_config.json"
LABEL_LAYOUT_CONFIG_PATH = _DATA_DIR / "etiq_label_layout_config.json"
LABEL_MODELS_PATH = _DATA_DIR / "etiq_label_models.json"

LABEL_CLIENTS_TABLE = "etiq_clients"


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

LABEL_WIDTH_MM = 80
LABEL_HEIGHT_MM = 20
LABEL_SAFE_MARGIN_MM = 1

CLIENTES_REQUIRED_COLUMNS = [
    "NumeroCliente",
    "CorRoteiro",
    "HorarioRoteiro",
    "DataImpressao",
]

HEADER_ALIASES = {
    "NumeroCliente": ("NumeroCliente",),
    "NomeCliente": ("NomeCliente", "Nome", "NomeDoCliente"),
    "CorRoteiro": ("CorRoteiro", "CorDoRoteiro"),
    "HorarioRoteiro": ("HorarioRoteiro",),
    "DataImpressao": ("DataImpressao",),
    "Entregador": ("Entregador",),
}

CODE128_LAYOUT_PARAM_KEYS = (
    "page_orientation",
    "text_orientation",
    "cols",
    "rows",
    "margin_x_mm",
    "margin_y_mm",
    "barcode_width_mm",
    "barcode_height_mm",
    "barcode_offset_y_mm",
    "text_gap_mm",
    "text_font_size_pt",
)

ROUTE_COLOR_MAP = {
    "amarelo claro": "#f5e14d",
    "amarelo escuro": "#d4a017",
    "azul": "#2f80ed",
    "azul claro": "#74b9ff",
    "laranja": "#f2994a",
    "marrom": "#a52a2a",
    "rosa": "#e86a92",
    "verde": "#27ae60",
    "verde escuro": "#1a6e3c",
}

# Thread-safety globals
DB_BOOTSTRAP_LOCK = Lock()
DB_FILE_LOCK = Lock()
DB_READY_KEY: tuple[str, int, int] | None = None
CLIENTS_CACHE: list[dict[str, Any]] | None = None
CLIENTS_CACHE_KEY: tuple[str, int, int] | None = None
LEGACY_IMPORT_DONE_KEY: tuple[str, int, int] | None = None
LABEL_LAYOUT_LOCK = Lock()
LABEL_MODELS_LOCK = Lock()

TRIAGE_SECTOR = "TRIAGEM"

ETIQ_FEATURE_PERMISSIONS = (
    "etiq_criar",
    "etiq_adicionar_cliente",
    "etiq_caixinhas",
    "etiq_envio",
    "etiq_reimprimir",
)


# ---------------------------------------------------------------------------
# Access control
# ---------------------------------------------------------------------------

@etq_bp.before_request
def _check_etiq_access():
    if request.endpoint and "static" in request.endpoint:
        return
    if "user" not in session:
        flash("Faça login para continuar.", "warning")
        return redirect(url_for("login"))
    if session.get("user", "").lower() == "admin":
        return
    permissions = session.get('permissions', [])
    if isinstance(permissions, list) and ('etiquetas' in permissions or any(feature in permissions for feature in ETIQ_FEATURE_PERMISSIONS)):
        return
    flash("Acesso restrito ao setor autorizado para etiquetas ou admin.", "danger")
    return redirect(url_for("dashboard"))


# ---------------------------------------------------------------------------
# Granular permissions — cada botão da navbar do sistema de etiquetas vira
# uma permissão independente, liberável em Gerenciamento de Células / Setores.
# ---------------------------------------------------------------------------

def can_access_feature(feature: str) -> bool:
    if session.get("user", "").lower() == "admin":
        return True
    permissions = session.get("permissions", [])
    return feature in permissions if isinstance(permissions, list) else False


@etq_bp.app_template_global("etiq_can_access")
def etiq_can_access(feature: str) -> bool:
    return can_access_feature(feature)


def _first_accessible_etiq_redirect():
    if can_access_feature("etiq_criar") or can_access_feature("etiq_adicionar_cliente"):
        return redirect(url_for("etiquetas.index"))
    if can_access_feature("etiq_caixinhas"):
        return redirect(url_for("etiquetas.barcode_code128"))
    if can_access_feature("etiq_envio"):
        return redirect(url_for("etiquetas.label_100x150_new"))
    if can_access_feature("etiq_reimprimir"):
        return redirect(url_for("etiquetas.impressos_list"))
    flash("Acesso restrito para os recursos de etiquetas.", "danger")
    return redirect(url_for("dashboard"))


def _etiq_feature_required(feature: str):
    """Protege uma rota exigindo uma permissão granular do sistema de etiquetas."""
    def decorator(view_func):
        @wraps(view_func)
        def wrapped(*args, **kwargs):
            if can_access_feature(feature):
                return view_func(*args, **kwargs)
            if request.method == "GET":
                flash("Acesso restrito para este recurso de etiquetas.", "danger")
                return _first_accessible_etiq_redirect()
            return jsonify({"error": "Acesso restrito para este recurso."}), 403
        return wrapped
    return decorator


# ---------------------------------------------------------------------------
# Database bootstrap
# ---------------------------------------------------------------------------

@etq_bp.before_request
def _prepare_database():
    try:
        ensure_database_ready()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def setup_error_message(exc: Exception) -> str:
    return f"Erro de banco de dados: {exc}"


def color_name_to_hex(color_name: str) -> str:
    normalized = color_name.strip().lower()
    return ROUTE_COLOR_MAP.get(normalized, "#7a0012")


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_horario(value: Any) -> str:
    import datetime as _dt
    if value is None:
        return ""
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, _dt.timedelta):
        total = int(value.total_seconds())
        if total < 0:
            total = 0
        hours = (total // 3600) % 24
        minutes = (total % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"
    if isinstance(value, (int, float, Decimal)):
        try:
            as_decimal = Decimal(str(value))
        except (InvalidOperation, ValueError):
            return str(value).strip()
        # Excel stores time as day fraction (e.g. 14:20 -> 0.59722...)
        if Decimal("0") <= as_decimal < Decimal("1"):
            total_minutes = int((as_decimal * Decimal("1440")).quantize(Decimal("1")))
            hours = (total_minutes // 60) % 24
            minutes = total_minutes % 60
            return f"{hours:02d}:{minutes:02d}"
    raw = str(value).strip()
    if not raw:
        return ""
    if "day" in raw and "," in raw:
        raw = raw.split(",", 1)[1].strip()
    if re.fullmatch(r"\d+(?:\.0+)?", raw):
        numeric = raw.split(".", 1)[0]
        if len(numeric) in {3, 4}:
            hh = int(numeric[:-2])
            mm = int(numeric[-2:])
            if 0 <= hh <= 23 and 0 <= mm <= 59:
                return f"{hh:02d}:{mm:02d}"
    parts = raw.split(":")
    if len(parts) == 3:
        try:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        except ValueError:
            return f"{parts[0]}:{parts[1]}"
    if len(parts) == 2:
        try:
            hh = int(parts[0])
            mm = int(parts[1])
            if 0 <= hh <= 23 and 0 <= mm <= 59:
                return f"{hh:02d}:{mm:02d}"
        except ValueError:
            pass
    float_match = re.fullmatch(r"\d+\.\d+", raw)
    if float_match:
        try:
            as_decimal = Decimal(raw)
            if Decimal("0") <= as_decimal < Decimal("1"):
                total_minutes = int((as_decimal * Decimal("1440")).quantize(Decimal("1")))
                hours = (total_minutes // 60) % 24
                minutes = total_minutes % 60
                return f"{hours:02d}:{minutes:02d}"
        except InvalidOperation:
            pass
    return raw


def _normalize_numero_cliente(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        pass
    raw = str(value).strip()
    if not raw:
        return None
    try:
        as_decimal = Decimal(raw)
        if as_decimal == as_decimal.to_integral_value():
            return int(as_decimal)
    except (InvalidOperation, ValueError):
        return None
    return None


def get_db_path() -> str:
    return db_mdb.get_db_path()


def _normalize_data_impressao(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None or value == "":
        return None
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
    return None


def _resolve_column_indexes(ws: Any) -> dict[str, int]:
    first_row = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
    normalized = [_normalize_header(cell) for cell in first_row]
    resolved: dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                resolved[canonical] = normalized.index(alias) + 1
                break
    for canonical in CLIENTES_REQUIRED_COLUMNS:
        if canonical in resolved:
            continue
        next_col = ws.max_column + 1
        ws.cell(row=1, column=next_col, value=canonical)
        resolved[canonical] = next_col
    return resolved


def _resolve_existing_column_indexes(ws: Any) -> dict[str, int]:
    first_row = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
    normalized = [_normalize_header(cell) for cell in first_row]
    resolved: dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                resolved[canonical] = normalized.index(alias) + 1
                break
    missing_columns = [col for col in CLIENTES_REQUIRED_COLUMNS if col not in resolved]
    if missing_columns:
        raise ValueError(f"Colunas obrigatorias ausentes na planilha: {', '.join(missing_columns)}")
    return resolved


def _find_existing_column_index(ws: Any, *aliases: str) -> int | None:
    first_row = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
    normalized = [_normalize_header(cell) for cell in first_row]
    for alias in aliases:
        if alias in normalized:
            return normalized.index(alias) + 1
    return None


def _build_db_cache_key() -> tuple[str, int, int] | None:
    try:
        db_path = Path(get_db_path())
        stat_info = db_path.stat()
    except OSError:
        return None
    return (str(db_path.resolve()), int(stat_info.st_mtime_ns), int(stat_info.st_size))


def _table_exists(cursor: Any, table_name: str) -> bool:
    try:
        cursor.tables(table=table_name, tableType="TABLE")
        return cursor.fetchone() is not None
    except Exception:
        return False


def _column_exists(cursor: Any, table_name: str, column_name: str) -> bool:
    try:
        cursor.columns(table=table_name, column=column_name)
        return cursor.fetchone() is not None
    except Exception:
        return False


def _run_ddl_on_conn(conn: Any, sql: str) -> None:
    old_autocommit = getattr(conn, "autocommit", False)
    try:
        conn.autocommit = True
        conn.execute(sql)
    finally:
        conn.autocommit = old_autocommit


def _get_legacy_labels_path() -> Path:
    env_path = os.environ.get("XLS_DB_PATH", "").strip()
    if env_path:
        return Path(env_path)
    return _DATA_DIR / LEGACY_XLS_FILENAME


def _import_integrador_opto():
    if getattr(sys, "frozen", False):
        opto_dir = Path(sys.executable).resolve().parent / "OPTO_INTEGRATIONS"
    else:
        opto_dir = (Path(__file__).resolve().parent.parent / "OPTO_INTEGRATIONS").resolve()
    if str(opto_dir) not in sys.path:
        sys.path.insert(0, str(opto_dir))
    import importlib
    import integrador_opto as _m  # type: ignore
    return importlib.reload(_m) if getattr(_m, "__spec__", None) else _m


def _fetch_tipo_lente_from_opto(os_id: str) -> str:
    os_id = str(os_id or "").strip()
    if not os_id:
        return ""
    try:
        opto = _import_integrador_opto()
        opto.init_database()
        if hasattr(opto, "resolve_txt_fields"):
            _, fields = opto.resolve_txt_fields(os_id)
        else:
            txt_path = opto.find_txt(os_id)
            fields = opto.parse_txt(txt_path)
        bd_prod = getattr(opto, "BD_PROD", {}) or {}
        fallback_codigo = ""
        for field_idx in (35, 36):
            codigo = str(fields.get(field_idx, "") or "").strip()
            if not codigo:
                continue
            if not fallback_codigo:
                fallback_codigo = codigo
            tipo_lente = str((bd_prod.get(codigo) or {}).get("C", "") or "").strip()
            if tipo_lente:
                return tipo_lente
        if fallback_codigo:
            return f"COD {fallback_codigo}"
        row = opto.build_row(fields)
        return str(row[4] or "").strip() if len(row) > 4 else ""
    except Exception:
        return ""


def _ensure_labels_schema(conn: Any) -> None:
    cursor = conn.cursor()
    if not _table_exists(cursor, LABEL_CLIENTS_TABLE):
        _run_ddl_on_conn(
            conn,
            f"""
            CREATE TABLE {LABEL_CLIENTS_TABLE} (
                id COUNTER PRIMARY KEY,
                numero_cliente LONG,
                nome_cliente TEXT(255),
                cor_roteiro TEXT(100),
                horario_roteiro TEXT(20),
                entregador TEXT(100),
                data_impressao TEXT(50),
                created_at TEXT(50),
                updated_at TEXT(50)
            )
            """,
        )

    expected_columns = {
        "numero_cliente": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN numero_cliente LONG",
        "nome_cliente": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN nome_cliente TEXT(255)",
        "endereco": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN endereco TEXT(255)",
        "numero": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN numero TEXT(50)",
        "complemento": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN complemento TEXT(100)",
        "bairro": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN bairro TEXT(100)",
        "cidade": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN cidade TEXT(100)",
        "estado": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN estado TEXT(50)",
        "cep": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN cep TEXT(20)",
        "cnpj": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN cnpj TEXT(50)",
        "cor_roteiro": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN cor_roteiro TEXT(100)",
        "horario_roteiro": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN horario_roteiro TEXT(20)",
        "entregador": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN entregador TEXT(100)",
        "data_impressao": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN data_impressao TEXT(50)",
        "created_at": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN created_at TEXT(50)",
        "updated_at": f"ALTER TABLE {LABEL_CLIENTS_TABLE} ADD COLUMN updated_at TEXT(50)",
    }
    for column_name, ddl in expected_columns.items():
        if not _column_exists(cursor, LABEL_CLIENTS_TABLE, column_name):
            try:
                _run_ddl_on_conn(conn, ddl)
            except Exception:
                pass
    conn.commit()

    # Ensure 'roteiros' table exists to persist created routes
    try:
        if not _table_exists(cursor, 'roteiros'):
            _run_ddl_on_conn(
                conn,
                f"""
                CREATE TABLE roteiros (
                    id COUNTER PRIMARY KEY,
                    nome TEXT(255),
                    cor TEXT(100),
                    horario TEXT(20),
                    entregador TEXT(100),
                    created_at TEXT(50),
                    updated_at TEXT(50)
                )
                """,
            )
    except Exception:
        # ignore failures here; table creation is best-effort
        pass
    # Ensure 'roteiro_cores' table exists to persist custom route colors
    try:
        if not _table_exists(cursor, 'roteiro_cores'):
            _run_ddl_on_conn(
                conn,
                f"""
                CREATE TABLE roteiro_cores (
                    id COUNTER PRIMARY KEY,
                    nome TEXT(100),
                    hex TEXT(20),
                    created_at TEXT(50),
                    updated_at TEXT(50)
                )
                """,
            )
    except Exception:
        pass


def _import_legacy_labels_if_needed(conn: Any) -> None:
    global LEGACY_IMPORT_DONE_KEY
    legacy_path = _get_legacy_labels_path()
    legacy_key = _build_db_cache_key() if legacy_path == Path(get_db_path()) else None
    if not legacy_path.exists():
        return
    try:
        cursor = conn.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM {LABEL_CLIENTS_TABLE}")
        if int(cursor.fetchone()[0] or 0) > 0:
            if legacy_key is not None:
                LEGACY_IMPORT_DONE_KEY = legacy_key
            return
    except Exception:
        return

    try:
        from openpyxl import load_workbook
    except Exception:
        return

    try:
        workbook = load_workbook(legacy_path, read_only=True, data_only=False)
    except Exception:
        return

    worksheet = workbook[workbook.sheetnames[0]]
    try:
        columns = _resolve_existing_column_indexes(worksheet)
        inserted = 0
        cursor = conn.cursor()
        for row_idx in range(2, worksheet.max_row + 1):
            numero_cliente = _normalize_numero_cliente(worksheet.cell(row=row_idx, column=columns["NumeroCliente"]).value)
            if numero_cliente is None:
                continue
            record = {
                "numero_cliente": numero_cliente,
                "cor_roteiro": _normalize_text(worksheet.cell(row=row_idx, column=columns["CorRoteiro"]).value),
                "horario_roteiro": _normalize_horario(worksheet.cell(row=row_idx, column=columns["HorarioRoteiro"]).value),
                "data_impressao": _normalize_data_impressao(worksheet.cell(row=row_idx, column=columns["DataImpressao"]).value),
                "nome_cliente": _normalize_text(
                    worksheet.cell(row=row_idx, column=columns["NomeCliente"]).value
                    if columns.get("NomeCliente") else None
                ),
                "entregador": _normalize_text(
                    worksheet.cell(row=row_idx, column=columns["Entregador"]).value
                    if columns.get("Entregador") else None
                ),
            }
            cursor.execute(
                f"""
                INSERT INTO {LABEL_CLIENTS_TABLE} (
                    numero_cliente, nome_cliente, cor_roteiro, horario_roteiro,
                    entregador, data_impressao, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(record["numero_cliente"]),
                    record.get("nome_cliente") or None,
                    str(record.get("cor_roteiro") or "").strip().upper(),
                    record.get("horario_roteiro") or None,
                    record.get("entregador") or None,
                    record["data_impressao"].strftime("%Y-%m-%d %H:%M:%S") if record.get("data_impressao") else None,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            inserted += 1
        conn.commit()
        if inserted and legacy_key is not None:
            LEGACY_IMPORT_DONE_KEY = legacy_key
    finally:
        workbook.close()


def _invalidate_clients_cache() -> None:
    global CLIENTS_CACHE, CLIENTS_CACHE_KEY
    CLIENTS_CACHE = None
    CLIENTS_CACHE_KEY = None


def ensure_database_ready() -> None:
    global DB_READY_KEY
    db_key = _build_db_cache_key()
    if db_key is not None and DB_READY_KEY == db_key:
        return
    with DB_BOOTSTRAP_LOCK:
        db_key = _build_db_cache_key()
        if db_key is not None and DB_READY_KEY == db_key:
            return
        conn = db_mdb.get_connection()
        _ensure_labels_schema(conn)
        _import_legacy_labels_if_needed(conn)
        DB_READY_KEY = _build_db_cache_key() or db_key


def fetch_client_label_base(numero_cliente: int) -> tuple[Any, Any, Any, Any]:
    ensure_database_ready()
    conn = db_mdb.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        f"""
        SELECT TOP 1 numero_cliente, cor_roteiro, horario_roteiro, entregador
        FROM {LABEL_CLIENTS_TABLE}
        WHERE numero_cliente = ?
        ORDER BY id DESC
        """,
        (int(numero_cliente),),
    )
    row = cursor.fetchone()
    if not row:
        return None, None, None, None
    return (
        _normalize_numero_cliente(row[0]),
        _normalize_text(row[1]),
        _normalize_horario(row[2]),
        _normalize_text(row[3]),
    )


def fetch_all_clients() -> list[dict[str, Any]]:
    global CLIENTS_CACHE, CLIENTS_CACHE_KEY
    ensure_database_ready()
    cache_key = _build_db_cache_key()
    if CLIENTS_CACHE is not None and CLIENTS_CACHE_KEY is not None and cache_key == CLIENTS_CACHE_KEY:
        return [row.copy() for row in CLIENTS_CACHE]
    with DB_FILE_LOCK:
        cache_key = _build_db_cache_key()
        if CLIENTS_CACHE is not None and CLIENTS_CACHE_KEY is not None and cache_key == CLIENTS_CACHE_KEY:
            return [row.copy() for row in CLIENTS_CACHE]
        conn = db_mdb.get_connection()
        cursor = conn.cursor()
        cursor.execute(
            f"""
            SELECT numero_cliente, nome_cliente, cor_roteiro, horario_roteiro, entregador, data_impressao
            FROM {LABEL_CLIENTS_TABLE}
            ORDER BY numero_cliente ASC, id ASC
            """
        )
        rows: list[dict[str, Any]] = []
        for row in cursor.fetchall():
            rows.append(
                {
                    "numero_cliente": _normalize_numero_cliente(row[0]),
                    "nome_cliente": _normalize_text(row[1]),
                    "cor_roteiro": _normalize_text(row[2]),
                    "horario_roteiro": _normalize_horario(row[3]),
                    "entregador": _normalize_text(row[4]),
                    "data_impressao": _normalize_data_impressao(row[5]),
                }
            )
    CLIENTS_CACHE = [row.copy() for row in rows]
    CLIENTS_CACHE_KEY = cache_key
    return [row.copy() for row in rows]


def fetch_clients_filtered(filter_color: str, filter_data: str) -> list[dict[str, Any]]:
    clients = fetch_all_clients()
    target_date = None
    if filter_data:
        target_date = datetime.strptime(filter_data, "%Y-%m-%d").date()
    filtered: list[dict[str, Any]] = []
    for client in clients:
        if filter_color and filter_color.lower() not in str(client["cor_roteiro"]).lower():
            continue
        data_impressao = client.get("data_impressao")
        if target_date:
            if not isinstance(data_impressao, datetime) or data_impressao.date() != target_date:
                continue
        filtered.append(client)
    return filtered


# ---------------------------------------------------------------------------
# Impressos/ — helper to persist every generated PDF
# ---------------------------------------------------------------------------

def _save_to_impressos(pdf_bytes: bytes, label: str) -> Path | None:
    """Save *pdf_bytes* to _DATA_DIR/Impressos/<timestamp>_<label>.pdf.
    Returns the path on success, None on failure (non-fatal).
    """
    try:
        impressos_dir = _DATA_DIR / "Impressos"
        impressos_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = impressos_dir / f"{ts}_{label}.pdf"
        dest.write_bytes(pdf_bytes)
        return dest
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 100x150 label routes
# ---------------------------------------------------------------------------


@etq_bp.route("/etiquetas/envio")
@_etiq_feature_required("etiq_envio")
def label_100x150_new():
    """Direct access from topbar — empty form, auto-fills enviado_por from session."""
    return render_template(
        "etiq/label_print_100x150.html",
        label_id=0,
        id_master=request.args.get("id_master", ""),
        enviado_por=session.get("user", ""),
        os_id=request.args.get("os_id", ""),
        endereco=request.args.get("endereco", ""),
        tratamento=request.args.get("tratamento", ""),
        caixa=request.args.get("caixa", ""),
        od_esf=request.args.get("od_esf", ""),
        od_cil=request.args.get("od_cil", ""),
        od_eixo=request.args.get("od_eixo", ""),
        od_ad=request.args.get("od_ad", ""),
        oe_esf=request.args.get("oe_esf", ""),
        oe_cil=request.args.get("oe_cil", ""),
        oe_eixo=request.args.get("oe_eixo", ""),
        oe_ad=request.args.get("oe_ad", ""),
    )


@etq_bp.route("/etiquetas/envio/pdf")
@_etiq_feature_required("etiq_envio")
def label_envio_pdf_quick():
    """Quick PDF from query params (used by auto-print modal on dashboard)."""
    rq = request.args
    from datetime import datetime as _dt
    data = {
        "id_master": rq.get("id_master", ""),
        "os_id": rq.get("os_id", ""),
        "endereco": rq.get("endereco", ""),
        "tratamento": rq.get("tratamento", ""),
        "tipo_lente": rq.get("tipo_lente", ""),
        "fotossensibilidade": rq.get("fotossensibilidade", ""),
        "material": rq.get("material", ""),
        "caixa": rq.get("caixa", ""),
        "enviado_por": rq.get("enviado_por") or session.get("user", ""),
        "od_esf": rq.get("od_esf", ""),
        "od_cil": rq.get("od_cil", ""),
        "od_eixo": rq.get("od_eixo", ""),
        "od_ad": rq.get("od_ad", ""),
        "oe_esf": rq.get("oe_esf", ""),
        "oe_cil": rq.get("oe_cil", ""),
        "oe_eixo": rq.get("oe_eixo", ""),
        "oe_ad": rq.get("oe_ad", ""),
        "data_impressao": _dt.now().strftime("%d/%m/%Y %H:%M"),
    }
    buf = draw_label_100x150_pdf(data)
    pdf_bytes = buf.getvalue()
    os_id = rq.get("os_id") or rq.get("id_master") or "envio"
    _save_to_impressos(pdf_bytes, f"envio_{os_id}")
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=False,
                     download_name=f"etiqueta_envio_{os_id}.pdf")


@etq_bp.route("/etiquetas/print_100x150/<int:id>")
@_etiq_feature_required("etiq_envio")
def label_100x150_view(id: int):
    """Preview page with form for all label fields."""
    numero, _cor, _horario, entregador = fetch_client_label_base(id)
    return render_template(
        "etiq/label_print_100x150.html",
        label_id=id,
        id_master=numero or "",
        enviado_por=entregador or session.get("user", ""),
        os_id=request.args.get("os_id", ""),
        endereco=request.args.get("endereco", ""),
        tratamento=request.args.get("tratamento", ""),
        tipo_lente=request.args.get("tipo_lente", ""),
        fotossensibilidade=request.args.get("fotossensibilidade", ""),
        material=request.args.get("material", ""),
        caixa=request.args.get("caixa", ""),
        od_esf=request.args.get("od_esf", ""),
        od_cil=request.args.get("od_cil", ""),
        od_eixo=request.args.get("od_eixo", ""),
        od_ad=request.args.get("od_ad", ""),
        oe_esf=request.args.get("oe_esf", ""),
        oe_cil=request.args.get("oe_cil", ""),
        oe_eixo=request.args.get("oe_eixo", ""),
        oe_ad=request.args.get("oe_ad", ""),
    )


@etq_bp.route("/etiquetas/print_100x150/<int:id>/pdf")
@_etiq_feature_required("etiq_envio")
def label_100x150_pdf(id: int):
    """Generate and return a 150x100mm landscape PDF, saving it to Impressos/."""
    numero, _cor, _horario, entregador = fetch_client_label_base(id)
    rq = request.args
    from datetime import datetime as _dt
    data = {
        "id_master": rq.get("id_master") or str(numero or ""),
        "os_id": rq.get("os_id", ""),
        "endereco": rq.get("endereco", ""),
        "tratamento": rq.get("tratamento", ""),
        "tipo_lente": rq.get("tipo_lente", ""),
        "fotossensibilidade": rq.get("fotossensibilidade", ""),
        "material": rq.get("material", ""),
        "caixa": rq.get("caixa", ""),
        "enviado_por": rq.get("enviado_por") or entregador or session.get("user", ""),
        "od_esf": rq.get("od_esf", ""),
        "od_cil": rq.get("od_cil", ""),
        "od_eixo": rq.get("od_eixo", ""),
        "od_ad": rq.get("od_ad", ""),
        "oe_esf": rq.get("oe_esf", ""),
        "oe_cil": rq.get("oe_cil", ""),
        "oe_eixo": rq.get("oe_eixo", ""),
        "oe_ad": rq.get("oe_ad", ""),
        "data_impressao": _dt.now().strftime("%d/%m/%Y %H:%M"),
    }
    buf = draw_label_100x150_pdf(data)
    pdf_bytes = buf.getvalue()
    os_id = data.get("os_id") or data.get("id_master") or str(id)
    _save_to_impressos(pdf_bytes, f"envio_{os_id}_id{id}")
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=f"etiqueta_{id}_paisagem.pdf",
    )


@etq_bp.route("/etiquetas/expedicao/preview")
@_etiq_feature_required("etiq_envio")
def expedicao_label_preview():
    """Preview rápido da etiqueta de expedição criada para o módulo de expedição.

    Aceita query params: `cliente_nome`, `cliente_codigo`, `cliente_endereco`.
    """
    cliente_nome = request.args.get("cliente_nome", "")
    cliente_codigo = request.args.get("cliente_codigo", "")
    cliente_endereco = request.args.get("cliente_endereco", "")
    return render_template(
        "etiq/expedicao_label_100x150.html",
        cliente_nome=cliente_nome,
        cliente_codigo=cliente_codigo,
        cliente_endereco=cliente_endereco,
    )


@etq_bp.route("/etiquetas/impressos")
@_etiq_feature_required("etiq_reimprimir")
def impressos_list():
    """List all saved PDFs in Impressos/ for reprinting."""
    impressos_dir = _DATA_DIR / "Impressos"
    files: list[dict] = []
    if impressos_dir.exists():
        for p in sorted(impressos_dir.glob("*.pdf"), reverse=True):
            stat = p.stat()
            files.append({
                "name": p.name,
                "size_kb": round(stat.st_size / 1024, 1),
                "mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M:%S"),
            })
    return render_template("etiq/impressos_list.html", files=files)


@etq_bp.route("/etiquetas/impressos/<path:filename>")
@_etiq_feature_required("etiq_reimprimir")
def impressos_serve(filename: str):
    """Serve a saved PDF from Impressos/ for view / reprint / download."""
    impressos_dir = (_DATA_DIR / "Impressos").resolve()
    # Security: ensure the resolved path is inside Impressos/ (prevent path traversal)
    target = (impressos_dir / filename).resolve()
    if impressos_dir not in target.parents and target != impressos_dir:
        return "Acesso negado.", 403
    if not target.exists() or not target.is_file():
        return "Arquivo não encontrado.", 404
    as_attachment = request.args.get("dl") == "1"
    return send_file(str(target), mimetype="application/pdf",
                     as_attachment=as_attachment, download_name=target.name)


def delete_client(numero_cliente: int) -> None:
    ensure_database_ready()
    with DB_FILE_LOCK:
        conn = db_mdb.get_connection()
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM {LABEL_CLIENTS_TABLE} WHERE numero_cliente = ?", (int(numero_cliente),))
        conn.commit()
        _invalidate_clients_cache()


def upsert_client(
    numero_cliente: int,
    cor_roteiro: str,
    horario_roteiro: str,
    nome_cliente: str = "",
    entregador: str = "",
    endereco: str = "",
    numero: str = "",
    complemento: str = "",
    bairro: str = "",
    cidade: str = "",
    estado: str = "",
    cep: str = "",
    cnpj: str = "",
) -> None:
    ensure_database_ready()
    numero_cliente_norm = _normalize_numero_cliente(numero_cliente)
    if numero_cliente_norm is None:
        raise ValueError("Numero do cliente invalido.")
    horario_roteiro_norm = _normalize_horario(horario_roteiro)
    with DB_FILE_LOCK:
        conn = db_mdb.get_connection()
        cursor = conn.cursor()
        cursor.execute(
            f"""
            SELECT TOP 1 id, data_impressao, created_at
            FROM {LABEL_CLIENTS_TABLE}
            WHERE numero_cliente = ?
            ORDER BY id DESC
            """,
            (int(numero_cliente_norm),),
        )
        row = cursor.fetchone()
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if row:
            cursor.execute(
                f"""
                UPDATE {LABEL_CLIENTS_TABLE}
                SET nome_cliente = ?, cor_roteiro = ?, horario_roteiro = ?, entregador = ?,
                    endereco = ?, numero = ?, complemento = ?, bairro = ?, cidade = ?, estado = ?, cep = ?, cnpj = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    nome_cliente or None,
                    cor_roteiro.upper(),
                    horario_roteiro_norm or None,
                    entregador or None,
                    endereco or None,
                    numero or None,
                    complemento or None,
                    bairro or None,
                    cidade or None,
                    estado or None,
                    cep or None,
                    cnpj or None,
                    now_str,
                    int(row[0]),
                ),
            )
        else:
            cursor.execute(
                f"""
                INSERT INTO {LABEL_CLIENTS_TABLE} (
                    numero_cliente, nome_cliente, cor_roteiro, horario_roteiro,
                    entregador, endereco, numero, complemento, bairro, cidade, estado, cep, cnpj,
                    data_impressao, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(numero_cliente_norm),
                    nome_cliente or None,
                    cor_roteiro.upper(),
                    horario_roteiro_norm or None,
                    entregador or None,
                    endereco or None,
                    numero or None,
                    complemento or None,
                    bairro or None,
                    cidade or None,
                    estado or None,
                    cep or None,
                    cnpj or None,
                    None,
                    now_str,
                    now_str,
                ),
            )
        conn.commit()
        _invalidate_clients_cache()


def _persist_print_date(numero_cliente: int, print_dt: datetime) -> None:
    with DB_FILE_LOCK:
        conn = db_mdb.get_connection()
        cursor = conn.cursor()
        cursor.execute(
            f"SELECT TOP 1 id FROM {LABEL_CLIENTS_TABLE} WHERE numero_cliente = ? ORDER BY id DESC",
            (int(numero_cliente),),
        )
        row = cursor.fetchone()
        if row is None:
            return
        timestamp = print_dt.strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            f"""
            UPDATE {LABEL_CLIENTS_TABLE}
            SET data_impressao = ?, updated_at = ?
            WHERE id = ?
            """,
            (timestamp, timestamp, int(row[0])),
        )
        conn.commit()
        _invalidate_clients_cache()


def _create_or_get_roteiro(cor: str, entregador: str, horario: str, nome: str | None = None) -> int | None:
    """Create a roteiro row if not exists and return its id (or None on error)."""
    try:
        ensure_database_ready()
        with DB_FILE_LOCK:
            conn = db_mdb.get_connection()
            cursor = conn.cursor()
            # normalize
            cor_n = (cor or "").strip()
            entregador_n = (entregador or "").strip()
            horario_n = _normalize_horario(horario)
            nome_n = (nome or "").strip() or None

            # try to find an existing roteiro with same cor+entregador+horario
            cursor.execute(
                "SELECT TOP 1 id FROM roteiros WHERE cor = ? AND entregador = ? AND horario = ?",
                (cor_n, entregador_n, horario_n),
            )
            row = cursor.fetchone()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if row:
                return int(row[0])
            # insert new
            cursor.execute(
                "INSERT INTO roteiros (nome, cor, horario, entregador, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
                (nome_n, cor_n or None, horario_n or None, entregador_n or None, now, now),
            )
            conn.commit()
            # fetch id of last inserted
            cursor.execute("SELECT TOP 1 id FROM roteiros ORDER BY id DESC")
            r = cursor.fetchone()
            return int(r[0]) if r else None
    except Exception:
        return None


def _create_or_get_cor(nome: str, hex_value: str | None = None) -> dict | None:
    """Create or return existing color record. Returns dict with id,nome,hex or None on error."""
    try:
        ensure_database_ready()
        with DB_FILE_LOCK:
            conn = db_mdb.get_connection()
            cursor = conn.cursor()
            nome_n = (nome or "").strip()
            hex_n = (hex_value or "").strip()
            if not nome_n:
                return None
            # try to find existing by name (case-insensitive)
            cursor.execute("SELECT TOP 1 id, nome, hex FROM roteiro_cores WHERE LCASE(nome) = ?", (nome_n.lower(),))
            row = cursor.fetchone()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if row:
                return {"id": int(row[0]), "nome": row[1], "hex": row[2]}
            cursor.execute(
                "INSERT INTO roteiro_cores (nome, hex, created_at, updated_at) VALUES (?, ?, ?, ?)",
                (nome_n, hex_n or None, now, now),
            )
            conn.commit()
            cursor.execute("SELECT TOP 1 id, nome, hex FROM roteiro_cores ORDER BY id DESC")
            r = cursor.fetchone()
            if not r:
                return None
            return {"id": int(r[0]), "nome": r[1], "hex": r[2]}
    except Exception:
        return None


def build_label_data(numero_cliente: int, persist_print_date: bool = True) -> dict[str, Any] | None:
    row_numero_cliente, row_cor_roteiro, row_horario_roteiro, row_entregador = fetch_client_label_base(numero_cliente)
    if not row_numero_cliente:
        return None
    print_dt = datetime.now()
    if persist_print_date:
        ensure_database_ready()
        _persist_print_date(row_numero_cliente, print_dt)
    return {
        "numero_cliente": row_numero_cliente,
        "data_impressao": print_dt.strftime("%d/%m/%Y %H:%M"),
        "cor_roteiro": row_cor_roteiro,
        "cor_hex": color_name_to_hex(str(row_cor_roteiro)),
        "horario_roteiro": row_horario_roteiro,
        "entregador": row_entregador or "",
    }


def build_label_data_by_os(os_id: int, numero_cliente: int, persist_print_date: bool = True, fetch_tipo: bool = True) -> dict[str, Any] | None:
    label_data = build_label_data(numero_cliente, persist_print_date=persist_print_date)
    if not label_data:
        return None
    label_data["os_id"] = os_id
    label_data["show_tipo_lente"] = fetch_tipo
    if fetch_tipo:
        label_data["tipo_lente"] = _fetch_tipo_lente_from_opto(str(os_id))
    else:
        label_data["tipo_lente"] = ""
    return label_data


def draw_label_pdf(pdf: Any, label: dict[str, Any], x: float, y: float) -> None:
    from reportlab.graphics.barcode import code128
    info_x = x
    barcode_x = x + (47 * mm)
    if "os_id" in label:
        barcode_value = str(label["os_id"])
        barcode = code128.Code128(barcode_value, barHeight=10.5 * mm, barWidth=0.34)
        barcode.drawOn(pdf, barcode_x, y - (12.2 * mm))
        pdf.setFont("Helvetica-Bold", 15)
        pdf.drawString(barcode_x + (2 * mm), y - (14.6 * mm), barcode_value)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.setLineWidth(0.3)
    pdf.rect(info_x, y - (5.9 * mm), 36.5 * mm, 4.5 * mm, stroke=1, fill=0)
    pdf.rect(info_x, y - (11.2 * mm), 36.5 * mm, 4.5 * mm, stroke=1, fill=0)
    pdf.drawString(info_x + 0.8 * mm, y - (2.75 * mm), str(label["horario_roteiro"]))
    pdf.drawString(info_x + 0.8 * mm, y - (8.05 * mm), str(label["cor_roteiro"]))
    pdf.setFont("Helvetica-Bold", 8.5)
    pdf.drawString(info_x, y - (14.6 * mm), f"CLI: {label['numero_cliente']}")
    dt_y = 17.8
    pdf.setFont("Helvetica-Bold", 7.6)
    pdf.drawString(info_x, y - (dt_y * mm), f"DT: {label['data_impressao']}")
    pdf.setFillColor(label.get("cor_hex", "#7a0012"))
    pdf.rect(info_x + (27 * mm), y - (17.4 * mm), 15 * mm, 7 * mm, stroke=0, fill=1)
    pdf.setFillColor("black")


def draw_label_caixinha_pdf(pdf: Any, label: dict[str, Any], x: float, y: float, width: float = 60 * mm, height: float = 30 * mm) -> None:
    from reportlab.graphics.barcode import code128
    numero_cliente = label.get("numero_cliente", "")
    pdf.setLineWidth(0.5)
    pdf.rect(x, y - height, width, height, stroke=1, fill=0)
    if numero_cliente:
        barcode_value = str(numero_cliente)
        barcode = code128.Code128(barcode_value, barHeight=8 * mm, barWidth=0.3)
        barcode.drawOn(pdf, x + (width / 2) - (25 * mm), y - height + (15 * mm))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(x + (width / 2) - (10 * mm), y - height + (4 * mm), barcode_value)


def build_labels_from_clients(clients: list[dict[str, Any]]) -> list[dict[str, Any]]:
    labels: list[dict[str, Any]] = []
    for client in clients:
        numero_cliente = int(client["numero_cliente"])
        label = build_label_data(numero_cliente)
        if label:
            labels.append(label)
    return labels


def parse_box_numbers(raw_values: str) -> list[str]:
    return [value.strip() for value in raw_values.splitlines() if value.strip()]


def _parse_int_value(raw_value: Any, default: int, min_value: int, max_value: int) -> int:
    try:
        value = int(str(raw_value).strip())
    except (TypeError, ValueError):
        value = default
    return max(min_value, min(max_value, value))


def _parse_float_value(raw_value: Any, default: float, min_value: float, max_value: float) -> float:
    try:
        value = float(str(raw_value).strip())
    except (TypeError, ValueError):
        value = default
    return max(min_value, min(max_value, value))


def parse_code128_layout_params(source: Any) -> dict[str, Any]:
    page_orientation = str(source.get("page_orientation", "portrait")).strip().lower()
    if page_orientation not in {"portrait", "landscape"}:
        page_orientation = "portrait"
    text_orientation = str(source.get("text_orientation", "horizontal")).strip().lower()
    if text_orientation not in {"horizontal", "vertical"}:
        text_orientation = "horizontal"
    return {
        "page_orientation": page_orientation,
        "cols": _parse_int_value(source.get("cols"), default=3, min_value=1, max_value=12),
        "rows": _parse_int_value(source.get("rows"), default=8, min_value=1, max_value=12),
        "margin_x_mm": _parse_float_value(source.get("margin_x_mm"), default=15.0, min_value=0.0, max_value=50.0),
        "margin_y_mm": _parse_float_value(source.get("margin_y_mm"), default=15.0, min_value=0.0, max_value=50.0),
        "barcode_width_mm": _parse_float_value(source.get("barcode_width_mm"), default=36.0, min_value=10.0, max_value=80.0),
        "barcode_height_mm": _parse_float_value(source.get("barcode_height_mm"), default=21.0, min_value=8.0, max_value=80.0),
        "barcode_offset_y_mm": _parse_float_value(source.get("barcode_offset_y_mm"), default=0.0, min_value=-20.0, max_value=20.0),
        "text_gap_mm": _parse_float_value(source.get("text_gap_mm"), default=2.5, min_value=0.0, max_value=20.0),
        "text_font_size_pt": _parse_float_value(source.get("text_font_size_pt"), default=11.0, min_value=6.0, max_value=24.0),
        "text_orientation": text_orientation,
    }


def load_code128_layout_config() -> dict[str, Any]:
    if not CODE128_LAYOUT_CONFIG_PATH.exists():
        return parse_code128_layout_params({})
    try:
        data = json.loads(CODE128_LAYOUT_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return parse_code128_layout_params({})
    if not isinstance(data, dict):
        return parse_code128_layout_params({})
    return parse_code128_layout_params(data)


def save_code128_layout_config(layout_config: dict[str, Any]) -> None:
    CODE128_LAYOUT_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    CODE128_LAYOUT_CONFIG_PATH.write_text(
        json.dumps(layout_config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _build_code128_query(values: str, layout_config: dict[str, Any], **extra: Any) -> dict[str, Any]:
    query: dict[str, Any] = {"values": values}
    for key in CODE128_LAYOUT_PARAM_KEYS:
        query[key] = layout_config[key]
    query.update(extra)
    return query


def load_label_models() -> list[dict[str, Any]]:
    if not LABEL_MODELS_PATH.exists():
        return []
    try:
        data = json.loads(LABEL_MODELS_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_label_models(models: list[dict[str, Any]]) -> None:
    LABEL_MODELS_PATH.parent.mkdir(parents=True, exist_ok=True)
    LABEL_MODELS_PATH.write_text(
        json.dumps(models, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_label_layout_config() -> dict[str, Any]:
    if not LABEL_LAYOUT_CONFIG_PATH.exists():
        return {}
    try:
        data = json.loads(LABEL_LAYOUT_CONFIG_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_label_layout_config(layout_config: dict[str, Any]) -> None:
    LABEL_LAYOUT_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    LABEL_LAYOUT_CONFIG_PATH.write_text(
        json.dumps(layout_config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def clear_label_layout_config() -> None:
    try:
        LABEL_LAYOUT_CONFIG_PATH.unlink()
    except FileNotFoundError:
        pass


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@etq_bp.get("/")
@_etiq_feature_required("etiq_criar")
def index():
    error = request.args.get("error")
    success = request.args.get("success")
    try:
        ensure_database_ready()
    except Exception as exc:
        error = setup_error_message(exc)
    return render_template(
        "etiq/index.html",
        error=error,
        success=success,
        active_tab="dashboard",
        route_colors=sorted(ROUTE_COLOR_MAP.items()),
        layout_config=load_label_layout_config(),
    )


@etq_bp.get("/codigo-barras/code128")
@_etiq_feature_required("etiq_caixinhas")
def barcode_code128():
    error = request.args.get("error")
    success = request.args.get("success")
    raw_values = request.args.get("values", "").strip()
    if not raw_values:
        raw_values = request.args.get("value", "").strip()
    has_custom_layout = any(key in request.args for key in CODE128_LAYOUT_PARAM_KEYS)
    if has_custom_layout:
        layout_config = parse_code128_layout_params(request.args)
    else:
        layout_config = load_code128_layout_config()
    return render_template(
        "etiq/barcode_code128.html",
        raw_values=raw_values,
        barcode_values=parse_box_numbers(raw_values),
        layout_config=layout_config,
        error=error,
        success=success,
        active_tab="code128",
    )


@etq_bp.post("/codigo-barras/code128/layout/salvar")
@_etiq_feature_required("etiq_caixinhas")
def barcode_code128_save_layout():
    raw_values = request.form.get("values", "").strip()
    layout_config = parse_code128_layout_params(request.form)
    save_code128_layout_config(layout_config)
    query = _build_code128_query(raw_values, layout_config, success="Layout salvo com sucesso.")
    return redirect(url_for("etiquetas.barcode_code128", **query))


@etq_bp.post("/codigo-barras/code128/pdf")
@_etiq_feature_required("etiq_caixinhas")
def barcode_code128_pdf():
    raw_values = request.form.get("values", "").strip()
    barcode_values = parse_box_numbers(raw_values)
    layout_config = parse_code128_layout_params(request.form)
    save_code128_layout_config(layout_config)
    if not barcode_values:
        query = _build_code128_query(raw_values, layout_config, error="Nenhum codigo informado.")
        return redirect(url_for("etiquetas.barcode_code128", **query))
    cols = layout_config["cols"]
    rows = layout_config["rows"]
    margin_x = layout_config["margin_x_mm"] * mm
    margin_y = layout_config["margin_y_mm"] * mm
    if layout_config["page_orientation"] == "landscape":
        page_width = 297 * mm
        page_height = 210 * mm
    else:
        page_width = 210 * mm
        page_height = 297 * mm
    available_width = page_width - (2 * margin_x)
    available_height = page_height - (2 * margin_y)
    if available_width <= 0 or available_height <= 0:
        query = _build_code128_query(raw_values, layout_config, error="Margens invalidas para a pagina.")
        return redirect(url_for("etiquetas.barcode_code128", **query))
    cell_width = available_width / cols
    cell_height = available_height / rows
    target_barcode_width = layout_config["barcode_width_mm"] * mm
    target_barcode_height = layout_config["barcode_height_mm"] * mm
    barcode_offset_y = layout_config["barcode_offset_y_mm"] * mm
    text_gap = layout_config["text_gap_mm"] * mm
    text_font_size = layout_config["text_font_size_pt"]
    text_orientation = layout_config["text_orientation"]
    codes_per_page = cols * rows
    from reportlab.pdfgen import canvas
    from reportlab.graphics.barcode import code128
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=(page_width, page_height))
    for idx, barcode_value in enumerate(barcode_values):
        cell_idx = idx % codes_per_page
        if cell_idx == 0 and idx > 0:
            pdf.showPage()
        row = cell_idx // cols
        col = cell_idx % cols
        x = margin_x + (col * cell_width)
        y = page_height - margin_y - ((row + 1) * cell_height)
        probe_barcode = code128.Code128(str(barcode_value), barHeight=target_barcode_height, barWidth=1.0)
        adjusted_bar_width = target_barcode_width / probe_barcode.width
        barcode = code128.Code128(str(barcode_value), barHeight=target_barcode_height, barWidth=adjusted_bar_width)
        barcode_x = x + ((cell_width - barcode.width) / 2)
        barcode_y = y + ((cell_height - target_barcode_height) / 2) + barcode_offset_y
        barcode.drawOn(pdf, barcode_x, barcode_y)
        pdf.setFont("Helvetica-Bold", text_font_size)
        if text_orientation == "vertical":
            pdf.saveState()
            text_x = barcode_x - text_gap
            text_y = barcode_y + (target_barcode_height * 0.1)
            pdf.translate(text_x, text_y)
            pdf.rotate(270)
            pdf.drawString(0, 0, str(barcode_value))
            pdf.restoreState()
        else:
            text_y = max(y + (1.5 * mm), barcode_y - text_gap)
            pdf.drawCentredString(x + (cell_width / 2), text_y, str(barcode_value))
    pdf.showPage()
    pdf.save()
    pdf_buffer.seek(0)
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name="caixinhas_codigos.pdf",
        mimetype="application/pdf",
    )


@etq_bp.post("/clientes")
@_etiq_feature_required("etiq_adicionar_cliente")
def add_or_update_client():
    try:
        numero_cliente = int(request.form["numero_cliente"])
        nome_cliente = request.form.get("nome_cliente", "").strip()
        cor_roteiro = request.form.get("cor_roteiro", "").strip()
        horario_roteiro = request.form.get("horario_roteiro", "").strip()
        entregador = request.form.get("entregador", "").strip()
        endereco = request.form.get("endereco", "").strip()
        numero = request.form.get("numero", "").strip()
        complemento = request.form.get("complemento", "").strip()
        bairro = request.form.get("bairro", "").strip()
        cidade = request.form.get("cidade", "").strip()
        estado = request.form.get("estado", "").strip()
        cnpj = request.form.get("cnpj", "").strip()
        cep = request.form.get("cep", "").strip()
        create_roteiro = request.form.get("create_roteiro", "") == "1"

        # address validation: require main address fields
        if not endereco or not bairro or not cidade or not estado or not cep:
            raise ValueError("Endereço completo (rua, bairro, cidade, estado, CEP) e obrigatório.")

        # if user asked to create a roteiro, ensure cor and entregador present
        if create_roteiro:
            if not cor_roteiro:
                raise ValueError("Cor do roteiro é obrigatória quando cadastrar roteiro.")
            if not entregador:
                raise ValueError("Entregador é obrigatório quando cadastrar roteiro.")

        # proceed to upsert client with address fields
        upsert_client(
            numero_cliente,
            cor_roteiro,
            horario_roteiro,
            nome_cliente,
            entregador,
            endereco=endereco,
            numero=numero,
            complemento=complemento,
            bairro=bairro,
            cidade=cidade,
            estado=estado,
            cep=cep,
            cnpj=cnpj,
        )
        # If requested, create a roteiro record (route) so user can manage routes separately
        if create_roteiro:
            # nome for roteiro: prefer entregador or fallback to 'Roteiro {cor}'
            roteiro_nome = entregador or f"Roteiro {cor_roteiro}"
            _create_or_get_roteiro(cor_roteiro, entregador, horario_roteiro, nome=roteiro_nome)
        return redirect(url_for("etiquetas.index", success="Cliente salvo com sucesso."))
    except Exception as exc:
        return redirect(url_for("etiquetas.index", error=str(exc)))


@etq_bp.get("/clientes/roteiro-info")
@_etiq_feature_required("etiq_adicionar_cliente")
def roteiro_info():
    cor = request.args.get("cor", "").strip().lower()
    if not cor:
        return jsonify({"entregador": "", "horario": ""})
    clients = fetch_all_clients()
    fallback: dict[str, Any] | None = None
    for c in reversed(clients):
        if str(c.get("cor_roteiro", "")).strip().lower() != cor:
            continue
        payload = {
            "entregador": c.get("entregador", "") or "",
            "horario": c.get("horario_roteiro", "") or "",
        }
        if payload["entregador"] or payload["horario"]:
            return jsonify(payload)
        if fallback is None:
            fallback = payload
    if fallback is not None:
        return jsonify(fallback)
    return jsonify({"entregador": "", "horario": ""})



@etq_bp.post("/clientes/roteiros")
@_etiq_feature_required("etiq_adicionar_cliente")
def create_roteiro_endpoint():
    try:
        payload = None
        if request.is_json:
            payload = request.get_json()
        else:
            payload = request.form

        cor = (payload.get("cor") or "").strip()
        entregador = (payload.get("entregador") or "").strip()
        horario = (payload.get("horario") or "").strip()
        nome = (payload.get("nome") or "").strip() or None

        if not cor:
            return jsonify({"ok": False, "error": "Cor é obrigatória."}), 400
        if not entregador:
            return jsonify({"ok": False, "error": "Entregador é obrigatório."}), 400

        roteiro_id = _create_or_get_roteiro(cor, entregador, horario, nome=nome)
        if roteiro_id is None:
            return jsonify({"ok": False, "error": "Não foi possível criar o roteiro."}), 500
        return jsonify({"ok": True, "id": roteiro_id})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500



@etq_bp.post("/clientes/cores")
@_etiq_feature_required("etiq_adicionar_cliente")
def create_roteiro_cor_endpoint():
    try:
        payload = None
        if request.is_json:
            payload = request.get_json()
        else:
            payload = request.form
        nome = (payload.get("nome") or "").strip()
        hexv = (payload.get("hex") or "").strip()
        if not nome:
            return jsonify({"ok": False, "error": "Nome da cor é obrigatório."}), 400
        rec = _create_or_get_cor(nome, hexv)
        if not rec:
            return jsonify({"ok": False, "error": "Não foi possível criar a cor."}), 500
        return jsonify({"ok": True, "id": rec.get("id"), "nome": rec.get("nome"), "hex": rec.get("hex")})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@etq_bp.post("/clientes/excluir/<int:numero_cliente>")
@_etiq_feature_required("etiq_criar")
def remove_client(numero_cliente: int):
    try:
        delete_client(numero_cliente)
        return redirect(url_for("etiquetas.index", success="Cliente removido com sucesso."))
    except Exception as exc:
        return redirect(url_for("etiquetas.index", error=str(exc)))


@etq_bp.get("/etiquetas/os")
@_etiq_feature_required("etiq_criar")
def label_by_os_id():
    try:
        os_id = int(request.args.get("os_id", "").strip())
        numero_cliente = int(request.args.get("numero_cliente", "").strip())
        auto_print = request.args.get("autoprint", "").strip().lower() in {"1", "true", "yes"}
        label_size = request.args.get("label_size", "80x20").strip()
        barcode_type = request.args.get("barcode_type", "CODE128").strip().upper()
        if barcode_type not in {"CODE128", "CODE39", "EAN13", "EAN8", "UPC", "QRCODE"}:
            barcode_type = "CODE128"
        show_tipo = request.args.get("show_tipo_lente", "1").strip().lower() in {"1", "true", "yes"}
        label_data = build_label_data_by_os(os_id, numero_cliente, persist_print_date=True, fetch_tipo=show_tipo)
        if not label_data:
            return redirect(url_for("etiquetas.index", error="Cliente nao encontrado."))
        return render_template(
            "etiq/label_print.html",
            label=label_data,
            auto_print=auto_print,
            label_size=label_size,
            barcode_type=barcode_type,
        )
    except Exception as exc:
        return redirect(url_for("etiquetas.index", error=setup_error_message(exc)))


@etq_bp.get("/etiquetas/os/preview")
@_etiq_feature_required("etiq_criar")
def label_preview_by_os_id():
    try:
        os_id = int(request.args.get("os_id", "").strip())
        numero_cliente = int(request.args.get("numero_cliente", "").strip())
        label_size = request.args.get("label_size", "80x20").strip()
        show_tipo = request.args.get("show_tipo_lente", "1").strip().lower() in {"1", "true", "yes"}
        label_data = build_label_data_by_os(os_id, numero_cliente, persist_print_date=False, fetch_tipo=show_tipo)
        if not label_data:
            return "Cliente nao encontrado.", 404
        return render_template(
            "etiq/label_strip_fragment.html",
            label=label_data,
            barcode_id="os-barcode-inline",
            label_size=label_size,
        )
    except Exception as exc:
        return setup_error_message(exc), 400


@etq_bp.get("/etiquetas/cliente/<int:numero_cliente>")
@_etiq_feature_required("etiq_criar")
def print_label(numero_cliente: int):
    label_data = build_label_data(numero_cliente, persist_print_date=True)
    auto_print = request.args.get("autoprint", "").strip().lower() in {"1", "true", "yes"}
    label_size = request.args.get("label_size", "80x20").strip()
    if not label_data:
        return redirect(url_for("etiquetas.index", error="Cliente nao encontrado."))
    return render_template(
        "etiq/label_print.html",
        label=label_data,
        auto_print=auto_print,
        label_size=label_size,
    )


@etq_bp.get("/etiquetas/cliente/<int:numero_cliente>/pdf")
@_etiq_feature_required("etiq_criar")
def label_pdf(numero_cliente: int):
    label_data = build_label_data(numero_cliente, persist_print_date=True)
    if not label_data:
        return redirect(url_for("etiquetas.index", error="Cliente nao encontrado."))
    from reportlab.pdfgen import canvas
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=(LABEL_WIDTH_MM * mm, LABEL_HEIGHT_MM * mm))
    draw_label_pdf(pdf, label_data, LABEL_SAFE_MARGIN_MM * mm, (LABEL_HEIGHT_MM - LABEL_SAFE_MARGIN_MM) * mm)
    pdf.showPage()
    pdf.save()
    # ── save to Impressos/ ───────────────────────────────────────────────────
    _save_to_impressos(pdf_buffer.getvalue(), f"cliente_{numero_cliente}")
    pdf_buffer.seek(0)
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"etiqueta_cliente_{numero_cliente}.pdf",
        mimetype="application/pdf",
    )


@etq_bp.get("/etiquetas/cliente/<int:numero_cliente>/pdf/caixinha")
@_etiq_feature_required("etiq_criar")
def label_pdf_caixinha(numero_cliente: int):
    label_data = build_label_data(numero_cliente, persist_print_date=True)
    if not label_data:
        return redirect(url_for("etiquetas.index", error="Cliente nao encontrado."))
    COLS = 3
    ROWS = 8
    LABEL_W = 60 * mm
    LABEL_H = 30 * mm
    MARGIN_X = 15 * mm
    MARGIN_Y = 15 * mm
    PAGE_W = 210 * mm
    PAGE_H = 297 * mm
    from reportlab.pdfgen import canvas
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=(PAGE_W, PAGE_H))
    for row in range(ROWS):
        for col in range(COLS):
            x = MARGIN_X + (col * LABEL_W)
            y = PAGE_H - MARGIN_Y - (row * LABEL_H)
            draw_label_caixinha_pdf(pdf, label_data, x, y, LABEL_W, LABEL_H)
    pdf.showPage()
    pdf.save()
    pdf_buffer.seek(0)
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"caixinha_cliente_{numero_cliente}.pdf",
        mimetype="application/pdf",
    )


@etq_bp.get("/etiquetas/lote")
@_etiq_feature_required("etiq_criar")
def print_batch_labels():
    try:
        filter_rota = request.args.get("filter_rota", "").strip()
        filter_data = request.args.get("filter_data", "").strip()
        clients = fetch_clients_filtered(filter_rota, filter_data)
        labels = build_labels_from_clients(clients)
        if not labels:
            return redirect(url_for("etiquetas.index", error="Nenhum cliente encontrado para os filtros informados."))
        return render_template("etiq/labels_batch_print.html", labels=labels)
    except Exception as exc:
        return redirect(url_for("etiquetas.index", error=str(exc)))


@etq_bp.get("/etiquetas/lote/pdf")
@_etiq_feature_required("etiq_criar")
def batch_labels_pdf():
    try:
        filter_rota = request.args.get("filter_rota", "").strip()
        filter_data = request.args.get("filter_data", "").strip()
        clients = fetch_clients_filtered(filter_rota, filter_data)
        labels = build_labels_from_clients(clients)
        if not labels:
            return redirect(url_for("etiquetas.index", error="Nenhum cliente encontrado para os filtros informados."))
        from reportlab.pdfgen import canvas
        pdf_buffer = BytesIO()
        pdf = canvas.Canvas(pdf_buffer, pagesize=(LABEL_WIDTH_MM * mm, LABEL_HEIGHT_MM * mm))
        for label in labels:
            draw_label_pdf(pdf, label, LABEL_SAFE_MARGIN_MM * mm, (LABEL_HEIGHT_MM - LABEL_SAFE_MARGIN_MM) * mm)
            pdf.showPage()
        pdf.save()
        pdf_buffer.seek(0)
        return Response(
            pdf_buffer.read(),
            mimetype="application/pdf",
            headers={"Content-Disposition": "attachment; filename=etiquetas_lote.pdf"},
        )
    except Exception as exc:
        return redirect(url_for("etiquetas.index", error=str(exc)))


@etq_bp.get("/etiquetas/manual")
@_etiq_feature_required("etiq_criar")
def label_manual_edit():
    cores = sorted(ROUTE_COLOR_MAP.items())
    return render_template("etiq/label_edit.html", cores=cores, layout_config=load_label_layout_config())


@etq_bp.get("/etiquetas/modelos")
@_etiq_feature_required("etiq_criar")
def list_label_models():
    return jsonify(load_label_models())


@etq_bp.post("/etiquetas/modelos/salvar")
@_etiq_feature_required("etiq_criar")
def save_label_model():
    data = request.get_json(force=True, silent=True) or {}
    name = str(data.get("name", "")).strip()
    config = data.get("config", {})
    if not name:
        return jsonify({"error": "Nome obrigatorio."}), 400
    if not isinstance(config, dict):
        return jsonify({"error": "Configuracao invalida."}), 400
    with LABEL_MODELS_LOCK:
        models = load_label_models()
        models = [m for m in models if m.get("name") != name]
        models.append({
            "name": name,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "config": config,
        })
        save_label_models(models)
    return jsonify({"ok": True})


@etq_bp.post("/etiquetas/modelos/renomear")
@_etiq_feature_required("etiq_criar")
def rename_label_model():
    data = request.get_json(force=True, silent=True) or {}
    old_name = str(data.get("old_name", "")).strip()
    new_name = str(data.get("new_name", "")).strip()
    if not old_name or not new_name:
        return jsonify({"error": "Nome obrigatorio."}), 400
    if old_name == new_name:
        return jsonify({"ok": True})
    with LABEL_MODELS_LOCK:
        models = load_label_models()
        found = False
        for m in models:
            if m.get("name") == old_name:
                m["name"] = new_name
                found = True
                break
        if not found:
            return jsonify({"error": "Modelo nao encontrado."}), 404
        save_label_models(models)
    return jsonify({"ok": True})


@etq_bp.post("/etiquetas/modelos/excluir")
@_etiq_feature_required("etiq_criar")
def delete_label_model():
    data = request.get_json(force=True, silent=True) or {}
    name = str(data.get("name", "")).strip()
    if not name:
        return jsonify({"error": "Nome obrigatorio."}), 400
    with LABEL_MODELS_LOCK:
        models = load_label_models()
        save_label_models([m for m in models if m.get("name") != name])
    return jsonify({"ok": True})


@etq_bp.get("/etiquetas/layout")
@_etiq_feature_required("etiq_criar")
def get_label_layout():
    return jsonify(load_label_layout_config())


@etq_bp.post("/etiquetas/layout/salvar")
@_etiq_feature_required("etiq_criar")
def save_label_layout():
    data = request.get_json(force=True, silent=True) or {}
    config = data.get("config", {})
    if not isinstance(config, dict):
        return jsonify({"error": "Configuracao invalida."}), 400
    with LABEL_LAYOUT_LOCK:
        save_label_layout_config(config)
    return jsonify({"ok": True})


@etq_bp.post("/etiquetas/layout/limpar")
@_etiq_feature_required("etiq_criar")
def clear_label_layout():
    with LABEL_LAYOUT_LOCK:
        clear_label_layout_config()
    return jsonify({"ok": True})
