"""
Blueprint: Confirmações de Ordens de Serviço (Conferência)
Módulo integrado ao WMS para validação e rastreamento de OS
"""

from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, send_file, flash
from functools import wraps
import json
from datetime import datetime
from io import BytesIO, StringIO
import csv

import db_mdb

confirmations_bp = Blueprint('confirmations', __name__, url_prefix='')


def login_required(f):
    """Decorator para proteger rotas que precisam de autenticação"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def get_current_user():
    """Retorna informações do usuário autenticado"""
    return session.get('user')


def get_current_unit():
    """Retorna unidade da sessão"""
    return session.get('unit', db_mdb.DEFAULT_UNIT)


def get_current_sector():
    """Retorna setor ativo da sessão"""
    return session.get('sector', db_mdb.DEFAULT_SECTOR)


def is_admin():
    """Verifica se usuário é admin"""
    return get_current_user() == 'admin'


def can_access_feature(feature):
    if is_admin():
        return True
    permissions = session.get('permissions', [])
    return feature in permissions if isinstance(permissions, list) else False


# ============================================================================
# ROTAS PRINCIPAIS
# ============================================================================

@confirmations_bp.route('/confirmations', methods=['GET'])
@login_required
def confirmations_page():
    """Página principal de conferência de OS"""
    if not can_access_feature('confirmations'):
        flash('Acesso restrito à conferência de OS.', 'danger')
        return redirect(url_for('dashboard'))
    user = get_current_user()
    unit = get_current_unit()
    sector = get_current_sector()
    
    # Busca informações do usuário do banco
    user_info = db_mdb.get_user_by_username(user, unit=unit)
    
    return render_template('confirmations.html',
        username=user,
        sector=sector or user_info.get('sector', 'Geral') if user_info else 'Geral',
        unit=unit
    )


@confirmations_bp.route('/confirmations/history', methods=['GET'])
@login_required
def confirmations_history_page():
    """Lista histórica das conferências do usuário logado"""
    if not (is_admin() or can_access_feature('confirmations_history')):
        flash('Acesso restrito ao histórico da conferência de OS.', 'danger')
        return redirect(url_for('dashboard'))

    user = get_current_user()
    unit = get_current_unit()
    sector = get_current_sector()

    confirmations = db_mdb.get_confirmations(
        unit=unit,
        username=user,
        limit=200
    )

    return render_template(
        'confirmations_history.html',
        username=user,
        sector=sector,
        unit=unit,
        confirmations=confirmations,
    )


@confirmations_bp.route('/admin/confirmations', methods=['GET'])
@login_required
def admin_confirmations_page():
    """Painel administrativo de confirmações"""
    if not is_admin():
        return "Acesso negado", 403
    
    unit = get_current_unit()
    
    # Busca todos os usuários para filtro
    all_users = db_mdb.get_all_users(unit=unit)
    usernames = [u.get('username') for u in all_users if u.get('username')]
    
    # Busca setores disponíveis
    try:
        with open('sectors.json', 'r', encoding='utf-8') as f:
            sectors_data = json.load(f)
            sectors = [s.get('name') for s in sectors_data.get('sectors', [])]
    except:
        sectors = ['AR', 'TRIAGEM']
    
    return render_template('admin_confirmations.html',
        usernames=usernames,
        sectors=sectors,
        unit=unit
    )


# ============================================================================
# API: CONFIRMAÇÕES
# ============================================================================

@confirmations_bp.route('/api/confirmations', methods=['POST'])
@login_required
def create_confirmation():
    """Cria novo registro de confirmação"""
    try:
        data = request.get_json()
        username = get_current_user()
        unit = get_current_unit()
        sector = get_current_sector()
        
        os_reference = str(data.get('os_reference', '')).strip()
        os_confirmation = str(data.get('os_confirmation', '')).strip()
        
        if not os_reference or not os_confirmation:
            return jsonify({'error': 'OS inválidos'}), 400
        
        # Determina resultado (ok ou erro)
        result = 'ok' if os_reference == os_confirmation else 'error'
        
        # Salva no banco
        confirmation = db_mdb.add_confirmation(
            username=username,
            sector=sector,
            os_reference=os_reference,
            os_confirmation=os_confirmation,
            result=result,
            unit=unit
        )
        
        return jsonify({
            'success': True,
            'result': result,
            'confirmation': confirmation
        }), 201
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@confirmations_bp.route('/api/confirmations', methods=['GET'])
@login_required
def get_user_confirmations():
    """Lista confirmações do usuário logado"""
    try:
        username = get_current_user()
        unit = get_current_unit()
        limit = request.args.get('limit', 100, type=int)
        
        confirmations = db_mdb.get_confirmations(
            unit=unit,
            username=username,
            limit=limit
        )
        
        return jsonify({
            'success': True,
            'confirmations': confirmations
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@confirmations_bp.route('/api/confirmations/stats', methods=['GET'])
@login_required
def get_confirmations_stats():
    """Retorna estatísticas de confirmações"""
    try:
        unit = get_current_unit()
        
        # Se for admin, retorna stats globais; senão apenas do usuário
        if is_admin():
            filters = None
        else:
            filters = {'username': get_current_user()}
        
        stats = db_mdb.get_confirmation_stats(unit=unit, filters=filters)
        
        return jsonify({
            'success': True,
            'stats': stats
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@confirmations_bp.route('/api/confirmations/all', methods=['GET'])
@login_required
def get_all_confirmations():
    """Lista todas as confirmações (admin)"""
    try:
        if not is_admin():
            return jsonify({'error': 'Unauthorized'}), 403
        
        unit = get_current_unit()
        
        # Busca com filtros opcionais
        filters = {
            'result': request.args.get('result', ''),
            'username': request.args.get('username', ''),
            'date_from': request.args.get('date_from', ''),
            'date_to': request.args.get('date_to', ''),
            'sector': request.args.get('sector', ''),
        }
        
        # Remove filtros vazios
        filters = {k: v for k, v in filters.items() if v}
        
        confirmations = db_mdb.get_confirmations_filtered(filters, unit=unit)
        stats = db_mdb.get_confirmation_stats(unit=unit, filters=filters)
        
        return jsonify({
            'success': True,
            'confirmations': confirmations,
            'stats': stats
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@confirmations_bp.route('/api/confirmations/export', methods=['POST'])
@login_required
def export_confirmations_xlsx():
    """Exporta confirmações em XLSX - APENAS ADMIN"""
    try:
        from flask import send_file
        
        # Apenas admin pode exportar
        if not is_admin():
            return jsonify({'error': 'Permissão negada. Apenas administrador pode exportar.'}), 403
        
        # Admin exporta com filtros
        unit = get_current_unit()
        data = request.get_json() or {}
        filters = {
            'result': data.get('result', ''),
            'username': data.get('username', ''),
            'date_from': data.get('date_from', ''),
            'date_to': data.get('date_to', ''),
            'sector': data.get('sector', ''),
        }
        filters = {k: v for k, v in filters.items() if v}
        confirmations = db_mdb.get_confirmations_filtered(filters, unit=unit)
        
        if not confirmations:
            return jsonify({'error': 'Nenhum registro para exportar'}), 400
        
        # Tenta usar openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Cria workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Confirmações OS"
            
            # Headers
            headers = ['#', 'Usuário', 'Setor', 'OS Referência', 'OS Confirmação', 'Data', 'Hora', 'Status']
            ws.append(headers)
            
            # Estilo header
            header_fill = PatternFill(start_color="1E2130", end_color="1E2130", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Dados
            for i, conf in enumerate(confirmations, 1):
                resultado = 'OK' if conf.get('result') == 'ok' else 'Divergente'
                ws.append([
                    i,
                    conf.get('username', ''),
                    conf.get('sector', ''),
                    conf.get('os_reference', ''),
                    conf.get('os_confirmation', ''),
                    conf.get('data', ''),
                    conf.get('hora', ''),
                    resultado,
                ])
            
            # Ajusta larguras de coluna
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 16
            
            # Sheet de resumo
            ws_stats = wb.create_sheet("Resumo")
            stats = db_mdb.get_confirmation_stats(unit=unit)
            
            ws_stats['A1'] = "Total de conferências:"
            ws_stats['B1'] = stats.get('total', 0)
            
            ws_stats['A2'] = "Conferidas com sucesso:"
            ws_stats['B2'] = stats.get('ok', 0)
            
            ws_stats['A3'] = "Divergentes:"
            ws_stats['B3'] = stats.get('error', 0)
            
            ws_stats['A4'] = "Taxa de acerto (%):"
            ws_stats['B4'] = stats.get('accuracy_percent', 0)
            
            ws_stats['A5'] = "Gerado em:"
            ws_stats['B5'] = db_mdb.datetime_now_str()
            
            ws_stats['A6'] = "Exportado por:"
            ws_stats['B6'] = get_current_user()
            
            # Salva em memória
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f'confirmacoes_os_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        except ImportError:
            # Fallback: xlsxwriter se disponível
            try:
                import xlsxwriter
                output = BytesIO()
                
                workbook = xlsxwriter.Workbook(output)
                worksheet = workbook.add_worksheet("Confirmações OS")
                header_format = workbook.add_format({
                    'bg_color': '#1E2130',
                    'font_color': 'white',
                    'bold': True,
                    'border': 1,
                })
                
                # Headers
                headers = ['#', 'Usuário', 'Setor', 'OS Referência', 'OS Confirmação', 'Data', 'Hora', 'Status']
                for col, header in enumerate(headers):
                    worksheet.write(0, col, header, header_format)
                
                # Dados
                for row, conf in enumerate(confirmations, 1):
                    resultado = 'OK' if conf.get('result') == 'ok' else 'Divergente'
                    worksheet.write(row, 0, row)
                    worksheet.write(row, 1, conf.get('username', ''))
                    worksheet.write(row, 2, conf.get('sector', ''))
                    worksheet.write(row, 3, conf.get('os_reference', ''))
                    worksheet.write(row, 4, conf.get('os_confirmation', ''))
                    worksheet.write(row, 5, conf.get('data', ''))
                    worksheet.write(row, 6, conf.get('hora', ''))
                    worksheet.write(row, 7, resultado)
                
                workbook.close()
                output.seek(0)
                
                filename = f'confirmacoes_os_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
            
            except ImportError:
                # Fallback final: CSV
                output = BytesIO()
                import csv
                
                # Escreve como bytes
                text_buffer = StringIO()
                writer = csv.writer(text_buffer)
                
                writer.writerow(['#', 'Usuário', 'Setor', 'OS Referência', 'OS Confirmação', 'Data', 'Hora', 'Status'])
                
                for i, conf in enumerate(confirmations, 1):
                    resultado = 'OK' if conf.get('result') == 'ok' else 'Divergente'
                    writer.writerow([
                        i,
                        conf.get('username', ''),
                        conf.get('sector', ''),
                        conf.get('os_reference', ''),
                        conf.get('os_confirmation', ''),
                        conf.get('data', ''),
                        conf.get('hora', ''),
                        resultado,
                    ])
                
                output.write(text_buffer.getvalue().encode('utf-8'))
                output.seek(0)
                
                filename = f'confirmacoes_os_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
                
                return send_file(
                    output,
                    mimetype='text/csv',
                    as_attachment=True,
                    download_name=filename
                )
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@confirmations_bp.route('/api/confirmations/search', methods=['GET'])
@login_required
def search_confirmations():
    """Busca confirmações por termo"""
    try:
        query = request.args.get('q', '')
        unit = get_current_unit()
        
        if not query:
            return jsonify({'error': 'Query vazio'}), 400
        
        if is_admin():
            confirmations = db_mdb.search_confirmations(query, unit=unit)
        else:
            # Operador só busca suas próprias confirmações
            confirmations = db_mdb.search_confirmations(query, unit=unit)
            confirmations = [c for c in confirmations if c.get('username') == get_current_user()]
        
        return jsonify({
            'success': True,
            'confirmations': confirmations
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
