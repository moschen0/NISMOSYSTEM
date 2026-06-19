"""
WMS (Warehouse Management System) - Web Application (versão MDB)
Sistema de Gerenciamento de Armazém com Flask + Access Database
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, Response
from functools import wraps
from datetime import datetime, timedelta
import json
import logging
import os
import re
import shutil
import sys
import threading
import time
import unicodedata
import uuid
from logging.handlers import RotatingFileHandler

# Carrega variáveis do arquivo .env ANTES de importar db_mdb,
# para que WMS_MDB_PATH já esteja disponível quando DB_PATH for resolvido.
# Quando rodando como EXE (frozen), __file__ aponta para _internal; usa sys.executable como base.
try:
    from dotenv import load_dotenv
    _env_base = (
        os.path.dirname(sys.executable)
        if getattr(sys, 'frozen', False)
        else os.path.dirname(os.path.abspath(__file__))
    )
    load_dotenv(os.path.join(_env_base, '.env'), override=False)
except ImportError:
    pass  # python-dotenv opcional; use variáveis de ambiente do sistema

import db_mdb

# ============================================================================
# CONFIGURAÇÃO INICIAL
# ============================================================================

def get_resource_base_dir():
    """Retorna a pasta de recursos (templates/static) para dev e executavel."""
    if not getattr(sys, 'frozen', False):
        return os.path.dirname(os.path.abspath(__file__))

    exe_dir = os.path.dirname(sys.executable)
    meipass_dir = getattr(sys, '_MEIPASS', None)
    candidates = [
        meipass_dir,
        os.path.join(exe_dir, '_internal'),
        exe_dir,
    ]

    for base_dir in candidates:
        if not base_dir:
            continue
        templates_dir = os.path.join(base_dir, 'templates')
        static_dir = os.path.join(base_dir, 'static')
        if os.path.isdir(templates_dir) and os.path.isdir(static_dir):
            return base_dir

    return exe_dir


def get_runtime_data_dir():
    """Retorna pasta para dados mutaveis em execucao."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


RESOURCE_BASE_DIR = get_resource_base_dir()
DATA_BASE_DIR = get_runtime_data_dir()
TEMPLATES_DIR = os.path.join(RESOURCE_BASE_DIR, 'templates')
ZONE_METADATA_PATH = os.path.join(DATA_BASE_DIR, 'zone_metadata.json')
TAG_CATALOG_PATH = os.path.join(DATA_BASE_DIR, 'zone_tag_catalog.json')
ZONE_TAGS_PATH = os.path.join(DATA_BASE_DIR, 'zone_tags_map.json')
SECTORS_PATH = os.path.join(DATA_BASE_DIR, 'sectors.json')
TIME_THRESHOLDS_PATH    = os.path.join(DATA_BASE_DIR, 'time_thresholds.json')
TELEGRAM_CONFIG_PATH   = os.path.join(DATA_BASE_DIR, 'telegram_config.json')
TELEGRAM_NOTIFIED_PATH = os.path.join(DATA_BASE_DIR, 'telegram_notified.json')
OPTO_SCHEDULER_PATH = os.path.join(DATA_BASE_DIR, 'opto_scheduler.json')
DB_MODE_FILE           = os.path.join(DATA_BASE_DIR, 'db_mode.json')
IP_ACL_PATH            = os.path.join(DATA_BASE_DIR, 'ip_acl.json')
AUDIT_HISTORY_PATH     = os.path.join(DATA_BASE_DIR, 'conference_history.json')

AUDIT_HISTORY_LOCK = threading.Lock()
AUDIT_HISTORY_LIMIT = 5

TAG_RULES = {
    'maintenance': 'Em manutencao (ignora na alocacao)',
    'priority': 'Prioridade (primeira da fila)',
    'none': 'Sem regra automatica'
}
TRIAGE_SECTOR = 'TRIAGEM'
AR_SECTOR = 'AR'
OPTO_OS_PREFIXES = ('9MA', '2BA', '6VA')

# PERMISSION_FLAGS será carregado dinamicamente de load_permissions()
# Inicializar com padrão vazio para evitar erro antes do Flask estar pronto
PERMISSION_FLAGS = {}

# ============================================================================
# LOGGER WMS
# ============================================================================

# WMS.log fica na pasta-pai (raiz do projeto), um nível acima de WMS_SISTEMA
_WMS_LOG_PATH = os.path.normpath(os.path.join(DATA_BASE_DIR, '..', 'WMS.log'))


def _setup_wms_logger() -> logging.Logger:
    """Configura e retorna o logger principal do WMS.

    Formato: [DD/MM/YYYY HH:MM:SS] LEVEL   | mensagem
    Arquivo rotativo: 2 MB por arquivo, 5 backups (WMS.log, WMS.log.1 … .5).
    """
    logger = logging.getLogger('wms')
    if logger.handlers:          # evita duplicação em reloads
        return logger
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        '[%(asctime)s] %(levelname)-7s | %(message)s',
        datefmt='%d/%m/%Y %H:%M:%S'
    )
    try:
        fh = RotatingFileHandler(
            _WMS_LOG_PATH,
            maxBytes=2 * 1024 * 1024,   # 2 MB
            backupCount=5,
            encoding='utf-8'
        )
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(fmt)
        logger.addHandler(fh)
    except Exception as _log_err:
        print(f'[LOGGER] Não foi possível criar WMS.log: {_log_err}')

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    return logger


wms_logger = _setup_wms_logger()

# ============================================================================
# MODEL — TEMPO / COR
# ============================================================================

# Defaults para thresholds de tempo de permanência (dias)
_DEFAULT_THRESHOLDS = {'green_days': 3, 'yellow_days': 4, 'red_days': 6}


def load_time_thresholds():
    """Carrega thresholds de colorização por tempo de permanência."""
    if os.path.exists(TIME_THRESHOLDS_PATH):
        try:
            with open(TIME_THRESHOLDS_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # garante que todas as chaves existam
            return {k: int(data.get(k, v)) for k, v in _DEFAULT_THRESHOLDS.items()}
        except Exception:
            pass
    return dict(_DEFAULT_THRESHOLDS)


def save_time_thresholds(green_days, yellow_days, red_days):
    """Persiste os thresholds de tempo no arquivo JSON."""
    with open(TIME_THRESHOLDS_PATH, 'w', encoding='utf-8') as f:
        json.dump(
            {'green_days': int(green_days), 'yellow_days': int(yellow_days), 'red_days': int(red_days)},
            f, indent=2
        )


# ── Telegram config ──────────────────────────────────────────────────────────

_DEFAULT_TELEGRAM_CONFIG = {
    'notify_status_alerts': False,
    'notify_tiers': ['urgent', 'critical'],
    'notify_daily_report': False,
    'daily_report_hour': 8,
    # Relatório de período (fecha mês)
    'scheduled_report_enabled': False,
    'scheduled_report_hour': 8,
    'scheduled_report_mode': 'month_to_date',   # 'month_to_date' | 'full_month' | 'custom_days'
    'scheduled_report_start_day': 1,
    'scheduled_report_end_day': 0,              # 0 = último dia do mês
}


def load_telegram_config():
    """Carrega configurações de notificação do Telegram."""
    if os.path.exists(TELEGRAM_CONFIG_PATH):
        try:
            with open(TELEGRAM_CONFIG_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return {k: data.get(k, v) for k, v in _DEFAULT_TELEGRAM_CONFIG.items()}
        except Exception:
            pass
    return dict(_DEFAULT_TELEGRAM_CONFIG)


def save_telegram_config(cfg):
    """Persiste as configurações de notificação do Telegram (escrita atômica)."""
    safe = {k: cfg.get(k, v) for k, v in _DEFAULT_TELEGRAM_CONFIG.items()}
    tmp = TELEGRAM_CONFIG_PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(safe, f, indent=2, ensure_ascii=False)
    os.replace(tmp, TELEGRAM_CONFIG_PATH)


# ============================================================================
# MODO DO BANCO DE DADOS (produção / teste)
# ============================================================================

def load_db_mode():
    """Retorna o modo salvo: 'production' ou 'test'. Padrão: 'production'."""
    try:
        if os.path.exists(DB_MODE_FILE):
            with open(DB_MODE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if data.get('mode') in ('production', 'test'):
                return data['mode']
    except Exception:
        pass
    return 'production'


def save_db_mode(mode: str):
    """Persiste o modo do banco (escrita atômica)."""
    tmp = DB_MODE_FILE + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump({'mode': mode}, f)
    os.replace(tmp, DB_MODE_FILE)


def apply_db_mode(mode: str):
    """Aplica o modo de banco chamando db_mdb.switch_database com o caminho correto."""
    path = db_mdb.DB_PATH_PROD if mode == 'production' else db_mdb.DB_PATH_TEST
    db_mdb.switch_database(path)
    wms_logger.info(f'DB | Modo de banco alterado para: {mode} ({path})')


def get_order_age_days(timestamp_str):
    """Retorna a idade do pedido em dias inteiros a partir do campo timestamp."""
    if not timestamp_str:
        return None
    for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            dt = datetime.strptime(str(timestamp_str).strip(), fmt)
            return max(0, (datetime.now() - dt).days)
        except ValueError:
            continue
    return None

def get_age_tier(age_days, thresholds):
    """Retorna o tier de cor: 'normal'|'attention'|'urgent'|'critical'."""
    if age_days is None:
        return 'normal'
    gd = thresholds.get('green_days', 3)
    yd = thresholds.get('yellow_days', 4)
    rd = thresholds.get('red_days', 6)
    if age_days < gd:
        return 'normal'
    if age_days < yd:
        return 'attention'
    if age_days < rd:
        return 'urgent'
    return 'critical'


def make_box_entry(order, thresholds):
    """Constrói os dados da caixa para visualização e tooltip de detalhes."""
    label = str(order.get('box') or order.get('order_id', '')).strip()
    order_id = str(order.get('order_id', '') or '').strip()
    box_number = str(order.get('box', '') or '').strip()
    os_opto = str(order.get('os_opto', '') or '').strip()
    triage_caixa = os_opto  # para triagem: os_opto guarda o número da caixa física
    created_by = str(order.get('created_by', '') or '').strip()
    created_at = str(order.get('timestamp') or order.get('date') or '').strip()
    age_days = get_order_age_days(order.get('timestamp'))
    tier = get_age_tier(age_days, thresholds)
    return {
        'label': label,
        'tier': tier,
        'age_days': age_days,
        'order_id': order_id,
        'box': box_number,
        'os_opto': os_opto,
        'triage_caixa': triage_caixa,
        'created_by': created_by,
        'created_at': created_at,
    }

# ============================================================================
# MODEL — SETORES
# ============================================================================

def _normalize_sector_permissions(sector):
    perms = sector.get('permissions', [])
    if isinstance(perms, str):
        perms = [p.strip() for p in perms.split(',') if p.strip()]
    if not isinstance(perms, list):
        return []
    return [p for p in perms if p in PERMISSION_FLAGS]


def _make_sector_entry(name, description, permissions=None):
    return {
        'name': name,
        'description': description,
        'status': 'active',
        'created_at': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        'permissions': permissions or []
    }


def load_sectors():
    """Carrega definições de células/setores do arquivo JSON."""
    default = {
        'AR': _make_sector_entry('AR', 'Setor AR', permissions=[]),
        TRIAGE_SECTOR: _make_sector_entry(TRIAGE_SECTOR, 'Setor de recebimento e triagem', permissions=['triage', 'etiquetas']),
        'VTA': _make_sector_entry('VTA', 'Setor VTA', permissions=[])
    }

    if not os.path.exists(SECTORS_PATH):
        save_sectors(default)
        return default
    try:
        with open(SECTORS_PATH, 'r', encoding='utf-8') as f:
            sectors = json.load(f)
    except Exception as e:
        print(f"Erro ao ler sectors.json: {e}")
        return default

    changed = False
    legacy_default_permissions = list(PERMISSION_FLAGS.keys())
    for key, sector in list(sectors.items()):
        if key in {'AR', 'VTA'} and not sector.get('permissions'):
            sector['permissions'] = list(legacy_default_permissions)
            changed = True
        if key == TRIAGE_SECTOR and 'permissions' not in sector:
            sector['permissions'] = ['triage', 'etiquetas']
            changed = True
        normalized = _normalize_sector_permissions(sector)
        if sector.get('permissions') != normalized:
            sector['permissions'] = normalized
            changed = True
        if 'status' not in sector:
            sector['status'] = 'active'
            changed = True
        if 'created_at' not in sector:
            sector['created_at'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            changed = True
        sectors[key] = sector

    if TRIAGE_SECTOR not in sectors:
        sectors[TRIAGE_SECTOR] = _make_sector_entry(TRIAGE_SECTOR, 'Setor de recebimento e triagem', permissions=['triage', 'etiquetas'])
        changed = True

    if changed:
        save_sectors(sectors)
    return sectors

def save_sectors(sectors):
    """Salva setores de forma atômica."""
    tmp_path = f"{SECTORS_PATH}.tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(sectors, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, SECTORS_PATH)

def get_active_sector_keys():
    """Retorna lista de chaves de setores ativos."""
    sectors = load_sectors()
    return [k for k, v in sectors.items() if v.get('status') == 'active']

# ============================================================================
# MODEL — PERMISSÕES
# ============================================================================

PERMISSIONS_PATH = os.path.join(DATA_BASE_DIR, 'permissions.json')

def load_permissions():
    """Carrega permissões do arquivo JSON."""
    try:
        if os.path.exists(PERMISSIONS_PATH):
            with open(PERMISSIONS_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return {p['id']: p for p in data.get('permissions', [])}
    except Exception as e:
        wms_logger.error(f"Erro ao carregar permissões: {e}")
    
    # Padrão: áreas principais do sistema
    return {
        'dashboard': {
            'id': 'dashboard',
            'name': 'Dashboard / Prateleiras',
            'description': 'Visualização do painel principal e das prateleiras',
            'icon': 'bi-speedometer2'
        },
        'audit': {
            'id': 'audit',
            'name': 'Conferência de Endereço',
            'description': 'Conferência de posições e endereços de estoque',
            'icon': 'bi-clipboard2-check'
        },
        'audit_expected_orders': {
            'id': 'audit_expected_orders',
            'name': 'Conferência de Endereço - Detalhes',
            'description': 'Visualização dos pedidos esperados na conferência de endereço',
            'icon': 'bi-list-check'
        },
        'confirmations': {
            'id': 'confirmations',
            'name': 'Conferência de OS',
            'description': 'Acesso à tela de conferência de ordens de serviço',
            'icon': 'bi-check2-circle'
        },
        'confirmations_history': {
            'id': 'confirmations_history',
            'name': 'Histórico da Conferência de OS',
            'description': 'Consulta à lista das próprias conferências realizadas',
            'icon': 'bi-list-check'
        },
        'search': {
            'id': 'search',
            'name': 'Busca',
            'description': 'Pesquisa de pedidos, posições e caixas',
            'icon': 'bi-search'
        },
        'search_cross_sector': {
            'id': 'search_cross_sector',
            'name': 'Busca entre setores',
            'description': 'Permite pesquisar pedidos e posições em todos os setores da unidade',
            'icon': 'bi-diagram-3'
        },
        'checkout': {
            'id': 'checkout',
            'name': 'Saída de Pedidos',
            'description': 'Registro de retirada de pedidos do estoque',
            'icon': 'bi-box-arrow-up'
        },
        'os_opto': {
            'id': 'os_opto',
            'name': 'OS OPTO',
            'description': 'Exige o campo OS OPTO no cadastro de pedidos do setor',
            'icon': 'bi-building'
        },
        'triage': {
            'id': 'triage',
            'name': 'Triagem',
            'description': 'Acesso à triagem de recebimento',
            'icon': 'bi-box2'
        },
        'etiquetas': {
            'id': 'etiquetas',
            'name': 'Etiquetas',
            'description': 'Geração e impressão de etiquetas',
            'icon': 'bi-tag'
        },
        'movements': {
            'id': 'movements',
            'name': 'Histórico',
            'description': 'Consulta ao histórico de movimentos',
            'icon': 'bi-clock-history'
        },
        'users': {
            'id': 'users',
            'name': 'Usuários',
            'description': 'Gerenciamento de usuários do sistema',
            'icon': 'bi-people'
        },
        'settings': {
            'id': 'settings',
            'name': 'Configurações',
            'description': 'Acesso às configurações do sistema',
            'icon': 'bi-gear'
        }
    }

def save_permissions(permissions_dict):
    """Salva permissões para arquivo JSON."""
    try:
        permissions_list = list(permissions_dict.values())
        tmp_path = f"{PERMISSIONS_PATH}.tmp"
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump({'permissions': permissions_list}, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, PERMISSIONS_PATH)
        return True
    except Exception as e:
        wms_logger.error(f"Erro ao salvar permissões: {e}")
        return False


PERMISSION_FLAGS = load_permissions()

# ============================================================================
# MODEL — BACKUP
# ============================================================================

BACKUP_DIR = r'\\192.168.1.210\apps master\BAKCUP BANDO WMS'
BACKUP_LOG = os.path.join(BACKUP_DIR, 'backup.log')

_backup_last_date = None


def _write_backup_log(message):
    """Escreve uma linha no log de backup, criando o arquivo se necessário."""
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        line = f'[{timestamp}] {message}\n'
        with open(BACKUP_LOG, 'a', encoding='utf-8') as f:
            f.write(line)
    except Exception as e:
        print(f'[BACKUP] Erro ao escrever log: {e}')


def perform_backup(triggered_by='sistema'):
    """Copia wms_database.mdb para o diretório de backup de rede com timestamp.
    Retorna (sucesso: bool, mensagem: str).
    """
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        src = db_mdb.get_db_path()
        if not os.path.isfile(src):
            msg = f'ERRO - Arquivo de origem não encontrado: {src}'
            _write_backup_log(msg)
            return False, 'Arquivo de banco de dados não encontrado.'
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename = f'wms_database_{ts}.mdb'
        dst = os.path.join(BACKUP_DIR, filename)
        shutil.copy2(src, dst)
        size_kb = os.path.getsize(dst) // 1024
        msg = f'SUCESSO - {filename} ({size_kb} KB) por {triggered_by}'
        _write_backup_log(msg)
        wms_logger.info(f'BACKUP OK | {filename} ({size_kb} KB) por {triggered_by}')
        return True, f'Backup realizado: {filename} ({size_kb} KB)'
    except Exception as e:
        msg = f'ERRO - {e} (por {triggered_by})'
        _write_backup_log(msg)
        wms_logger.error(f'BACKUP ERRO | {e} | por {triggered_by}')
        return False, str(e)


def get_backup_log_tail(n=15):
    """Retorna as últimas N linhas do log de backup."""
    try:
        if not os.path.isfile(BACKUP_LOG):
            return []
        with open(BACKUP_LOG, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        return [line.rstrip() for line in lines[-n:]]
    except Exception:
        return []


def _daily_backup_worker():
    """Thread de backup automático diário às 02:00."""
    global _backup_last_date
    backup_hour = 2
    while True:
        try:
            now = datetime.now()
            today = now.date()
            if now.hour >= backup_hour and _backup_last_date != today:
                _backup_last_date = today
                print('[BACKUP] Iniciando backup automático diário...')
                ok, msg = perform_backup(triggered_by='auto-diário')
                print(f'[BACKUP] {msg}')
        except Exception as e:
            print(f'[BACKUP] Erro no worker: {e}')
        time.sleep(600)  # verifica a cada 10 minutos


def start_daily_backup_scheduler():
    """Inicia o thread de backup automático diário em background."""
    t = threading.Thread(target=_daily_backup_worker, daemon=True, name='backup-scheduler')
    t.start()
    wms_logger.info('BACKUP SCHEDULER | Agendador de backup diário iniciado (02:00)')
    print('[BACKUP] Agendador de backup diário iniciado (02:00).')


# ============================================================================
# TELEGRAM — SCHEDULERS
# ============================================================================

# Rastreia quais pedidos já foram notificados em cada tier (evita re-envio).
# Formato: { order_id: 'attention'|'urgent'|'critical' }
_telegram_notified_state: dict = {}


def _load_telegram_notified():
    global _telegram_notified_state
    if os.path.exists(TELEGRAM_NOTIFIED_PATH):
        try:
            with open(TELEGRAM_NOTIFIED_PATH, 'r', encoding='utf-8') as f:
                _telegram_notified_state = json.load(f)
        except Exception:
            _telegram_notified_state = {}
    return _telegram_notified_state


def _save_telegram_notified(data):
    global _telegram_notified_state
    _telegram_notified_state = data
    tmp = TELEGRAM_NOTIFIED_PATH + '.tmp'
    try:
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        os.replace(tmp, TELEGRAM_NOTIFIED_PATH)
    except Exception as exc:
        wms_logger.error(f'TELEGRAM | Erro ao salvar estado de notificados: {exc}')


def _telegram_status_alert_worker():
    """Thread que verifica pedidos com tier alto e envia alertas ao Telegram."""
    import telegram_notifier as tg
    _load_telegram_notified()
    _TIER_RANK = {'normal': 0, 'attention': 1, 'urgent': 2, 'critical': 3}

    while True:
        try:
            cfg = load_telegram_config()
            if cfg.get('notify_status_alerts') and tg.is_configured():
                notify_tiers = cfg.get('notify_tiers', ['urgent', 'critical'])
                thresholds = load_time_thresholds()
                gd = int(thresholds.get('green_days', 3))
                yd = int(thresholds.get('yellow_days', 4))
                rd = int(thresholds.get('red_days', 6))

                try:
                    all_orders = db_mdb.get_all_orders(status_filter='add')
                except TypeError:
                    all_orders = db_mdb.get_all_orders()

                notified = dict(_telegram_notified_state)
                orders_by_tier = {'attention': [], 'urgent': [], 'critical': []}

                for o in all_orders:
                    age = get_order_age_days(o.get('timestamp'))
                    if age is None:
                        continue
                    if age < gd:
                        tier = 'normal'
                    elif age < yd:
                        tier = 'attention'
                    elif age < rd:
                        tier = 'urgent'
                    else:
                        tier = 'critical'

                    if tier not in notify_tiers:
                        continue

                    oid = str(o.get('order_id', ''))
                    last_tier = notified.get(oid, 'normal')
                    if _TIER_RANK.get(tier, 0) > _TIER_RANK.get(last_tier, 0):
                        o['_age_days'] = age
                        orders_by_tier[tier].append(o)
                        notified[oid] = tier

                if any(orders_by_tier.values()):
                    msg = tg.build_status_alert_message(orders_by_tier)
                    if msg:
                        ok, err = tg.send_message(msg)
                        if ok:
                            _save_telegram_notified(notified)
                            wms_logger.info('TELEGRAM | Alerta de status enviado')
                        else:
                            wms_logger.warning(f'TELEGRAM | Falha no alerta de status: {err}')
        except Exception as exc:
            wms_logger.error(f'TELEGRAM | Erro no worker de alertas: {exc}')
        time.sleep(3600)  # verifica a cada hora


_telegram_last_daily_date = None


def _telegram_daily_report_worker():
    """Thread que envia relatório diário de pedidos ao Telegram."""
    global _telegram_last_daily_date
    import telegram_notifier as tg

    while True:
        try:
            cfg = load_telegram_config()
            if cfg.get('notify_daily_report') and tg.is_configured():
                now = datetime.now()
                today = now.date()
                report_hour = int(cfg.get('daily_report_hour', 8))
                if now.hour >= report_hour and _telegram_last_daily_date != today:
                    _telegram_last_daily_date = today
                    try:
                        all_orders = db_mdb.get_all_orders(status_filter='add')
                    except TypeError:
                        all_orders = db_mdb.get_all_orders()
                    thresholds = load_time_thresholds()
                    msg = tg.build_daily_report_message(all_orders, thresholds)
                    ok, err = tg.send_message(msg)
                    if ok:
                        wms_logger.info('TELEGRAM | Relatório diário enviado')
                    else:
                        wms_logger.warning(f'TELEGRAM | Falha no relatório diário: {err}')
        except Exception as exc:
            wms_logger.error(f'TELEGRAM | Erro no worker de relatório: {exc}')
        time.sleep(600)  # verifica a cada 10 minutos


_telegram_last_period_report_date = None


def _telegram_scheduled_report_worker():
    """Thread que envia o relatório de período (fecha mês) automaticamente."""
    global _telegram_last_period_report_date
    import telegram_notifier as tg

    while True:
        try:
            cfg = load_telegram_config()
            if cfg.get('scheduled_report_enabled') and tg.is_configured():
                now = datetime.now()
                today = now.date()
                report_hour = int(cfg.get('scheduled_report_hour', 8))
                if now.hour >= report_hour and _telegram_last_period_report_date != today:
                    _telegram_last_period_report_date = today
                    try:
                        try:
                            all_orders = db_mdb.get_all_orders()
                        except Exception:
                            all_orders = []
                        dt_from, dt_to = tg.resolve_report_period(cfg)
                        added, removed, active = tg.filter_orders_by_period(all_orders, dt_from, dt_to)
                        unit = DEFAULT_UNIT
                        csv_bytes = tg.build_period_report_csv(added, removed, active, dt_from, dt_to, unit=unit)
                        caption = tg.build_period_report_message(added, removed, active, dt_from, dt_to, unit=unit)
                        fname = f'relatorio_{dt_from.strftime("%Y%m%d")}_{dt_to.strftime("%Y%m%d")}.csv'
                        ok, err = tg.send_document(csv_bytes, fname, caption=caption)
                        if ok:
                            wms_logger.info(f'TELEGRAM | Relatório de período enviado ({dt_from.strftime("%d/%m/%Y")} → {dt_to.strftime("%d/%m/%Y")})')
                        else:
                            wms_logger.warning(f'TELEGRAM | Falha no relatório de período: {err}')
                    except Exception as exc:
                        wms_logger.error(f'TELEGRAM | Erro ao gerar relatório de período: {exc}')
        except Exception as exc:
            wms_logger.error(f'TELEGRAM | Erro no worker de relatório de período: {exc}')
        time.sleep(600)  # verifica a cada 10 minutos


def start_telegram_schedulers():
    """Inicia as threads de alerta de status, relatório diário e relatório de período."""
    t1 = threading.Thread(target=_telegram_status_alert_worker, daemon=True, name='telegram-alerts')
    t1.start()
    t2 = threading.Thread(target=_telegram_daily_report_worker, daemon=True, name='telegram-daily')
    t2.start()
    t3 = threading.Thread(target=_telegram_scheduled_report_worker, daemon=True, name='telegram-period')
    t3.start()
    wms_logger.info('TELEGRAM | Schedulers de alertas, relatório diário e relatório de período iniciados')
    print('[TELEGRAM] Schedulers iniciados.')


# ============================================================================
# OPTO SCHEDULER
# ============================================================================
_OPTO_SCHEDULER_DEFAULTS = {
    'enabled': False,
    'hour': 16,
    'minute': 30,
    'companies': ['2BA', '6VA', '9MA'],
}


def load_opto_scheduler_config() -> dict:
    """Carrega config do agendador OPTO; aplica defaults se ausentes."""
    data = {}
    if os.path.exists(OPTO_SCHEDULER_PATH):
        try:
            with open(OPTO_SCHEDULER_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            data = {}
    # mescla defaults → garante campos ausentes
    return {k: data.get(k, v) for k, v in _OPTO_SCHEDULER_DEFAULTS.items()}


def save_opto_scheduler_config(cfg: dict) -> None:
    """Persiste config do agendador OPTO (escrita atômica)."""
    safe = {k: cfg.get(k, v) for k, v in _OPTO_SCHEDULER_DEFAULTS.items()}
    tmp = OPTO_SCHEDULER_PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(safe, f, indent=2, ensure_ascii=False)
    os.replace(tmp, OPTO_SCHEDULER_PATH)


def _import_integrador_opto():
    """Importa integrador_opto adicionando OPTO_INTEGRATIONS ao sys.path."""
    if getattr(sys, 'frozen', False):
        # Rodando como EXE: OPTO_INTEGRATIONS fica ao lado do executável
        opto_dir = os.path.join(os.path.dirname(sys.executable), 'OPTO_INTEGRATIONS')
    else:
        opto_dir = os.path.normpath(
            os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'OPTO_INTEGRATIONS')
        )
    if opto_dir not in sys.path:
        sys.path.insert(0, opto_dir)
    import importlib
    import integrador_opto as _m  # type: ignore
    return importlib.reload(_m) if _m.__spec__ else _m


def _build_envio_label_data(order_id: str, os_opto: str, position: str,
                             box: str, user: str) -> dict:
    """Monta payload da etiqueta de envio consultando o parser do OPTO.

    Retorna dict parcial (sem campos dioptria/tratamento) se o .txt
    não estiver disponível — não gera exceção.
    """
    base = {
        "os_id": os_opto, "id_master": order_id, "endereco": position,
        "caixa": box, "enviado_por": user, "tratamento": "",
        "od_esf": "", "od_cil": "", "od_eixo": "", "od_ad": "",
        "oe_esf": "", "oe_cil": "", "oe_eixo": "", "oe_ad": "",
    }
    try:
        opto = _import_integrador_opto()
        opto.init_database()
        if hasattr(opto, 'resolve_txt_fields'):
            txt_path, fields = opto.resolve_txt_fields(order_id)
        else:
            txt_path = opto.find_txt(order_id)
            fields   = opto.parse_txt(txt_path)
        row      = opto.build_row(fields)
        from datetime import datetime as _dt
        base.update({
            "tratamento": row[1]  if len(row) > 1  else "",
            "tipo_lente": row[4] if len(row) > 4 else "",
            "fotossensibilidade": row[6] if len(row) > 6 else "",
            "material": row[8] if len(row) > 8 else "",
            "od_esf":     row[10] if len(row) > 10 else "",
            "od_cil":     row[11] if len(row) > 11 else "",
            "od_eixo":    row[12] if len(row) > 12 else "",
            "od_ad":      row[13] if len(row) > 13 else "",
            "oe_esf":     row[14] if len(row) > 14 else "",
            "oe_cil":     row[15] if len(row) > 15 else "",
            "oe_eixo":    row[16] if len(row) > 16 else "",
            "oe_ad":      row[17] if len(row) > 17 else "",
            "data_impressao": _dt.now().strftime("%d/%m/%Y %H:%M"),
        })
    except Exception:
        pass  # não-fatal: etiqueta ainda é gerada com dados disponíveis
    return base


def _opto_scheduler_worker():
    """Worker que executa exportação OPTO programada no horário configurado.
    Verifica a cada 60 s; dispara uma vez por dia quando hora:minuto bater."""
    last_run_date = None
    while True:
        try:
            cfg = load_opto_scheduler_config()
            if cfg.get('enabled'):
                target_h = int(cfg.get('hour', 16))
                target_m = int(cfg.get('minute', 30))
                companies = cfg.get('companies') or None
                now = datetime.now()
                today = now.date()
                # dispara quando hora bate e minuto está na janela de 1 min
                if (now.hour == target_h
                        and target_m <= now.minute < target_m + 2
                        and last_run_date != today):
                    try:
                        integrador_opto = _import_integrador_opto()
                        integrador_opto.init_database()
                        date_str = now.strftime('%d/%m/%Y')
                        res = integrador_opto.generate_scheduled_export(
                            companies=companies, date_str=date_str
                        )
                        files = res.get('files', {}) if isinstance(res, dict) else {}
                        errors = res.get('errors', []) if isinstance(res, dict) else []
                        wms_logger.info(
                            f'OPTO SCHEDULER | {now:%H:%M} | Gerados {len(files)} arquivo(s) | '
                            f'{len(errors)} erro(s): {list(files.keys())}'
                        )
                        last_run_date = today
                    except Exception as exc:
                        wms_logger.error(f'OPTO SCHEDULER | Erro ao gerar exportação: {exc}')
                        last_run_date = today  # não tenta de novo no mesmo dia
        except Exception as exc:
            wms_logger.error(f'OPTO SCHEDULER | Erro no worker: {exc}')
        time.sleep(60)


def start_opto_scheduler():
    """Inicia a thread do agendador OPTO em background."""
    t = threading.Thread(target=_opto_scheduler_worker, daemon=True, name='opto-scheduler')
    t.start()
    cfg = load_opto_scheduler_config()
    status = 'ativado' if cfg.get('enabled') else 'desativado'
    wms_logger.info(
        f'OPTO SCHEDULER | Iniciado (horario={cfg["hour"]:02d}:{cfg["minute"]:02d}, {status})'
    )
    print(f'[OPTO] Agendador iniciado ({cfg["hour"]:02d}:{cfg["minute"]:02d}, {status}).')


def init_db_mode():
    """Aplica o modo de banco salvo (produção/teste) no arranque do servidor."""
    mode = load_db_mode()
    apply_db_mode(mode)
    print(f'[DB] Modo de banco: {mode.upper()} → {db_mdb.get_db_path()}')


# ============================================================================
# APLICAÇÃO FLASK
# ============================================================================

app = Flask(
    __name__,
    template_folder=TEMPLATES_DIR,
    static_folder=os.path.join(RESOURCE_BASE_DIR, 'static'),
    static_url_path='/static'
)
from etiquetas_bp import etq_bp
app.register_blueprint(etq_bp)

from confirmations_bp import confirmations_bp
app.register_blueprint(confirmations_bp)

app.secret_key = os.environ.get('WMS_SECRET_KEY', 'wms-dev-key-insecure')
MASTER_PASSWORD = os.environ.get('WMS_MASTER_PASSWORD', 'masterkey')
DEFAULT_UNIT = db_mdb.DEFAULT_UNIT
DEFAULT_SECTOR = db_mdb.DEFAULT_SECTOR
AVAILABLE_UNITS = list(db_mdb.AVAILABLE_UNITS)
app.jinja_env.globals['permission_labels'] = PERMISSION_FLAGS


# ============================================================================
# RASTREAMENTO DE SESSÕES ATIVAS
# ============================================================================

_active_sessions: dict = {}          # sid → {user, last_seen, ip, unit, sector}
_active_sessions_lock = threading.Lock()
_SESSION_TIMEOUT_MIN  = 30


# ============================================================================
# CONTROLE DE ACESSO POR IP (BLACKLIST / WHITELIST)
# ============================================================================

_ip_acl_lock = threading.Lock()


def _load_ip_acl_file() -> dict:
    try:
        if os.path.exists(IP_ACL_PATH):
            with open(IP_ACL_PATH, 'r', encoding='utf-8') as f:
                d = json.load(f)
                return {
                    'blacklist':      list(d.get('blacklist', [])),
                    'whitelist':      list(d.get('whitelist', [])),
                    'whitelist_mode': bool(d.get('whitelist_mode', False)),
                }
    except Exception:
        pass
    return {'blacklist': [], 'whitelist': [], 'whitelist_mode': False}


def _save_ip_acl_file(data: dict):
    try:
        with open(IP_ACL_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as exc:
        wms_logger.error(f'IP_ACL | Erro ao salvar: {exc}')


def get_ip_acl() -> dict:
    """Retorna cópia atual do ACL de IPs."""
    with _ip_acl_lock:
        return _load_ip_acl_file()


def _valid_ipv4(ip: str) -> bool:
    """Valida formato IPv4 básico."""
    parts = ip.split('.')
    if len(parts) != 4:
        return False
    try:
        return all(0 <= int(p) <= 255 for p in parts)
    except ValueError:
        return False


def check_ip_access(ip: str) -> tuple:
    """Retorna (allowed: bool, reason: str)."""
    with _ip_acl_lock:
        acl = _load_ip_acl_file()
    if ip in acl['whitelist']:
        return True, 'whitelist'
    if ip in acl['blacklist']:
        return False, 'blacklist'
    if acl.get('whitelist_mode') and acl['whitelist']:
        return False, 'whitelist_mode'
    return True, 'ok'


# ============================================================================
# MODO MANUTENÇÃO
# ============================================================================

_maintenance: dict = {
    'active':  False,
    'until':   None,
    'message': 'Sistema em manutenção. Aguarde.',
}
_maintenance_lock = threading.Lock()


def get_maintenance_state() -> dict:
    """Retorna estado atual do modo manutenção (expira automaticamente)."""
    with _maintenance_lock:
        if _maintenance['active'] and _maintenance['until'] and datetime.now() > _maintenance['until']:
            _maintenance['active'] = False
        return dict(_maintenance)


def set_maintenance_state(active: bool, minutes: int = 30, message: str = ''):
    with _maintenance_lock:
        _maintenance['active'] = active
        _maintenance['until'] = datetime.now() + timedelta(minutes=int(minutes)) if active else None
        if message:
            _maintenance['message'] = message
        elif not active:
            _maintenance['message'] = 'Sistema em manutenção. Aguarde.'


def _cleanup_sessions():
    cutoff = datetime.now() - timedelta(minutes=_SESSION_TIMEOUT_MIN)
    with _active_sessions_lock:
        for k in [k for k, v in _active_sessions.items() if v['last_seen'] < cutoff]:
            del _active_sessions[k]


def get_active_users() -> list:
    """Retorna lista de usuários com sessão ativa (últimos 30 min)."""
    _cleanup_sessions()
    with _active_sessions_lock:
        return sorted(_active_sessions.values(), key=lambda x: x['user'])


@app.before_request
def _check_ip_and_maintenance():
    """Verifica blacklist/whitelist de IP e modo manutenção antes de qualquer request."""
    ep = request.endpoint
    # Endpoints sempre isentos
    if ep in ('static', 'maintenance_page', 'maintenance_status'):
        return

    ip = request.remote_addr or ''
    is_admin = session.get('user', '').lower() == 'admin'

    # ── Modo manutenção ──────────────────────────────────────────────────────
    maint = get_maintenance_state()
    if maint['active'] and not is_admin:
        if ep != 'login':  # admin pode fazer login normalmente
            session.clear()
        return redirect(url_for('maintenance_page'))

    # ── ACL de IP ────────────────────────────────────────────────────────────
    if not is_admin:  # admin nunca é bloqueado por IP
        allowed, reason = check_ip_access(ip)
        if not allowed:
            msg = (
                f'Seu IP (<code>{ip}</code>) está na blacklist e não tem permissão de acesso.'
                if reason == 'blacklist' else
                f'Acesso restrito. Seu IP (<code>{ip}</code>) não está na whitelist autorizada.'
            )
            return (
                f'<!doctype html><html lang="pt-BR"><head><meta charset="utf-8">'
                f'<title>Acesso Bloqueado - WMS</title>'
                f'<style>body{{font-family:sans-serif;text-align:center;padding:80px;background:#f8f9fa}}'
                f'h1{{color:#dc3545;font-size:3rem}}p{{font-size:1.2rem;color:#6c757d}}</style></head>'
                f'<body><h1>&#128683; 403</h1><h2>Acesso Bloqueado</h2><p>{msg}</p>'
                f'<p style="margin-top:40px;font-size:.9rem">Entre em contato com o administrador do sistema.</p>'
                f'</body></html>',
                403,
            )


@app.before_request
def _track_active_session():
    """Registra/atualiza sessão ativa a cada request autenticado."""
    if 'user' not in session:
        return
    import secrets as _sec
    sid = session.get('_sid')
    if not sid:
        session['_sid'] = _sec.token_hex(10)
        sid = session['_sid']
    with _active_sessions_lock:
        _active_sessions[sid] = {
            'user':      session['user'],
            'last_seen': datetime.now(),
            'ip':        request.remote_addr or '?',
            'unit':      session.get('unit', ''),
            'sector':    session.get('sector', ''),
        }


def get_sector_permissions(sector_name):
    sectors = load_sectors()
    if sector_name is None:
        return set()
    sector_name = str(sector_name).strip().upper()
    if sector_name == 'ALL':
        return set(PERMISSION_FLAGS.keys())
    sector = sectors.get(sector_name, {})
    return set(sector.get('permissions', []) if isinstance(sector.get('permissions', []), list) else [])


def can_access_feature(feature):
    if session.get('user', '').lower() == 'admin':
        return True
    sector = session.get('sector', DEFAULT_SECTOR)
    if sector == 'ALL':
        return True
    live_permissions = get_sector_permissions(sector)
    if feature in live_permissions:
        return True
    session_permissions = session.get('permissions', [])
    if isinstance(session_permissions, list) and session_permissions:
        return feature in session_permissions
    return False


def get_search_sector_scope():
    """Retorna o setor usado na busca; None libera pesquisa em todos os setores."""
    if can_access_feature('search_cross_sector'):
        return None
    return get_current_sector()


def require_feature_access(feature, message=None, redirect_endpoint='dashboard'):
    if can_access_feature(feature):
        return None
    flash(message or 'Acesso restrito para este setor.', 'danger')
    if redirect_endpoint:
        return redirect(url_for(redirect_endpoint))
    return render_template(
        'error.html',
        title='Acesso restrito',
        message=message or 'Acesso restrito para este setor.'
    ), 403


app.jinja_env.globals['can_access_feature'] = can_access_feature


@app.context_processor
def inject_admin_context():
    """Injeta variáveis globais úteis em todos os templates."""
    sectors = load_sectors()
    current_sec = session.get('sector', DEFAULT_SECTOR)
    # Carregar permissões - versão simplificada para evitar problemas
    try:
        permissions = load_permissions()
    except:
        permissions = {}
    permission_labels = {perm_id: perm['name'] for perm_id, perm in permissions.items()}
    return {
        'is_admin': session.get('user', '').lower() == 'admin',
        'all_units': AVAILABLE_UNITS,
        'current_unit': session.get('unit', DEFAULT_UNIT),
        'all_sectors': sectors,
        'current_sector': current_sec,
        'sector_is_all': current_sec == 'ALL',
        'can_access_triage': can_access_feature('triage'),
        'can_access_etiquetas': can_access_feature('etiquetas'),
        'can_access_dashboard': can_access_feature('dashboard'),
        'can_access_audit': can_access_feature('audit'),
        'can_access_audit_expected_orders': can_access_feature('audit_expected_orders'),
        'can_access_confirmations': can_access_feature('confirmations'),
        'can_access_confirmations_history': can_access_feature('confirmations_history'),
        'can_access_search': can_access_feature('search'),
        'can_access_checkout': can_access_feature('checkout'),
        'can_access_movements': can_access_feature('movements'),
        'can_access_users': can_access_feature('users'),
        'can_access_settings': can_access_feature('settings'),
        'permission_labels': permission_labels,
        'all_permissions': permissions,
    }


# ============================================================================
# MIDDLEWARE — AUTENTICAÇÃO
# ============================================================================

def login_required(f):
    """Decorator para proteger rotas que precisam de autenticação"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('Faça login para continuar', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def is_master(password):
    """Valida se a senha mestre está correta"""
    return password == MASTER_PASSWORD


def get_current_unit():
    """Retorna a unidade associada ao usuário autenticado."""
    return db_mdb.normalize_unit(session.get('unit', DEFAULT_UNIT))


def get_current_sector():
    """Retorna o setor ativo na sessão. None = sem filtro (admin ALL)."""
    sec = session.get('sector', DEFAULT_SECTOR)
    if sec == 'ALL':
        return None
    return sec or DEFAULT_SECTOR


def is_admin_user():
    """Retorna True se o usuário logado é admin."""
    return session.get('user', '').lower() == 'admin'


def can_access_triage():
    """Permite triagem para admin ou usuario do setor TRIAGEM."""
    return can_access_feature('triage')


def require_triage_access():
    """Retorna redirect quando nao ha permissao de triagem."""
    if can_access_triage():
        return None
    flash('Acesso permitido apenas ao setor autorizado para triagem ou admin.', 'danger')
    return redirect(url_for('dashboard'))


def is_valid_unit(unit):
    """Valida se a unidade selecionada existe na lista permitida."""
    return db_mdb.normalize_unit(unit) in AVAILABLE_UNITS

# ============================================================================
# MODEL — UTILITÁRIOS
# ============================================================================

def validate_username(username):
    """Valida formato do nome de usuário"""
    if not username or len(username) < 3:
        return False, "Usuário deve ter pelo menos 3 caracteres"
    if not re.match(r'^[a-zA-Z0-9_-]+$', username):
        return False, "Usuário pode conter apenas letras, números, hífen e underscore"
    return True, ""

def validate_password(password):
    """Valida força da senha"""
    if not password or len(password) < 4:
        return False, "Senha deve ter pelo menos 4 caracteres"
    return True, ""

def is_valid_triage_os(order_id):
    """OS da triagem deve ter exatamente 8 digitos numericos."""
    return bool(re.fullmatch(r'\d{8}', str(order_id or '').strip()))


def is_valid_opto_os(os_opto):
    """OS OPTO deve começar com uma das siglas permitidas."""
    value = str(os_opto or '').strip().upper()
    if not value:
        return False
    return any(value.startswith(prefix) for prefix in OPTO_OS_PREFIXES)

def is_valid_box_number(box):
    """Numero da caixa/cliente aceita somente 1 a 5 digitos."""
    return bool(re.fullmatch(r'\d{1,5}', str(box or '').strip()))

def is_triage_zone(zone, unit):
    """Retorna True quando a zona pertence ao setor TRIAGEM."""
    normalized_zone = str(zone or '').strip().upper()
    if not normalized_zone:
        return False

    triage_shelves = db_mdb.get_all_shelves(unit=unit, sector=TRIAGE_SECTOR)
    for shelf in triage_shelves:
        shelf_zone = str(shelf.get('zone', '')).strip().upper()
        if shelf_zone == normalized_zone:
            return True
    return False

def find_user(username, unit=None):
    """Encontra usuário no banco de dados"""
    return db_mdb.get_user_by_username(username, unit=unit)

def find_shelf(zone, module, unit=None, sector=None):
    """Encontra prateleira específica"""
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    return next((s for s in shelves 
                if s.get('zone') == zone and s.get('module') == module), None)

def get_shelf_positions(zone, module, levels, columns):
    """Gera lista de posições disponíveis em uma prateleira"""
    positions = []
    for level in range(levels, 0, -1):
        if columns == 1:
            positions.append(f"{zone}-{module}-{level:02d}")
        else:
            for col in range(1, columns + 1):
                positions.append(f"{zone}-{module}-{level:02d}-{col:02d}")
    return positions


def shelf_sort_key(shelf):
    """Ordena prateleiras numericamente quando possível e alfabeticamente no restante."""
    module = str(shelf.get('module', '')).strip()
    if module.isdigit():
        return (0, int(module))
    return (1, module.upper())

def count_orders_at_position(position, unit=None, sector=None):
    """Conta quantos pedidos ativos (status add) estão em uma posição"""
    return db_mdb.count_orders_in_position(position, unit=unit, sector=sector)


def get_positions_for_address(address, unit=None, sector=None):
    """Retorna posições do DB que correspondem ao endereço de auditoria.

    Regras de correspondência (address em maiúsculas):
    - 'P-01'       → todas as posições que começam com 'P-01-'
    - 'P-01-02'    → posição exata 'P-01-02' OU que começam com 'P-01-02-'
    - 'P-01-02-03' → exatamente 'P-01-02-03'
    """
    addr = address.strip().upper()
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    all_positions = []
    for shelf in shelves:
        zone = shelf.get('zone', '')
        module = shelf.get('module', '')
        levels = int(shelf.get('levels', 1) or 1)
        columns = int(shelf.get('columns', 1) or 1)
        all_positions.extend(get_shelf_positions(zone, module, levels, columns))
    return [p for p in all_positions
            if p.upper() == addr or p.upper().startswith(addr + '-')]


def normalize_header(text):
    """Normaliza cabecalho para comparacao de planilha."""
    raw = str(text or '').strip().lower()
    raw = unicodedata.normalize('NFKD', raw)
    raw = ''.join(c for c in raw if not unicodedata.combining(c))
    raw = re.sub(r'[^a-z0-9]+', '_', raw)
    return raw.strip('_')


def parse_int(value, default=1):
    """Converte valor para inteiro positivo."""
    text = str(value or '').strip().replace(',', '.')
    if not text:
        return default
    try:
        num = int(float(text))
        return num if num > 0 else default
    except Exception:
        return default


def parse_row_heights(value, levels, default_height=72):
    """Converte alturas das linhas em lista de inteiros (de cima para baixo)."""
    total_levels = max(1, int(levels or 1))
    default_height = max(32, int(default_height or 72))
    raw_text = str(value or '').strip()
    if not raw_text:
        return [default_height] * total_levels

    if raw_text.startswith('['):
        try:
            parsed_json = json.loads(raw_text)
            numbers = parsed_json if isinstance(parsed_json, list) else []
        except Exception:
            numbers = re.split(r'[\s,;]+', raw_text)
    else:
        numbers = re.split(r'[\s,;]+', raw_text)

    heights = []
    for item in numbers:
        try:
            height = int(float(str(item).strip()))
        except Exception:
            continue
        if height > 0:
            heights.append(max(32, height))

    if not heights:
        return [default_height] * total_levels
    if len(heights) == 1 and total_levels > 1:
        return heights * total_levels
    if len(heights) < total_levels:
        heights.extend([heights[-1]] * (total_levels - len(heights)))
    return heights[:total_levels]


def get_shelf_row_heights(shelf, levels, default_height=72):
    """Retorna alturas visuais por linha para uma prateleira."""
    return parse_row_heights(shelf.get('row_heights', ''), levels, default_height=default_height)


def parse_triage_excel_rows(file_storage):
    """Le e mapeia linhas da planilha de recebimento de triagem."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], ['Biblioteca openpyxl nao instalada. Execute: pip install openpyxl']

    wb = load_workbook(filename=file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ['Planilha vazia.']

    header_values = rows[0]
    normalized = [normalize_header(x) for x in header_values]

    aliases = {
        'customer_code': {'codigo_cliente', 'cod_cliente', 'cliente_codigo', 'codigo', 'cliente_id'},
        'customer_name': {'nome_cliente', 'cliente', 'razao_social', 'nome'},
        'quantity': {'quantidade', 'qtd', 'qtde'},
        'received_at': {'data_recebimento', 'data', 'recebido_em', 'data_recebido'},
        'notes': {'observacao', 'observacoes', 'obs', 'comentario'},
        # Mantidos como opcionais por compatibilidade com planilhas antigas.
        'order_id': {'order_id', 'pedido', 'id_pedido', 'numero_pedido', 'n_pedido'},
        'service_name': {'servico', 'servico_nome', 'tipo_servico', 'tipo_de_servico'},
    }

    mapped = {}
    for idx, name in enumerate(normalized):
        for field, options in aliases.items():
            if name in options and field not in mapped:
                mapped[field] = idx

    required = ['customer_code', 'customer_name', 'quantity', 'received_at']
    missing = [x for x in required if x not in mapped]
    if missing:
        return [], [f'Cabecalhos obrigatorios ausentes: {", ".join(missing)}']

    parsed = []
    errors = []
    for row_idx, row in enumerate(rows[1:], start=2):
        order_id = ''
        if 'order_id' in mapped:
            order_id = str(row[mapped['order_id']] or '').strip().upper()

        customer_code = str(row[mapped['customer_code']] or '').strip().upper()
        customer_name = str(row[mapped['customer_name']] or '').strip()
        service_name = ''
        if 'service_name' in mapped:
            service_name = str(row[mapped['service_name']] or '').strip()
        quantity = parse_int(row[mapped['quantity']], default=1)
        received_at = str(row[mapped['received_at']] or '').strip()
        notes = ''
        if 'notes' in mapped:
            notes = str(row[mapped['notes']] or '').strip()

        if not customer_code or not customer_name or not received_at:
            errors.append(f'Linha {row_idx}: campos obrigatorios incompletos.')
            continue

        parsed.append({
            'order_id': order_id,
            'customer_code': customer_code,
            'customer_name': customer_name,
            'service_name': service_name,
            'quantity': quantity,
            'received_at': received_at,
            'notes': notes,
        })

    return parsed, errors

# ============================================================================
# MODEL — ZONAS E TAGS
# ============================================================================

def load_zone_metadata():
    """Carrega descrições de zona salvas localmente."""
    if not os.path.exists(ZONE_METADATA_PATH):
        return {}

    try:
        with open(ZONE_METADATA_PATH, 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
    except Exception as e:
        print(f"Erro ao ler metadados de zona: {e}")
        return {}

    if not isinstance(raw_data, dict):
        return {}

    cleaned = {}
    for zone, name in raw_data.items():
        zone_key = str(zone).strip().upper()
        zone_name = str(name).strip()
        if zone_key and zone_name:
            cleaned[zone_key] = zone_name
    return cleaned

def save_zone_metadata(zone_map):
    """Salva descrições de zona de forma atômica para evitar arquivo corrompido."""
    safe_map = {}
    for zone, name in (zone_map or {}).items():
        zone_key = str(zone).strip().upper()
        zone_name = str(name).strip()
        if zone_key and zone_name:
            safe_map[zone_key] = zone_name

    tmp_path = f"{ZONE_METADATA_PATH}.tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(safe_map, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, ZONE_METADATA_PATH)

def upsert_zone_name(zone, zone_name):
    """Cria/atualiza a descrição de uma zona."""
    zone_key = (zone or '').strip().upper()
    name = (zone_name or '').strip()
    if not zone_key or not name:
        return

    zone_map = load_zone_metadata()
    zone_map[zone_key] = name
    save_zone_metadata(zone_map)

def delete_zone_name(zone):
    """Remove a descrição de uma zona quando ela for excluída."""
    zone_key = (zone or '').strip().upper()
    if not zone_key:
        return

    zone_map = load_zone_metadata()
    if zone_key in zone_map:
        del zone_map[zone_key]
        save_zone_metadata(zone_map)

def normalize_tag_key(tag_name):
    """Padroniza chave interna de tag."""
    name = (tag_name or '').strip()
    if not name:
        return ''
    return re.sub(r'\s+', ' ', name).upper()

def load_tag_catalog():
    """Carrega catalogo de tags com regras."""
    if not os.path.exists(TAG_CATALOG_PATH):
        return {}

    try:
        with open(TAG_CATALOG_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Erro ao ler catalogo de tags: {e}")
        return {}

    if not isinstance(data, dict):
        return {}

    cleaned = {}
    for key, info in data.items():
        tag_key = normalize_tag_key(key)
        if not tag_key or not isinstance(info, dict):
            continue

        display_name = str(info.get('name', '')).strip() or key.title()
        rule = str(info.get('rule', 'none')).strip().lower()
        if rule not in TAG_RULES:
            rule = 'none'
        extra_rules = str(info.get('extra_rules', '')).strip()

        cleaned[tag_key] = {
            'name': display_name,
            'rule': rule,
            'extra_rules': extra_rules
        }
    return cleaned

def save_tag_catalog(tag_catalog):
    """Salva catalogo de tags."""
    safe_data = {}
    for key, info in (tag_catalog or {}).items():
        tag_key = normalize_tag_key(key)
        if not tag_key or not isinstance(info, dict):
            continue

        display_name = str(info.get('name', '')).strip()
        if not display_name:
            continue

        rule = str(info.get('rule', 'none')).strip().lower()
        if rule not in TAG_RULES:
            rule = 'none'

        safe_data[tag_key] = {
            'name': display_name,
            'rule': rule,
            'extra_rules': str(info.get('extra_rules', '')).strip()
        }

    tmp_path = f"{TAG_CATALOG_PATH}.tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(safe_data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, TAG_CATALOG_PATH)

def load_zone_tags_map():
    """Carrega relacao zona -> tags."""
    if not os.path.exists(ZONE_TAGS_PATH):
        return {}

    try:
        with open(ZONE_TAGS_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Erro ao ler mapa de tags por zona: {e}")
        return {}

    if not isinstance(data, dict):
        return {}

    cleaned = {}
    for zone, tags in data.items():
        zone_key = str(zone).strip().upper()
        if not zone_key:
            continue
        if not isinstance(tags, list):
            tags = []

        tag_keys = []
        for tag in tags:
            tag_key = normalize_tag_key(tag)
            if tag_key and tag_key not in tag_keys:
                tag_keys.append(tag_key)

        if tag_keys:
            cleaned[zone_key] = tag_keys
    return cleaned

def save_zone_tags_map(zone_tags_map):
    """Salva relacao zona -> tags."""
    safe_data = {}
    for zone, tags in (zone_tags_map or {}).items():
        zone_key = str(zone).strip().upper()
        if not zone_key:
            continue

        tag_keys = []
        for tag in (tags or []):
            tag_key = normalize_tag_key(tag)
            if tag_key and tag_key not in tag_keys:
                tag_keys.append(tag_key)

        if tag_keys:
            safe_data[zone_key] = tag_keys

    tmp_path = f"{ZONE_TAGS_PATH}.tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(safe_data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, ZONE_TAGS_PATH)

def upsert_tag_definition(tag_name, rule, extra_rules=''):
    """Cria ou atualiza definicao de uma tag."""
    tag_key = normalize_tag_key(tag_name)
    display_name = (tag_name or '').strip()
    tag_rule = (rule or 'none').strip().lower()

    if not tag_key or not display_name:
        return False
    if tag_rule not in TAG_RULES:
        tag_rule = 'none'

    tag_catalog = load_tag_catalog()
    tag_catalog[tag_key] = {
        'name': display_name,
        'rule': tag_rule,
        'extra_rules': (extra_rules or '').strip()
    }
    save_tag_catalog(tag_catalog)
    return True

def attach_tags_to_zone(zone, selected_tag_keys):
    """Anexa tags a uma zona (sem duplicar)."""
    zone_key = (zone or '').strip().upper()
    if not zone_key:
        return

    normalized_keys = []
    for key in (selected_tag_keys or []):
        tag_key = normalize_tag_key(key)
        if tag_key and tag_key not in normalized_keys:
            normalized_keys.append(tag_key)

    if not normalized_keys:
        return

    zone_tags_map = load_zone_tags_map()
    current = zone_tags_map.get(zone_key, [])

    for tag_key in normalized_keys:
        if tag_key not in current:
            current.append(tag_key)

    zone_tags_map[zone_key] = current
    save_zone_tags_map(zone_tags_map)

def remove_zone_tags(zone):
    """Remove vinculos de tags de uma zona apagada."""
    zone_key = (zone or '').strip().upper()
    if not zone_key:
        return

    zone_tags_map = load_zone_tags_map()
    if zone_key in zone_tags_map:
        del zone_tags_map[zone_key]
        save_zone_tags_map(zone_tags_map)

def detach_tags_from_zone(zone, selected_tag_keys):
    """Remove tags especificas de uma zona."""
    zone_key = (zone or '').strip().upper()
    if not zone_key:
        return 0

    zone_tags_map = load_zone_tags_map()
    current = zone_tags_map.get(zone_key, [])
    if not current:
        return 0

    to_remove = []
    for key in (selected_tag_keys or []):
        normalized = normalize_tag_key(key)
        if normalized:
            to_remove.append(normalized)

    if not to_remove:
        return 0

    updated = [key for key in current if key not in to_remove]
    removed_count = len(current) - len(updated)

    if updated:
        zone_tags_map[zone_key] = updated
    else:
        zone_tags_map.pop(zone_key, None)

    save_zone_tags_map(zone_tags_map)
    return removed_count

def delete_tag_definition(tag_key):
    """Exclui uma tag do catalogo e remove de todas as zonas."""
    normalized = normalize_tag_key(tag_key)
    if not normalized:
        return False

    tag_catalog = load_tag_catalog()
    if normalized not in tag_catalog:
        return False

    del tag_catalog[normalized]
    save_tag_catalog(tag_catalog)

    zone_tags_map = load_zone_tags_map()
    changed = False
    for zone, tags in list(zone_tags_map.items()):
        filtered = [tag for tag in tags if tag != normalized]
        if len(filtered) != len(tags):
            changed = True
            if filtered:
                zone_tags_map[zone] = filtered
            else:
                del zone_tags_map[zone]

    if changed:
        save_zone_tags_map(zone_tags_map)

    return True

def zone_has_rule(zone, rule, tag_catalog=None, zone_tags_map=None):
    """Valida se uma zona possui ao menos uma tag com a regra solicitada."""
    zone_key = (zone or '').strip().upper()
    rule_key = (rule or '').strip().lower()
    if not zone_key or not rule_key:
        return False

    tag_catalog = tag_catalog if tag_catalog is not None else load_tag_catalog()
    zone_tags_map = zone_tags_map if zone_tags_map is not None else load_zone_tags_map()
    tag_keys = zone_tags_map.get(zone_key, [])

    for tag_key in tag_keys:
        info = tag_catalog.get(tag_key, {})
        if info.get('rule') == rule_key:
            return True
    return False

def sort_zones_by_priority(zones, tag_catalog=None, zone_tags_map=None):
    """Ordena zonas priorizando as que possuem regra de prioridade."""
    tag_catalog = tag_catalog if tag_catalog is not None else load_tag_catalog()
    zone_tags_map = zone_tags_map if zone_tags_map is not None else load_zone_tags_map()

    def zone_sort_key(zone):
        priority_rank = 0 if zone_has_rule(zone, 'priority', tag_catalog, zone_tags_map) else 1
        return (priority_rank, zone)

    return sorted(zones, key=zone_sort_key)

def get_best_position_for_zone(zone):
    """
    Retorna a melhor posição para armazenar um pedido em uma zona específica.
    Regras de prioridade:
    1. Menor módulo para maior módulo
    2. Dentro do módulo, maior andar para menor andar
    3. Dentro do mesmo andar, menor coluna para maior coluna
    4. Retorna a primeira posição com vaga
    """
    if zone_has_rule(zone, 'maintenance'):
        return None

    unit = get_current_unit()
    sector = get_current_sector()
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    position_counts = db_mdb.count_all_orders_in_positions(unit=unit, sector=sector)
    
    # Filtrar prateleiras da zona especificada
    zone_shelves = [s for s in shelves if s.get('zone', '').upper() == zone.upper()]
    
    if not zone_shelves:
        return None

    # Ordenar módulos do menor para o maior (fallback para string se não for numérico).
    def module_sort_key(shelf):
        module = str(shelf.get('module', '')).strip()
        return (0, int(module)) if module.isdigit() else (1, module)

    zone_shelves = sorted(zone_shelves, key=module_sort_key)

    for shelf in zone_shelves:
        zone_code = shelf.get('zone', '')
        module = shelf.get('module', '')
        levels = shelf.get('levels', 1)
        columns = shelf.get('columns', 1)
        slots = shelf.get('slots', 7)
        
        # Gerar todas as posições desta prateleira
        positions = get_shelf_positions(zone_code, module, levels, columns)
        
        # get_shelf_positions ja retorna do maior andar para o menor.
        # Para colunas, ja retorna da menor para a maior.
        for position in positions:
            count = position_counts.get(position, 0)
            
            # Pular posições cheias
            if count >= slots:
                continue

            return position

    return None


def _audit_now_iso():
    return datetime.now().isoformat(timespec='seconds')


def _audit_sector_key(sector):
    return sector or DEFAULT_SECTOR


def _coerce_scanned_ids(scanned_ids):
    if isinstance(scanned_ids, list):
        raw_lines = [str(item).strip() for item in scanned_ids]
    else:
        raw_lines = str(scanned_ids or '').splitlines()

    # Deduplica preservando ordem para não perder a sequência de bipagem.
    deduped = []
    seen = set()
    for line in raw_lines:
        value = str(line).strip()
        if not value or value in seen:
            continue
        seen.add(value)
        deduped.append(value)
    return deduped


def _load_audit_history_file():
    if not os.path.exists(AUDIT_HISTORY_PATH):
        return {'version': 1, 'drafts': []}

    try:
        with open(AUDIT_HISTORY_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {'version': 1, 'drafts': []}
        drafts = data.get('drafts', [])
        if not isinstance(drafts, list):
            drafts = []
        return {'version': 1, 'drafts': drafts}
    except Exception:
        return {'version': 1, 'drafts': []}


def _save_audit_history_file(data):
    tmp = AUDIT_HISTORY_PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, AUDIT_HISTORY_PATH)


def _trim_audit_history_for_context(drafts, unit, sector, limit=AUDIT_HISTORY_LIMIT):
    sector_key = _audit_sector_key(sector)
    context_items = [
        d for d in drafts
        if str(d.get('unit', '')) == str(unit) and str(d.get('sector', '')) == str(sector_key)
    ]
    other_items = [d for d in drafts if d not in context_items]
    context_items.sort(key=lambda x: str(x.get('updated_at', '')), reverse=True)
    return other_items + context_items[:limit]


def _audit_history_list(unit, sector, limit=AUDIT_HISTORY_LIMIT):
    with AUDIT_HISTORY_LOCK:
        data = _load_audit_history_file()

    sector_key = _audit_sector_key(sector)
    items = [
        d for d in data.get('drafts', [])
        if str(d.get('unit', '')) == str(unit) and str(d.get('sector', '')) == str(sector_key)
    ]
    items.sort(key=lambda x: str(x.get('updated_at', '')), reverse=True)
    return items[:limit]


def _audit_history_get(draft_id, unit, sector):
    with AUDIT_HISTORY_LOCK:
        data = _load_audit_history_file()

    sector_key = _audit_sector_key(sector)
    for item in data.get('drafts', []):
        if str(item.get('draft_id', '')) != str(draft_id):
            continue
        if str(item.get('unit', '')) != str(unit):
            continue
        if str(item.get('sector', '')) != str(sector_key):
            continue
        return item
    return None


def _audit_history_upsert(*, draft_id, address, scanned_ids, username, unit, sector, status='draft', result_summary=None):
    now_iso = _audit_now_iso()
    sector_key = _audit_sector_key(sector)
    clean_address = str(address or '').strip().upper()
    scan_list = _coerce_scanned_ids(scanned_ids)

    with AUDIT_HISTORY_LOCK:
        data = _load_audit_history_file()
        drafts = data.get('drafts', [])

        existing = None
        for item in drafts:
            if str(item.get('draft_id', '')) == str(draft_id or ''):
                existing = item
                break

        if existing is None:
            existing = {
                'draft_id': draft_id or uuid.uuid4().hex,
                'created_at': now_iso,
            }
            drafts.append(existing)

        existing['address'] = clean_address
        existing['username'] = str(username or '')
        existing['unit'] = str(unit or '')
        existing['sector'] = str(sector_key or '')
        existing['status'] = str(status or 'draft')
        existing['updated_at'] = now_iso
        existing['scanned_ids'] = scan_list
        existing['scan_count'] = len(scan_list)
        if result_summary is not None:
            existing['result_summary'] = result_summary

        data['drafts'] = _trim_audit_history_for_context(drafts, unit, sector_key, AUDIT_HISTORY_LIMIT)
        _save_audit_history_file(data)

    return existing


def _to_br_datetime(iso_text):
    text = str(iso_text or '').strip()
    if not text:
        return ''
    try:
        return datetime.fromisoformat(text).strftime('%d/%m/%Y %H:%M:%S')
    except ValueError:
        return text

# ============================================================================
# ROTAS DE API (JSON)
# ============================================================================

@app.route('/api/best-position/<zone>')
@login_required
def api_best_position(zone):
    """Retorna a melhor posição disponível em uma zona (JSON)"""
    if zone_has_rule(zone, 'maintenance'):
        return jsonify({'success': False, 'message': f'Zona {zone} está em manutenção e foi ignorada na alocação'})

    position = get_best_position_for_zone(zone)
    if position:
        return jsonify({'success': True, 'position': position, 'zone': zone})
    return jsonify({'success': False, 'message': f'Nenhuma posição disponível na zona {zone}'})

@app.route('/logo')
def get_logo():
    """Serve a logo da empresa do diretorio local configurado"""
    logo_paths = [
        r"C:\APPS MASTER\IMG\Master_Logo_1.png",
        r"C:\APPS MASTER\IMG\Master_logo_1.png",
        r"\\192.168.1.210\apps master\IMG\Master_Logo_1.png",
        r"\\192.168.1.210\apps master\IMG\Master_logo_1.png",
    ]
    
    try:
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                return send_file(logo_path, mimetype='image/png')
    except Exception as e:
        print(f"Erro ao acessar logo: {e}")
    
    # Fallback: Retornar SVG inline
    svg_content = '''<svg width="400" height="100" xmlns="http://www.w3.org/2000/svg">
        <rect width="400" height="100" fill="#ff9800"/>
        <text x="50%" y="50%" font-size="48" font-weight="bold" fill="white" text-anchor="middle" dy=".3em">WMS</text>
    </svg>'''
    
    return Response(svg_content, mimetype='image/svg+xml')

# ============================================================================
# ROTAS DE AUTENTICAÇÃO
# ============================================================================

@app.route('/')
def index():
    """Página inicial - redireciona para dashboard ou login"""
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        unit = db_mdb.normalize_unit(request.form.get('unit', ''))

        if not is_valid_unit(unit):
            flash('Selecione uma unidade válida', 'danger')
            return render_template(
                'login.html',
                default_unit=DEFAULT_UNIT,
                selected_unit=DEFAULT_UNIT,
                available_units=AVAILABLE_UNITS
            ), 400

        try:
            # Admin pode fazer login em qualquer unidade sem restrição de unit no BD
            search_unit = None if username.lower() == 'admin' else unit
            user = find_user(username, unit=search_unit)
        except RuntimeError as exc:
            # Erro esperado quando driver ODBC do Access nao esta instalado.
            flash(f'Erro de infraestrutura do banco: {exc}', 'danger')
            return render_template(
                'login.html',
                default_unit=DEFAULT_UNIT,
                selected_unit=unit,
                available_units=AVAILABLE_UNITS
            ), 503

        # Compatibilidade/diagnostico: hashes scrypt truncados em VARCHAR(100)
        # sao irrecuperaveis e precisam de redefinicao de senha.
        if user:
            stored_password = user.get('password', '') or ''
            if stored_password.startswith('scrypt:') and len(stored_password) <= 100:
                flash('Senha deste usuario precisa ser redefinida. Use o reset de senha do admin.', 'warning')
                return render_template(
                    'login.html',
                    default_unit=DEFAULT_UNIT,
                    selected_unit=unit,
                    available_units=AVAILABLE_UNITS
                ), 401
        
        if user and db_mdb.verify_password(password, user.get('password', '')):
            session['user'] = username
            session['sector'] = user.get('sector', '')
            # Admin usa a unidade selecionada no login; outros usam a unidade do cadastro
            if username.lower() == 'admin':
                session['unit'] = unit
                session['sector'] = 'ALL'  # Admin vê todos os setores por padrão
                session['permissions'] = list(PERMISSION_FLAGS.keys())
            else:
                session['unit'] = db_mdb.normalize_unit(user.get('unit', unit))
                session['sector'] = user.get('sector', DEFAULT_SECTOR) or DEFAULT_SECTOR
                session['permissions'] = list(get_sector_permissions(session['sector']))
            flash(f'Bem-vindo, {username}!', 'success')
            wms_logger.info(f'LOGIN OK | user={username} unit={session["unit"]} ip={request.remote_addr}')
            return redirect(url_for('dashboard'))
        
        wms_logger.warning(f'LOGIN FALHOU | user={username} unit={unit} ip={request.remote_addr}')
        flash('Usuário ou senha incorretos', 'danger')
    
    return render_template(
        'login.html',
        default_unit=DEFAULT_UNIT,
        selected_unit=DEFAULT_UNIT,
        available_units=AVAILABLE_UNITS
    )

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    """Página de registro de novo usuário (apenas admin)."""
    if not is_admin_user():
        flash('Apenas o admin pode criar novos usuários.', 'danger')
        return redirect(url_for('list_users'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')
        sector = request.form.get('sector', 'geral').strip()
        unit = db_mdb.normalize_unit(request.form.get('unit', ''))

        if not is_valid_unit(unit):
            flash('Selecione uma unidade válida', 'danger')
            return redirect(url_for('register'))
        
        # Validações
        valid, msg = validate_username(username)
        if not valid:
            flash(msg, 'danger')
            return redirect(url_for('register'))
        
        valid, msg = validate_password(password)
        if not valid:
            flash(msg, 'danger')
            return redirect(url_for('register'))
        
        if password != password_confirm:
            flash('As senhas não coincidem', 'danger')
            return redirect(url_for('register'))
        
        if find_user(username, unit=unit):
            flash(f'Este usuário já existe na unidade {unit}', 'danger')
            return redirect(url_for('register'))
        
        # Criar novo usuário no MDB
        db_mdb.add_user(
            username=username,
            password=password,
            sector=sector,
            created_at=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            active=True,
            unit=unit
        )
        
        wms_logger.info(f'REGISTER | novo_user={username} unit={unit} setor={sector}')
        flash(f'Usuário {username} registrado com sucesso na unidade {unit}!', 'success')
        return redirect(url_for('list_users'))
    
    return render_template('register.html', available_units=AVAILABLE_UNITS, default_unit=DEFAULT_UNIT,
                           available_sectors=get_active_sector_keys())

@app.route('/logout')
def logout():
    """Faz logout do usuário"""
    username = session.get('user', 'Usuário')
    sid = session.get('_sid')
    if sid:
        with _active_sessions_lock:
            _active_sessions.pop(sid, None)
    wms_logger.info(f'LOGOUT | user={username}')
    session.clear()
    flash(f'{username} desconectado com sucesso', 'info')
    return redirect(url_for('login'))


@app.route('/switch-unit/<unit_name>')
@login_required
def switch_unit(unit_name):
    """Troca a unidade ativa na sessão (apenas admin)"""
    if not is_admin_user():
        flash('Apenas o admin pode trocar de unidade', 'danger')
        return redirect(url_for('dashboard'))
    unit_name = db_mdb.normalize_unit(unit_name)
    if not is_valid_unit(unit_name):
        flash('Unidade inválida', 'danger')
        return redirect(url_for('dashboard'))
    session['unit'] = unit_name
    flash(f'Unidade alterada para {unit_name}', 'success')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/switch-sector/<sector_name>')
@login_required
def switch_sector(sector_name):
    """Troca o setor ativo na sessão (apenas admin)"""
    if not is_admin_user():
        flash('Apenas o admin pode trocar de setor', 'danger')
        return redirect(url_for('dashboard'))
    if sector_name != 'ALL':
        sectors = load_sectors()
        if sector_name not in sectors:
            flash('Setor inválido', 'danger')
            return redirect(url_for('dashboard'))
    session['sector'] = sector_name
    if sector_name == 'ALL':
        session['permissions'] = list(PERMISSION_FLAGS.keys())
    else:
        session['permissions'] = list(get_sector_permissions(sector_name))
    label = 'Todos os setores' if sector_name == 'ALL' else sector_name
    flash(f'Setor alterado para {label}', 'success')
    return redirect(request.referrer or url_for('dashboard'))


# ============================================================================
# ROTAS DE GERENCIAMENTO DE CÉLULAS/SETORES
# ============================================================================

@app.route('/cells')
@login_required
def list_cells():
    """Lista todas as células/setores cadastrados"""
    sectors = load_sectors()
    return render_template('cells.html', sectors=sectors, permission_labels=PERMISSION_FLAGS)


@app.route('/cells/add', methods=['POST'])
@login_required
def add_cell():
    """Cria uma nova célula/setor"""
    master_key = request.form.get('master_key', '')
    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('list_cells'))

    name = request.form.get('name', '').strip().upper()
    description = request.form.get('description', '').strip()
    permissions = [p for p in request.form.getlist('permissions') if p in PERMISSION_FLAGS]

    if not name:
        flash('Nome do setor é obrigatório', 'danger')
        return redirect(url_for('list_cells'))

    if not re.match(r'^[A-Z0-9_-]+$', name):
        flash('Nome do setor deve conter apenas letras, números, hífen e underscore', 'danger')
        return redirect(url_for('list_cells'))

    sectors = load_sectors()
    if name in sectors:
        flash(f'Setor {name} já existe', 'warning')
        return redirect(url_for('list_cells'))

    sectors[name] = {
        'name': name,
        'description': description,
        'status': 'active',
        'created_at': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        'permissions': permissions
    }
    save_sectors(sectors)

    db_mdb.add_movement(
        username=session.get('user'),
        action='sector_create',
        details=f'Setor {name} criado',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=get_current_unit(),
        sector=DEFAULT_SECTOR
    )

    flash(f'Setor {name} criado com sucesso!', 'success')
    return redirect(url_for('list_cells'))


@app.route('/cells/edit', methods=['POST'])
@login_required
def edit_cell():
    """Edita descrição ou status de um setor"""
    master_key = request.form.get('master_key', '')
    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('list_cells'))

    name = request.form.get('name', '').strip().upper()
    description = request.form.get('description', '').strip()
    status = request.form.get('status', 'active').strip()
    permissions = [p for p in request.form.getlist('permissions') if p in PERMISSION_FLAGS]

    sectors = load_sectors()
    if name not in sectors:
        flash(f'Setor {name} não encontrado', 'danger')
        return redirect(url_for('list_cells'))

    sectors[name]['description'] = description
    sectors[name]['permissions'] = permissions
    sectors[name]['status'] = status if status in ('active', 'inactive') else 'active'
    save_sectors(sectors)

    flash(f'Setor {name} atualizado com sucesso!', 'success')
    return redirect(url_for('list_cells'))


@app.route('/cells/delete', methods=['POST'])
@login_required
def delete_cell():
    """Remove um setor (apenas se não for o padrão)"""
    master_key = request.form.get('master_key', '')
    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('list_cells'))

    name = request.form.get('name', '').strip().upper()

    if name == DEFAULT_SECTOR:
        flash(f'O setor padrão {DEFAULT_SECTOR} não pode ser removido', 'danger')
        return redirect(url_for('list_cells'))

    sectors = load_sectors()
    if name not in sectors:
        flash(f'Setor {name} não encontrado', 'warning')
        return redirect(url_for('list_cells'))

    del sectors[name]
    save_sectors(sectors)

    db_mdb.add_movement(
        username=session.get('user'),
        action='sector_delete',
        details=f'Setor {name} removido',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=get_current_unit(),
        sector=DEFAULT_SECTOR
    )

    flash(f'Setor {name} removido com sucesso!', 'success')
    return redirect(url_for('list_cells'))


# ============================================================================
# ROTAS DE GERENCIAMENTO DE PERMISSÕES
# ============================================================================

@app.route('/permissions/add', methods=['POST'])
@login_required
def add_permission():
    """Cria uma nova permissão"""
    if not session.get('user', '').lower() == 'admin':
        return jsonify({'error': 'Apenas administradores podem criar permissões'}), 403
    
    try:
        perm_id = request.form.get('perm_id', '').strip().lower()
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        icon = request.form.get('icon', 'bi-gear')
        
        if not perm_id or not name:
            flash('ID e Nome da permissão são obrigatórios', 'danger')
            return redirect(url_for('list_cells'))
        
        # Verificar se já existe
        permissions = load_permissions()
        if perm_id in permissions:
            flash(f'Permissão "{perm_id}" já existe', 'danger')
            return redirect(url_for('list_cells'))
        
        # Adicionar nova permissão
        permissions[perm_id] = {
            'id': perm_id,
            'name': name,
            'description': description,
            'icon': icon,
            'created_at': datetime.now().isoformat()
        }
        
        if save_permissions(permissions):
            wms_logger.info(f'PERMISSION CREATE | perm={perm_id} user={session.get("user")}')
            flash(f'Permissão "{name}" criada com sucesso!', 'success')
        else:
            flash('Erro ao salvar permissão', 'danger')
    except Exception as e:
        wms_logger.error(f"Erro ao criar permissão: {e}")
        flash(f'Erro: {str(e)}', 'danger')
    
    return redirect(url_for('list_cells'))


@app.route('/permissions/edit/<perm_id>', methods=['POST'])
@login_required
def edit_permission(perm_id):
    """Edita uma permissão existente"""
    if not session.get('user', '').lower() == 'admin':
        return jsonify({'error': 'Apenas administradores podem editar permissões'}), 403
    
    try:
        permissions = load_permissions()
        if perm_id not in permissions:
            flash('Permissão não encontrada', 'danger')
            return redirect(url_for('list_cells'))
        
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        icon = request.form.get('icon', permissions[perm_id].get('icon', 'bi-gear'))
        
        if not name:
            flash('Nome da permissão é obrigatório', 'danger')
            return redirect(url_for('list_cells'))
        
        permissions[perm_id].update({
            'name': name,
            'description': description,
            'icon': icon
        })
        
        if save_permissions(permissions):
            wms_logger.info(f'PERMISSION EDIT | perm={perm_id} user={session.get("user")}')
            flash(f'Permissão "{name}" atualizada com sucesso!', 'success')
        else:
            flash('Erro ao salvar permissão', 'danger')
    except Exception as e:
        wms_logger.error(f"Erro ao editar permissão: {e}")
        flash(f'Erro: {str(e)}', 'danger')
    
    return redirect(url_for('list_cells'))


@app.route('/permissions/delete/<perm_id>', methods=['POST'])
@login_required
def delete_permission(perm_id):
    """Deleta uma permissão (se não estiver em uso)"""
    if not session.get('user', '').lower() == 'admin':
        return jsonify({'error': 'Apenas administradores podem deletar permissões'}), 403
    
    try:
        permissions = load_permissions()
        if perm_id not in permissions:
            flash('Permissão não encontrada', 'danger')
            return redirect(url_for('list_cells'))
        
        # Verificar se permissão está em uso em algum setor
        sectors = load_sectors()
        in_use = any(perm_id in sector.get('permissions', []) for sector in sectors.values())
        
        if in_use:
            flash(f'Não é possível deletar - permissão "{perm_id}" está em uso em setores', 'danger')
            return redirect(url_for('list_cells'))
        
        perm_name = permissions[perm_id].get('name', perm_id)
        del permissions[perm_id]
        
        if save_permissions(permissions):
            wms_logger.info(f'PERMISSION DELETE | perm={perm_id} user={session.get("user")}')
            flash(f'Permissão "{perm_name}" deletada com sucesso!', 'success')
        else:
            flash('Erro ao salvar permissões', 'danger')
    except Exception as e:
        wms_logger.error(f"Erro ao deletar permissão: {e}")
        flash(f'Erro: {str(e)}', 'danger')
    
    return redirect(url_for('list_cells'))

# ============================================================================
# ROTAS DO DASHBOARD PRINCIPAL
# ============================================================================

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principal com visualização de prateleiras e pedidos"""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.', redirect_endpoint=None)
    if access_denied:
        return access_denied
    try:
        unit = get_current_unit()
        sector = get_current_sector()
        shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
        orders = db_mdb.get_all_orders(status_filter='add', unit=unit, sector=sector)
        
        # Agrupar pedidos ativos por posição
        order_map = {}
        for order in orders:
            pos = order.get('position', '').strip()
            if pos:  # já filtramos apenas 'add' na query
                order_map.setdefault(pos, []).append(order)
        
        # Obter contagem de pedidos para TODAS as posições em uma única query
        position_counts = db_mdb.count_all_orders_in_positions(unit=unit, sector=sector)
        
        # Preparar dados das prateleiras (otimizado - evita N+1 queries)
        shelf_data = []
        for shelf in shelves:
            zone = shelf.get('zone', '').strip()
            module = shelf.get('module', '').strip()
            levels = int(shelf.get('levels', 1) or 1)
            columns = int(shelf.get('columns', 1) or 1)
            slots = int(shelf.get('slots', 7) or 7)
            
            positions = get_shelf_positions(zone, module, levels, columns)
            
            # Usar o dicionário de contagens ao invés de fazer queries individuais
            occupancy = sum(position_counts.get(pos, 0) for pos in positions)
            capacity = len(positions) * slots
            
            shelf_data.append({
                'zone': zone,
                'module': module,
                'levels': levels,
                'columns': columns,
                'slots': slots,
                'positions': positions,
                'occupancy': occupancy,
                'capacity': capacity,
                'usage_percent': int((occupancy / capacity * 100) if capacity > 0 else 0)
            })
        
        # Agrupar prateleiras por zona e coletar nomes/tags
        zone_names = load_zone_metadata()
        tag_catalog = load_tag_catalog()
        zone_tags_map = load_zone_tags_map()

        shelves_by_zone = {}
        for shelf in shelf_data:
            zone = shelf.get('zone', '')
            shelves_by_zone.setdefault(zone, []).append(shelf)

        zone_tags_display = {}
        zone_tag_entries = {}
        zone_rule_flags = {}
        ordered_zones = sort_zones_by_priority(list(shelves_by_zone.keys()), tag_catalog, zone_tags_map)
        zones_data = {}

        for zone in ordered_zones:
            zones_data[zone] = shelves_by_zone.get(zone, [])
            tag_names = []
            tag_entries = []
            for tag_key in zone_tags_map.get(zone, []):
                tag_info = tag_catalog.get(tag_key)
                if tag_info:
                    display_name = tag_info.get('name', tag_key.title())
                    tag_names.append(display_name)
                    tag_entries.append({
                        'key': tag_key,
                        'name': display_name,
                        'rule': tag_info.get('rule', 'none')
                    })
            zone_tags_display[zone] = tag_names
            zone_tag_entries[zone] = tag_entries
            zone_rule_flags[zone] = {
                'maintenance': zone_has_rule(zone, 'maintenance', tag_catalog, zone_tags_map),
                'priority': zone_has_rule(zone, 'priority', tag_catalog, zone_tags_map)
            }

        tag_options = [
            {
                'key': key,
                'name': info.get('name', key.title()),
                'rule': info.get('rule', 'none'),
                'extra_rules': info.get('extra_rules', '')
            }
            for key, info in sorted(tag_catalog.items(), key=lambda item: item[1].get('name', item[0]))
        ]
        all_zone_codes = ordered_zones
        
        # ── Visualização de prateleiras (tema Prateleiras) ───────────────────────
        _thresholds_dash = load_time_thresholds()
        preview_zones_dash = {}
        for shelf in shelf_data:
            z = shelf['zone']
            m = shelf['module']
            lv = shelf['levels']
            cl = shelf['columns']
            sl = shelf['slots']
            row_heights = get_shelf_row_heights(shelf, lv)
            rows_vis = []
            for row_index, level in enumerate(range(lv, 0, -1)):
                cells = []
                for col in range(1, cl + 1):
                    if cl == 1:
                        position = f"{z}-{m}-{level:02d}"
                    else:
                        position = f"{z}-{m}-{level:02d}-{col:02d}"
                    raw_boxes = [make_box_entry(o, _thresholds_dash) for o in order_map.get(position, [])]
                    raw_boxes = list(reversed(raw_boxes))  # mais antiga primeiro → fundo-esquerda
                    cells.append({'position': position, 'boxes': raw_boxes, 'count': len(raw_boxes)})
                rows_vis.append({'level': level, 'cells': cells, 'height_px': row_heights[row_index]})
            preview_zones_dash.setdefault(z, []).append({
                'zone': z,
                'module': m,
                'levels': lv,
                'columns': cl,
                'has_modules': cl > 1,
                'slots': sl,
                'vis_slots': min(sl, 7),
                'rows': rows_vis,
                'occupancy_percent': shelf['usage_percent'],
            })
        ordered_preview_zones_dash = [
            {'zone': z, 'shelves': sorted(preview_zones_dash[z], key=shelf_sort_key)}
            for z in sorted(preview_zones_dash.keys())
        ]

        return render_template('dashboard.html',
                             current_user=session.get('user'),
                             zones=zones_data,
                             zone_names=zone_names,
                             zone_tags=zone_tags_display,
                             zone_tag_entries=zone_tag_entries,
                             zone_rule_flags=zone_rule_flags,
                             tag_rules=TAG_RULES,
                             tag_options=tag_options,
                             order_map=order_map,
                             total_orders=len(orders),
                             preview_zones=ordered_preview_zones_dash,
                             time_thresholds=_thresholds_dash,
                             is_triage_sector=str(sector or '').strip().upper() == TRIAGE_SECTOR)
    except Exception as e:
        flash(f'Erro ao carregar dashboard: {str(e)}', 'danger')
        print(f"ERRO NO DASHBOARD: {e}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('index'))


@app.route('/prototype/shelves')
def shelf_preview():
    """Protótipo público de visualização física das prateleiras."""
    unit = get_current_unit()
    sector = get_current_sector()
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    active_orders = db_mdb.get_all_orders(status_filter='add', unit=unit, sector=sector)
    source_shelves = shelves if shelves else [
        {'zone': 'A', 'module': '01', 'levels': 6, 'columns': 1, 'slots': 7},
        {'zone': 'A', 'module': '02', 'levels': 6, 'columns': 3, 'slots': 3},
        {'zone': 'A', 'module': '03', 'levels': 6, 'columns': 4, 'slots': 3},
        {'zone': 'A', 'module': '04', 'levels': 6, 'columns': 2, 'slots': 4},
        {'zone': 'A', 'module': '05', 'levels': 6, 'columns': 1, 'slots': 7},
    ]

    orders_by_position = {}
    _thresholds_sp = load_time_thresholds()
    for order in active_orders:
        position = order.get('position', '').strip().upper()
        if not position:
            continue
        orders_by_position.setdefault(position, []).append(make_box_entry(order, _thresholds_sp))

    demo_mode = not bool(active_orders)
    demo_positions = {}

    if demo_mode:
        # Preenche ~70% com variação. Módulos de coluna única recebem até 4 colunas
        # visuais de 'slots' caixas — o template distribui horizontalmente (spreading).
        fill_pattern = [1.0, 1.0, 0.75, 0.5, 1.0, 0.75, 0.0, 1.0, 0.5, 0.75]
        box_num = 1000
        pos_idx = 0
        for s in source_shelves:
            z = str(s.get('zone', '')).strip().upper()
            m = str(s.get('module', '')).strip().upper()
            lv = int(s.get('levels', 1) or 1)
            cl = int(s.get('columns', 1) or 1)
            sl = int(s.get('slots', 7) or 7)
            # coluna única: preenche até 4 × slots boxes por posição para mostrar spreading
            vis_cols = 4 if cl == 1 else 1
            for pos in get_shelf_positions(z, m, lv, cl):
                ratio = fill_pattern[pos_idx % len(fill_pattern)]
                count = round(sl * vis_cols * ratio)
                if count > 0:
                    demo_positions[pos] = [f'CX-{box_num + j}' for j in range(count)]
                    box_num += count
                pos_idx += 1

    preview_zones = {}
    for shelf in source_shelves:
        zone = str(shelf.get('zone', '')).strip().upper() or 'SEM ZONA'
        module = str(shelf.get('module', '')).strip().upper() or '01'
        levels = int(shelf.get('levels', 1) or 1)
        columns = int(shelf.get('columns', 1) or 1)
        slots = int(shelf.get('slots', 7) or 7)
        row_heights = get_shelf_row_heights(shelf, levels)

        rows = []
        shelf_display_count = 0
        for row_index, level in enumerate(range(levels, 0, -1)):
            cells = []
            for col in range(1, columns + 1):
                if columns == 1:
                    position = f"{zone}-{module}-{level:02d}"
                else:
                    position = f"{zone}-{module}-{level:02d}-{col:02d}"

                raw_boxes = orders_by_position.get(position, [])
                if demo_mode and not raw_boxes:
                    raw_boxes = [{
                        'label': b,
                        'tier': 'normal',
                        'age_days': None,
                        'order_id': '',
                        'box': b,
                        'created_by': 'Demo',
                        'created_at': ''
                    }
                                 for b in demo_positions.get(position, [])]
                else:
                    raw_boxes = list(reversed(raw_boxes))  # mais antiga primeiro → fundo-esquerda

                shelf_display_count += min(len(raw_boxes), slots)

                cells.append({
                    'position': position,
                    'boxes': raw_boxes,
                    'count': len(raw_boxes)
                })

            rows.append({
                'level': level,
                'cells': cells,
                'height_px': row_heights[row_index],
            })

        preview_zones.setdefault(zone, []).append({
            'zone': zone,
            'module': module,
            'levels': levels,
            'columns': columns,
            'has_modules': columns > 1,
            'slots': slots,
            'vis_slots': min(slots, 7),
            'rows': rows,
            'position_count': levels * columns,
            'occupied_count': shelf_display_count,
            'capacity': levels * columns * slots,
            'occupancy_percent': min(100, int((shelf_display_count / (levels * columns * slots)) * 100)) if (levels * columns * slots) > 0 else 0
        })

    ordered_preview_zones = []
    for zone in sorted(preview_zones.keys()):
        ordered_preview_zones.append({
            'zone': zone,
            'shelves': sorted(preview_zones[zone], key=shelf_sort_key)
        })

    return render_template(
        'shelf_preview.html',
        preview_zones=ordered_preview_zones,
        shelf_total=sum(len(item['shelves']) for item in ordered_preview_zones),
        order_total=len(active_orders),
        demo_mode=demo_mode,
        unit=unit,
        sector=sector or DEFAULT_SECTOR,
        time_thresholds=_thresholds_sp,
    )


@app.route('/prototype/shelves6')
@login_required
def shelf_preview6():
    """Visualização de prateleiras — idêntica ao shelf_preview mas com altura máxima de 6 caixas por slot."""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.', redirect_endpoint=None)
    if access_denied:
        return access_denied

    unit = get_current_unit()
    sector = get_current_sector()
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    active_orders = db_mdb.get_all_orders(status_filter='add', unit=unit, sector=sector)
    source_shelves = shelves if shelves else [
        {'zone': 'A', 'module': '01', 'levels': 6, 'columns': 1, 'slots': 6},
        {'zone': 'A', 'module': '02', 'levels': 6, 'columns': 3, 'slots': 6},
        {'zone': 'A', 'module': '03', 'levels': 6, 'columns': 4, 'slots': 6},
        {'zone': 'A', 'module': '04', 'levels': 6, 'columns': 2, 'slots': 6},
        {'zone': 'A', 'module': '05', 'levels': 6, 'columns': 1, 'slots': 6},
    ]

    _thresholds_s6 = load_time_thresholds()
    orders_by_position = {}
    for order in active_orders:
        position = order.get('position', '').strip().upper()
        if not position:
            continue
        orders_by_position.setdefault(position, []).append(make_box_entry(order, _thresholds_s6))

    demo_mode = not bool(active_orders)
    demo_positions = {}

    if demo_mode:
        fill_pattern = [1.0, 1.0, 0.75, 0.5, 1.0, 0.75, 0.0, 1.0, 0.5, 0.75]
        box_num = 1000
        pos_idx = 0
        for s in source_shelves:
            z = str(s.get('zone', '')).strip().upper()
            m = str(s.get('module', '')).strip().upper()
            lv = int(s.get('levels', 1) or 1)
            cl = int(s.get('columns', 1) or 1)
            # vis_slots fixed at 6
            vis_cols = 4 if cl == 1 else 1
            for pos in get_shelf_positions(z, m, lv, cl):
                ratio = fill_pattern[pos_idx % len(fill_pattern)]
                count = round(6 * vis_cols * ratio)
                if count > 0:
                    demo_positions[pos] = [f'CX-{box_num + j}' for j in range(count)]
                    box_num += count
                pos_idx += 1

    preview_zones = {}
    for shelf in source_shelves:
        zone    = str(shelf.get('zone', '')).strip().upper() or 'SEM ZONA'
        module  = str(shelf.get('module', '')).strip().upper() or '01'
        levels  = int(shelf.get('levels', 1) or 1)
        columns = int(shelf.get('columns', 1) or 1)
        slots   = int(shelf.get('slots', 6) or 6)
        row_heights = get_shelf_row_heights(shelf, levels)

        rows = []
        shelf_display_count = 0
        for row_index, level in enumerate(range(levels, 0, -1)):
            cells = []
            for col in range(1, columns + 1):
                if columns == 1:
                    position = f"{zone}-{module}-{level:02d}"
                else:
                    position = f"{zone}-{module}-{level:02d}-{col:02d}"

                raw_boxes = orders_by_position.get(position, [])
                if demo_mode and not raw_boxes:
                    raw_boxes = [
                        {
                            'label': b, 'tier': 'normal', 'age_days': None,
                            'order_id': '', 'box': b, 'created_by': 'Demo', 'created_at': ''
                        }
                        for b in demo_positions.get(position, [])
                    ]
                else:
                    raw_boxes = list(reversed(raw_boxes))

                shelf_display_count += min(len(raw_boxes), 6)
                cells.append({'position': position, 'boxes': raw_boxes, 'count': len(raw_boxes)})

            rows.append({'level': level, 'cells': cells, 'height_px': row_heights[row_index]})

        capacity = levels * columns * slots
        preview_zones.setdefault(zone, []).append({
            'zone': zone,
            'module': module,
            'levels': levels,
            'columns': columns,
            'has_modules': columns > 1,
            'slots': slots,
            'vis_slots': 6,          # fixo em 6
            'rows': rows,
            'position_count': levels * columns,
            'occupied_count': shelf_display_count,
            'capacity': capacity,
            'occupancy_percent': min(100, int((shelf_display_count / capacity) * 100)) if capacity > 0 else 0,
        })

    ordered_preview_zones = [
        {'zone': z, 'shelves': sorted(preview_zones[z], key=shelf_sort_key)}
        for z in sorted(preview_zones.keys())
    ]

    return render_template(
        'shelf_preview6.html',
        preview_zones=ordered_preview_zones,
        shelf_total=sum(len(item['shelves']) for item in ordered_preview_zones),
        order_total=len(active_orders),
        demo_mode=demo_mode,
        unit=unit,
        sector=sector or DEFAULT_SECTOR,
        time_thresholds=_thresholds_s6,
    )


@app.route('/prototype/triage')
@login_required
def triage_preview():
    """Visualização física das prateleiras da triagem (16 fileiras × 6 caixas de altura)."""
    block = require_triage_access()
    if block:
        return block

    # Constantes físicas da triagem: caixas por fileira e máximo de fileiras por andar
    TRIAGE_VIS_SLOTS = 6   # altura máxima de caixas por fileira
    TRIAGE_VIS_COLS  = 16  # máximo de fileiras por andar

    unit = get_current_unit()
    # Sempre carrega especificamente o setor TRIAGEM
    triage_shelves = db_mdb.get_all_shelves(unit=unit, sector=TRIAGE_SECTOR)
    active_orders = db_mdb.get_all_orders(status_filter='add', unit=unit, sector=TRIAGE_SECTOR)

    # Demo: 2 módulos de triagem — 16 fileiras por andar, 6 caixas de altura
    source_shelves = triage_shelves if triage_shelves else [
        {'zone': 'T', 'module': '01', 'levels': 1, 'columns': 16, 'slots': 6},
        {'zone': 'T', 'module': '02', 'levels': 1, 'columns': 16, 'slots': 6},
    ]

    _thresholds_tp = load_time_thresholds()
    orders_by_position = {}
    for order in active_orders:
        position = order.get('position', '').strip().upper()
        if not position:
            continue
        orders_by_position.setdefault(position, []).append(make_box_entry(order, _thresholds_tp))

    demo_mode = not bool(active_orders)
    demo_positions = {}

    if demo_mode:
        fill_pattern = [1.0, 0.83, 0.67, 0.5, 1.0, 0.0, 0.83, 0.5, 1.0, 0.67,
                        0.5, 1.0, 0.33, 0.83, 0.67, 0.0]
        box_num = 2000
        for s in source_shelves:
            z = str(s.get('zone', '')).strip().upper()
            m = str(s.get('module', '')).strip().upper()
            lv = int(s.get('levels', 1) or 1)
            cl = int(s.get('columns', 1) or 1)
            sl = int(s.get('slots', 6) or 6)
            for pos_idx, pos in enumerate(get_shelf_positions(z, m, lv, cl)):
                # Total de caixas na posição = total capacity para esta posição
                count = sl
                demo_positions[pos] = [f'{box_num + j}' for j in range(count)]
                box_num += count

    preview_zones = {}
    for shelf in source_shelves:
        zone    = str(shelf.get('zone', '')).strip().upper() or 'T'
        module  = str(shelf.get('module', '')).strip().upper() or '01'
        levels  = int(shelf.get('levels', 1) or 1)
        columns = int(shelf.get('columns', 1) or 1)
        slots   = int(shelf.get('slots', 6) or 6)
        row_heights = get_shelf_row_heights(shelf, levels)

        # ── Visual remapping ─────────────────────────────────────────────────
        # Quando a prateleira tem coluna única com muitos slots (ex: columns=1, slots=96),
        # reagrupamos visualmente em TRIAGE_VIS_COLS fileiras de TRIAGE_VIS_SLOTS caixas.
        # columns > 1 já está configurado corretamente no banco.
        if columns == 1:
            vis_cols  = TRIAGE_VIS_COLS
            vis_slots = TRIAGE_VIS_SLOTS
        else:
            vis_cols  = min(columns, TRIAGE_VIS_COLS)
            vis_slots = TRIAGE_VIS_SLOTS

        rows = []
        shelf_display_count = 0
        for row_index, level in enumerate(range(levels, 0, -1)):
            cells = []
            for col in range(1, columns + 1):
                if columns == 1:
                    position = f"{zone}-{module}-{level:02d}"
                else:
                    position = f"{zone}-{module}-{level:02d}-{col:02d}"

                raw_boxes = orders_by_position.get(position, [])
                if demo_mode and not raw_boxes:
                    raw_boxes = [
                        {
                            'label': b,
                            'tier': 'normal',
                            'age_days': None,
                            'order_id': '',
                            'box': b,
                            'created_by': 'Demo',
                            'created_at': '',
                        }
                        for b in demo_positions.get(position, [])
                    ]
                else:
                    raw_boxes = list(reversed(raw_boxes))  # mais antiga → fundo

                shelf_display_count += min(len(raw_boxes), slots)
                cells.append({'position': position, 'boxes': raw_boxes, 'count': len(raw_boxes)})

            rows.append({'level': level, 'cells': cells, 'height_px': row_heights[row_index]})

        capacity = levels * columns * slots
        preview_zones.setdefault(zone, []).append({
            'zone': zone,
            'module': module,
            'levels': levels,
            'columns': columns,       # real DB columns
            'vis_cols': vis_cols,     # visual columns (max 16)
            'vis_slots': vis_slots,   # visual height per column (max 6)
            'has_modules': columns > 1,
            'slots': slots,
            'rows': rows,
            'position_count': levels * columns,
            'occupied_count': shelf_display_count,
            'capacity': capacity,
            'occupancy_percent': min(100, int((shelf_display_count / capacity) * 100)) if capacity > 0 else 0,
        })

    ordered_preview_zones = [
        {'zone': z, 'shelves': sorted(preview_zones[z], key=shelf_sort_key)}
        for z in sorted(preview_zones.keys())
    ]

    return render_template(
        'triage_preview.html',
        preview_zones=ordered_preview_zones,
        order_total=len(active_orders),
        demo_mode=demo_mode,
        unit=unit,
        time_thresholds=_thresholds_tp,
    )


@app.route('/zone/add', methods=['POST'])
@login_required
def add_zone():
    """Cria uma nova zona (apenas para registrar metadados)"""
    master_key = request.form.get('master_key', '')
    
    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('dashboard'))
    
    zone = request.form.get('zone', '').strip().upper()
    zone_name = request.form.get('zone_name', '').strip()
    
    if not zone:
        flash('Código da zona é obrigatório', 'danger')
        return redirect(url_for('dashboard'))

    unit = get_current_unit()

    if zone_name:
        upsert_zone_name(zone, zone_name)
    
    # Registrar movimento de criação de zona
    db_mdb.add_movement(
        username=session.get('user'),
        action='zone_create',
        position=zone,
        order_id='',
        box='',
        details=f'Nova zona criada: {zone_name}' if zone_name else f'Nova zona: {zone}',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=unit,
        sector=get_current_sector() or DEFAULT_SECTOR
    )
    
    flash(f'Zona {zone} criada com sucesso!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/tag/add', methods=['POST'])
@login_required
def add_tag():
    """Cria ou atualiza uma TAG com regra de negocio."""
    master_key = request.form.get('master_key', '')

    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('dashboard'))

    unit = get_current_unit()

    tag_name = request.form.get('tag_name', '').strip()
    tag_rule = request.form.get('tag_rule', 'none').strip().lower()
    extra_rules = request.form.get('extra_rules', '').strip()

    if not tag_name:
        flash('Nome da TAG é obrigatório', 'danger')
        return redirect(url_for('dashboard'))

    if tag_rule not in TAG_RULES:
        flash('Regra da TAG inválida', 'danger')
        return redirect(url_for('dashboard'))

    saved = upsert_tag_definition(tag_name, tag_rule, extra_rules)
    if not saved:
        flash('Não foi possível salvar a TAG', 'danger')
        return redirect(url_for('dashboard'))

    db_mdb.add_movement(
        username=session.get('user'),
        action='tag_upsert',
        position='',
        order_id='',
        box='',
        details=f'TAG {tag_name} ({tag_rule}) criada/atualizada',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=unit,
        sector=get_current_sector() or DEFAULT_SECTOR
    )

    flash(f'TAG {tag_name} salva com sucesso', 'success')
    return redirect(url_for('dashboard'))

@app.route('/tag/remove-zone', methods=['POST'])
@login_required
def remove_tag_from_zone():
    """Remove o vinculo de uma ou mais TAGs de uma zona."""
    zone = request.form.get('zone', '').strip().upper()
    tag_keys = request.form.getlist('tag_keys')
    single_tag_key = request.form.get('tag_key', '').strip()
    if single_tag_key:
        tag_keys.append(single_tag_key)

    if not zone:
        flash('Zona é obrigatória para remover TAG', 'danger')
        return redirect(url_for('dashboard'))

    if not tag_keys:
        flash('Selecione ao menos uma TAG para remover da zona', 'warning')
        return redirect(url_for('dashboard'))

    unit = get_current_unit()

    removed_count = detach_tags_from_zone(zone, tag_keys)
    if removed_count == 0:
        flash(f'Nenhuma TAG removida da zona {zone}', 'warning')
        return redirect(url_for('dashboard'))

    db_mdb.add_movement(
        username=session.get('user'),
        action='tag_zone_remove',
        position=zone,
        order_id='',
        box='',
        details=f'{removed_count} TAG(s) removida(s) da zona {zone}',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=unit,
        sector=get_current_sector() or DEFAULT_SECTOR
    )

    flash(f'{removed_count} TAG(s) removida(s) da zona {zone}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/tag/attach-zone', methods=['POST'])
@login_required
def attach_tag_to_zone():
    """Anexa uma tag existente em uma zona."""
    zone = request.form.get('zone', '').strip().upper()
    tag_key = request.form.get('tag_key', '').strip()

    if not zone:
        flash('Zona é obrigatória', 'danger')
        return redirect(url_for('dashboard'))

    unit = get_current_unit()

    normalized_tag_key = normalize_tag_key(tag_key)
    tag_catalog = load_tag_catalog()
    if normalized_tag_key not in tag_catalog:
        flash('TAG selecionada não existe no catálogo', 'warning')
        return redirect(url_for('dashboard'))

    attach_tags_to_zone(zone, [normalized_tag_key])
    tag_name = tag_catalog[normalized_tag_key].get('name', normalized_tag_key.title())

    db_mdb.add_movement(
        username=session.get('user'),
        action='tag_zone_attach',
        position=zone,
        order_id='',
        box='',
        details=f'TAG {tag_name} anexada na zona {zone}',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=unit,
        sector=get_current_sector() or DEFAULT_SECTOR
    )

    flash(f'TAG {tag_name} adicionada na zona {zone}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/tag/create-attach-zone', methods=['POST'])
@login_required
def create_and_attach_tag_to_zone():
    """Cria uma TAG e anexa diretamente em uma zona."""
    zone = request.form.get('zone', '').strip().upper()
    tag_name = request.form.get('tag_name', '').strip()
    tag_rule = request.form.get('tag_rule', 'none').strip().lower()
    extra_rules = request.form.get('extra_rules', '').strip()

    if not zone:
        flash('Zona é obrigatória', 'danger')
        return redirect(url_for('dashboard'))

    unit = get_current_unit()

    if not tag_name:
        flash('Nome da TAG é obrigatório', 'danger')
        return redirect(url_for('dashboard'))

    if tag_rule not in TAG_RULES:
        flash('Regra da TAG inválida', 'danger')
        return redirect(url_for('dashboard'))

    saved = upsert_tag_definition(tag_name, tag_rule, extra_rules)
    if not saved:
        flash('Não foi possível criar a TAG', 'danger')
        return redirect(url_for('dashboard'))

    normalized_tag_key = normalize_tag_key(tag_name)
    attach_tags_to_zone(zone, [normalized_tag_key])

    db_mdb.add_movement(
        username=session.get('user'),
        action='tag_zone_create_attach',
        position=zone,
        order_id='',
        box='',
        details=f'TAG {tag_name} ({tag_rule}) criada e anexada na zona {zone}',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=unit,
        sector=get_current_sector() or DEFAULT_SECTOR
    )

    flash(f'TAG {tag_name} criada e adicionada na zona {zone}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/tag/delete', methods=['POST'])
@login_required
def delete_tag():
    """Exclui TAG do catalogo e remove vinculos nas zonas."""
    master_key = request.form.get('master_key', '')

    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('dashboard'))

    unit = get_current_unit()

    tag_key = request.form.get('tag_key', '').strip()
    if not tag_key:
        flash('Selecione uma TAG para excluir', 'danger')
        return redirect(url_for('dashboard'))

    deleted = delete_tag_definition(tag_key)
    if not deleted:
        flash('TAG não encontrada no catálogo', 'warning')
        return redirect(url_for('dashboard'))

    db_mdb.add_movement(
        username=session.get('user'),
        action='tag_delete',
        position='',
        order_id='',
        box='',
        details=f'TAG {normalize_tag_key(tag_key)} excluida do catálogo',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=unit,
        sector=get_current_sector() or DEFAULT_SECTOR
    )

    flash('TAG excluída do catálogo com sucesso', 'success')
    return redirect(url_for('dashboard'))

# ============================================================================
# ROTAS DE GERENCIAMENTO DE PRATELEIRAS
# ============================================================================

@app.route('/shelf/add', methods=['POST'])
@login_required
def add_shelf():
    """Adiciona nova prateleira (requer senha mestre)"""
    master_key = request.form.get('master_key', '')
    
    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('dashboard'))
    
    zone = request.form.get('zone', '').strip().upper()
    zone_name = request.form.get('zone_name', '').strip()
    selected_tags = request.form.getlist('zone_tags')
    module = request.form.get('module', '').strip().zfill(2)
    levels = max(1, int(request.form.get('levels', 1) or 1))
    columns = max(1, int(request.form.get('columns', 1) or 1))
    slots = max(1, int(request.form.get('slots', 7) or 7))
    row_heights = parse_row_heights(request.form.get('row_heights', ''), levels)
    
    if not zone or not module:
        flash('Zona e Módulo são obrigatórios', 'danger')
        return redirect(url_for('dashboard'))

    if zone_name:
        upsert_zone_name(zone, zone_name)

    if selected_tags:
        tag_catalog = load_tag_catalog()
        valid_tag_keys = []
        for tag_key in selected_tags:
            normalized = normalize_tag_key(tag_key)
            if normalized in tag_catalog and normalized not in valid_tag_keys:
                valid_tag_keys.append(normalized)
        attach_tags_to_zone(zone, valid_tag_keys)
    
    unit = get_current_unit()

    # Verificar se já existe
    existing = find_shelf(zone, module, unit=unit, sector=get_current_sector())
    if existing:
        if zone_name:
            flash(f'Descrição da zona {zone} atualizada com sucesso', 'info')
        flash(f'Prateleira {zone}-{module} já existe', 'warning')
    else:
        db_mdb.add_shelf(
            zone,
            module,
            levels,
            columns,
            slots,
            unit=unit,
            sector=get_current_sector(),
            row_heights=json.dumps(row_heights, ensure_ascii=False),
        )
        db_mdb.add_movement(
            username=session.get('user'),
            action='shelf_add',
            position=f'{zone}-{module}',
            order_id='',
            box='',
            details=f'Nova prateleira ({zone_name})' if zone_name else 'Nova prateleira',
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            unit=unit,
            sector=get_current_sector() or DEFAULT_SECTOR
        )
        flash(f'Prateleira {zone}-{module} criada com sucesso', 'success')
    
    return redirect(url_for('dashboard'))

@app.route('/shelf/remove', methods=['POST'])
@login_required
def remove_shelf():
    """Remove prateleira (requer senha mestre)"""
    master_key = request.form.get('master_key', '')
    
    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('dashboard'))
    
    zone = request.form.get('zone', '').strip().upper()
    module = request.form.get('module', '').strip().zfill(2)
    
    unit = get_current_unit()
    sector = get_current_sector()
    shelf = find_shelf(zone, module, unit=unit, sector=sector)
    if shelf:
        # ANTES de deletar a prateleira, remover todos os pedidos dela
        levels = shelf.get('levels', 1)
        columns = shelf.get('columns', 1)
        positions = get_shelf_positions(zone, module, levels, columns)
        
        removed_count = 0
        current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        current_user = session.get('user', 'admin')
        
        for position in positions:
            orders = db_mdb.get_orders_by_position(position, unit=unit, sector=sector)
            for order in orders:
                order_id = order.get('order_id', '')
                box = order.get('box', '')
                order_s = order.get('sector') or sector or DEFAULT_SECTOR
                db_mdb.update_order_status(order_id=order_id, status='removed', removed_at=current_time, removed_by=current_user, unit=unit, sector=order_s)
                db_mdb.clear_order_position(order_id, unit=unit, sector=order_s)
                db_mdb.add_movement(username=current_user, action='order_checkout', position=position, order_id=order_id, box=box,
                    details=f'Pedido removido automaticamente - Prateleira {zone}-{module} deletada', timestamp=current_time, unit=unit, sector=order_s)
                removed_count += 1
        
        all_orders = db_mdb.get_all_orders(unit=unit, sector=sector)
        orphaned_shelf_orders = [
            o for o in all_orders 
            if o.get('status') == 'add' and 
            o.get('position', '').upper().startswith(f"{zone}-{module}-")
            and o.get('position', '').upper() not in positions
        ]
        
        for order in orphaned_shelf_orders:
            order_id = order.get('order_id', '')
            box = order.get('box', '')
            position = order.get('position', '')
            order_s = order.get('sector') or sector or DEFAULT_SECTOR
            db_mdb.update_order_status(order_id=order_id, status='removed', removed_at=current_time, removed_by=current_user, unit=unit, sector=order_s)
            db_mdb.add_movement(username=current_user, action='order_checkout', position=position, order_id=order_id, box=box,
                details=f'Pedido removido automaticamente - Prateleira {zone}-{module} deletada (posição órfã)', timestamp=current_time, unit=unit, sector=order_s)
            removed_count += 1
        
        db_mdb.delete_shelf(zone, module, unit=unit)
        db_mdb.add_movement(username=current_user, action='shelf_remove', position=f'{zone}-{module}', order_id='', box='',
            details=f'Prateleira removida - {removed_count} pedido(s) foram dados saída automaticamente', timestamp=current_time, unit=unit, sector=sector or DEFAULT_SECTOR)
        
        if removed_count > 0:
            flash(f'Prateleira {zone}-{module} removida com sucesso. {removed_count} pedido(s) foram dados saída do sistema.', 'success')
        else:
            flash(f'Prateleira {zone}-{module} removida com sucesso', 'success')
    else:
        flash('Prateleira não encontrada', 'warning')
    
    return redirect(url_for('dashboard'))

@app.route('/zone/remove', methods=['POST'])
@login_required
def remove_zone():
    """Remove zona inteira com todos seus módulos e pedidos (requer senha mestre)"""
    master_key = request.form.get('master_key', '')
    
    if not is_master(master_key):
        flash('Senha mestre incorreta', 'danger')
        return redirect(url_for('dashboard'))
    
    zone = request.form.get('zone', '').strip().upper()
    
    if not zone:
        flash('Zona não informada', 'danger')
        return redirect(url_for('dashboard'))
    
    # Buscar todas as prateleiras da zona
    unit = get_current_unit()
    sector = get_current_sector()
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    zone_shelves = [s for s in shelves if s.get('zone', '').upper() == zone]
    
    # Remover todos os pedidos de todas as prateleiras da zona
    total_removed_orders = 0
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    current_user = session.get('user', 'admin')
    
    for shelf in zone_shelves:
        zone_code = shelf.get('zone', '')
        module = shelf.get('module', '')
        levels = shelf.get('levels', 1)
        columns = shelf.get('columns', 1)
        
        positions = get_shelf_positions(zone_code, module, levels, columns)
        
        # Para cada posição da prateleira
        for position in positions:
            # Buscar pedidos ativos naquela posição
            orders = db_mdb.get_orders_by_position(position, unit=unit)
            
            for order in orders:
                order_id = order.get('order_id', '')
                box = order.get('box', '')
                order_s = order.get('sector') or sector or DEFAULT_SECTOR
                
                # Marcar pedido como removido
                db_mdb.update_order_status(
                    order_id=order_id,
                    status='removed',
                    removed_at=current_time,
                    removed_by=current_user,
                    unit=unit,
                    sector=order_s
                )
                
                # Registrar movimento de saída
                db_mdb.add_movement(
                    username=current_user,
                    action='order_checkout',
                    position=position,
                    order_id=order_id,
                    box=box,
                    details=f'Pedido removido automaticamente - Zona {zone_code} deletada',
                    timestamp=current_time,
                    unit=unit,
                    sector=order_s
                )
                
                total_removed_orders += 1
        
        # Deletar a prateleira
        db_mdb.delete_shelf(zone_code, module, unit=unit)
    
    # IMPORTANTE: Também remover QUALQUER pedido que esteja na zona mas em posições órfãs
    # (posições sem prateleira associada)
    all_orders = db_mdb.get_all_orders(unit=unit)
    orphaned_zone_orders = [
        o for o in all_orders 
        if o.get('status') == 'add' and o.get('position', '').strip().upper().startswith(f"{zone}-")
    ]
    
    for order in orphaned_zone_orders:
        order_id = order.get('order_id', '')
        box = order.get('box', '')
        position = order.get('position', '')
        order_s = order.get('sector') or sector or DEFAULT_SECTOR
        
        # Marcar pedido como removido
        db_mdb.update_order_status(
            order_id=order_id,
            status='removed',
            removed_at=current_time,
            removed_by=current_user,
            unit=unit,
            sector=order_s
        )
        db_mdb.clear_order_position(order_id, unit=unit, sector=order_s)
        
        # Registrar movimento de saída
        db_mdb.add_movement(
            username=current_user,
            action='order_checkout',
            position=position,
            order_id=order_id,
            box=box,
            details=f'Pedido removido automaticamente - Zona {zone} deletada (posição órfã)',
            timestamp=current_time,
            unit=unit,
            sector=order_s
        )
        
        total_removed_orders += 1

    if not zone_shelves and total_removed_orders == 0:
        flash('Nenhuma prateleira ou pedido ativo encontrado para esta zona', 'warning')
        return redirect(url_for('dashboard'))
    
    # Registrar movimento de zona removida
    db_mdb.add_movement(
        username=current_user,
        action='zone_remove',
        position=zone,
        order_id='',
        box='',
        details=f'Zona removida - {len(zone_shelves)} módulo(s), {total_removed_orders} pedido(s) foram dados saída automaticamente',
        timestamp=current_time,
        unit=unit,
        sector=sector or DEFAULT_SECTOR
    )

    delete_zone_name(zone)
    remove_zone_tags(zone)
    
    flash(f'Zona {zone} removida com sucesso. {len(zone_shelves)} módulo(s) e {total_removed_orders} pedido(s) foram dados saída do sistema.', 'success')
    
    return redirect(url_for('dashboard'))

# ============================================================================

@app.route('/order/add', methods=['GET', 'POST'])
@login_required
def add_order():
    """Adiciona novo pedido"""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    sector = get_current_sector()
    is_triage_sector = str(sector or '').strip().upper() == TRIAGE_SECTOR
    is_ar_sector = str(sector or '').strip().upper() == AR_SECTOR
    show_os_opto_field = is_ar_sector and can_access_feature('os_opto')
    # Triage zones for this unit — needed when admin (sector=None) uses a triage zone
    _triage_shelves_form = db_mdb.get_all_shelves(unit=unit, sector=TRIAGE_SECTOR)
    triage_zones_for_unit = sorted({ts.get('zone', '').strip().upper() for ts in _triage_shelves_form if ts.get('zone', '').strip()})
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    position_counts = db_mdb.count_all_orders_in_positions(unit=unit, sector=sector)
    quick_mode_from_query = request.args.get('quick', '0') == '1'

    def redirect_add_order(zone_value='', bipador=False, quick=False, level_filled=False, print_params=None):
        """Redireciona para o formulario preservando contexto do modo bipador."""
        params = {}
        if zone_value:
            params['zone'] = zone_value
        if bipador:
            params['bipador'] = '1'
        if quick:
            params['quick'] = '1'
        if level_filled:
            params['lf'] = '1'
        if print_params:
            params.update(print_params)
        return redirect(url_for('add_order', **params))
    
    if request.method == 'POST':
        order_id = request.form.get('order_id', '').strip()
        box = request.form.get('box', '').strip()
        zone = request.form.get('zone', '').strip().upper()
        os_opto = request.form.get('os_opto', '').strip().upper()
        triage_caixa = request.form.get('triage_caixa', '').strip()
        bipador_mode = request.form.get('bipador_mode', '0') == '1'
        quick_mode = request.form.get('quick_mode', '0') == '1'
        
        if not zone:
            flash('Selecione uma zona para alocar o pedido', 'danger')
            return redirect_add_order(bipador=bipador_mode, quick=quick_mode)

        # Bloquear adição de pedidos no modo "Todos os setores" — setor ambíguo
        if sector is None:
            flash('Selecione um setor específico para adicionar pedidos. Use "Trocar Setor" no menu superior.', 'warning')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
        
        if not order_id:
            if is_triage_sector or is_triage_zone(zone, unit):
                flash('Ordem de Servico da empresa e obrigatoria', 'danger')
            else:
                flash('ID do Pedido é obrigatório', 'danger')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)

        if not box:
            flash('Número da Caixa é obrigatório (1 a 5 dígitos numéricos)', 'danger')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)

        if show_os_opto_field:
            if not os_opto:
                flash('A OS OPTO é obrigatória no setor AR.', 'danger')
                return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
            if not is_valid_opto_os(os_opto):
                flash('A OS OPTO deve começar com 9MA, 2BA ou 6VA.', 'danger')
                return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
        elif is_triage_sector or zone in triage_zones_for_unit:
            # Caixa física da triagem — armazenada em os_opto
            if not triage_caixa:
                flash('Número da Caixa é obrigatório para triagem.', 'danger')
                return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
            if not is_valid_box_number(triage_caixa):
                flash('Número da Caixa deve conter apenas dígitos, com no máximo 5 dígitos.', 'danger')
                return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
            os_opto = triage_caixa
        else:
            os_opto = ''

        if not is_valid_triage_os(order_id):
            flash('A OS/ID do pedido deve conter exatamente 8 digitos numericos', 'danger')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)

        if not is_valid_box_number(box):
            flash('Numero da caixa deve conter apenas digitos, com no maximo 5 digitos', 'danger')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
        
        # Regra de negocio: o backend sempre define o endereco automaticamente.
        # Isso evita que o front-end envie uma posicao desatualizada.
        position = get_best_position_for_zone(zone)
        if not position:
            flash(f'Nenhuma posição disponível na zona {zone}', 'warning')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
        
        # Verificar se pedido ja existe NO SETOR ATUAL (permite mesmo ID em setor diferente)
        existing_order = db_mdb.get_order_by_id(order_id, unit=unit, sector=sector)
        if existing_order and existing_order.get('status', 'add') == 'add':
            flash(f'Pedido {order_id} já está ativo no setor {sector}!', 'danger')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
        
        # Validar capacidade usando o dicionário de contagens
        count = position_counts.get(position, 0)
        shelf_info = None
        for s in shelves:
            positions = get_shelf_positions(s.get('zone'), s.get('module'),
                                          s.get('levels', 1), s.get('columns', 1))
            if position in positions:
                shelf_info = s
                break
        
        if shelf_info and count >= shelf_info.get('slots', 7):
            flash(f'Posição {position} está cheia', 'danger')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
        
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        try:
            # Sempre insere novo registro — permite mesmo ID em setores distintos
            db_mdb.add_order(
                position=position,
                order_id=order_id,
                box=box,
                date=now_str,
                timestamp=now_str,
                created_by=session.get('user', 'Sistema'),
                status='add',
                unit=unit,
                sector=sector,
                os_opto=os_opto
            )
            movement_action = 'order_add'
            movement_details = 'Pedido adicionado'
        except Exception:
            flash(f'Falha ao salvar o pedido {order_id}. Tente novamente.', 'danger')
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode)
        
        # Registrar movimento
        db_mdb.add_movement(
            username=session.get('user'),
            action=movement_action,
            position=position,
            order_id=order_id,
            box=box,
            details=movement_details,
            timestamp=now_str,
            unit=unit,
            sector=sector or DEFAULT_SECTOR
        )
        
        flash(f'Pedido {order_id} adicionado à posição {position}', 'success')
        wms_logger.info(f'ORDER ADD | pedido={order_id} cx={box} pos={position} user={session.get("user")} unit={unit}')

        # Build print_params if user asked for auto-print of the envio label
        autoprint_envio = request.form.get('autoprint_envio', '0') == '1'
        print_params = None
        if autoprint_envio:
            print_params = {
                'print_envio': '1',
                'pe_os_id':    os_opto,   # OS OPTO (e.g. 2BA-123456)
                'pe_id_master': order_id, # 8-digit order ID
                'pe_endereco': position,  # allocated position (e.g. P-03-03)
                'pe_caixa':    box,       # box number
            }

        if bipador_mode or quick_mode:
            # Detectar se o andar ficou cheio após este pedido
            slots = shelf_info.get('slots', 7) if shelf_info else 7
            new_count = position_counts.get(position, 0) + 1
            level_just_filled = new_count >= slots
            return redirect_add_order(zone_value=zone, bipador=bipador_mode, quick=quick_mode, level_filled=level_just_filled, print_params=print_params)
        return redirect(url_for('position_detail', code=position))
    
    # GET - Preparar dados para form
    positions = []
    zones = set()
    tag_catalog = load_tag_catalog()
    zone_tags_map = load_zone_tags_map()
    for shelf in shelves:
        zone = shelf.get('zone', '')
        if not zone_has_rule(zone, 'maintenance', tag_catalog, zone_tags_map):
            zones.add(zone)
        module = shelf.get('module', '')
        levels = shelf.get('levels', 1)
        columns = shelf.get('columns', 1)
        positions.extend(get_shelf_positions(zone, module, levels, columns))
    
    requested_zone = request.args.get('zone', '').strip().upper()
    suggested_position = None
    
    if requested_zone and not zone_has_rule(requested_zone, 'maintenance', tag_catalog, zone_tags_map):
        suggested_position = get_best_position_for_zone(requested_zone)
    
    return render_template('order_form.html', 
                         positions=positions, 
                         zones=sort_zones_by_priority(list(zones), tag_catalog, zone_tags_map),
                         suggested_position=suggested_position,
                         requested_zone=requested_zone,
                         quick_mode=quick_mode_from_query,
                         is_triage_sector=is_triage_sector,
                         is_ar_sector=is_ar_sector,
                         show_os_opto_field=show_os_opto_field,
                         triage_zones=triage_zones_for_unit)


@app.route('/order/add/envio/pdf')
@login_required
def order_add_envio_pdf():
    """Gera e serve a etiqueta de envio PDF enriquecida pelo parser do OPTO."""
    import os as _os, sys as _sys
    from io import BytesIO
    from pathlib import Path

    # Garante que o diretório do app está no path para importar o gerador
    _etiq_dir = _os.path.dirname(_os.path.abspath(__file__))
    if _etiq_dir not in _sys.path:
        _sys.path.insert(0, _etiq_dir)
    from etiquetas_100x150 import draw_label_100x150_pdf  # type: ignore

    rq         = request.args
    order_id   = rq.get("id_master",  "").strip()
    os_opto    = rq.get("os_id",       "").strip()
    position   = rq.get("endereco",    "").strip()
    box        = rq.get("caixa",       "").strip()
    user       = rq.get("enviado_por", "") or session.get("user", "")

    # Monta payload enriquecido pelo parser do OPTO (fallback seguro)
    data = _build_envio_label_data(order_id, os_opto, position, box, user)
    # Permite sobrescrever campos via query string (edição manual)
    for k in ("tratamento", "tipo_lente", "fotossensibilidade", "material",
              "od_esf", "od_cil", "od_eixo", "od_ad",
              "oe_esf", "oe_cil", "oe_eixo", "oe_ad"):
        if rq.get(k, ""):
            data[k] = rq.get(k)

    buf = draw_label_100x150_pdf(data)
    pdf_bytes = buf.getvalue()

    # Salva em Impressos/ (não-fatal)
    try:
        from datetime import datetime as _dt
        impressos = Path(DATA_BASE_DIR) / "Impressos"
        impressos.mkdir(parents=True, exist_ok=True)
        ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        fname = f"{ts}_envio_{os_opto or order_id}.pdf"
        (impressos / fname).write_bytes(pdf_bytes)
    except Exception:
        pass

    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=False,
                     download_name=f"etiqueta_envio_{os_opto or order_id}.pdf")


@app.route('/position/<code>')
@login_required
def position_detail(code):
    """Página detalhada de uma posição"""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    sector = get_current_sector()
    orders = db_mdb.get_orders_by_position(code, unit=unit, sector=sector)
    
    # Obter todas as posições disponíveis
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    all_positions = []
    for shelf in shelves:
        zone = shelf.get('zone', '')
        module = shelf.get('module', '')
        levels = shelf.get('levels', 1)
        columns = shelf.get('columns', 1)
        all_positions.extend(get_shelf_positions(zone, module, levels, columns))
    
    return render_template('position_detail.html',
                         code=code,
                         orders=orders,
                         all_positions=all_positions,
                         current_user=session.get('user'))


@app.route('/triagem/recebimento', methods=['GET', 'POST'])
@login_required
def triage_receiving():
    """Cadastro e importacao de recebimento da triagem."""
    access_denied = require_triage_access()
    if access_denied:
        return access_denied

    unit = get_current_unit()
    current_user = session.get('user', 'Sistema')
    query_string = request.args.to_dict(flat=True)
    query_string.pop('edit_id', None)
    return_to = url_for('triage_receiving', **query_string) if query_string else url_for('triage_receiving')

    def format_datetime_local(value):
        text = str(value or '').strip()
        if not text:
            return ''

        for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S'):
            try:
                return datetime.strptime(text, fmt).strftime('%Y-%m-%dT%H:%M')
            except ValueError:
                pass

        try:
            return datetime.fromisoformat(text.replace('Z', '+00:00')).strftime('%Y-%m-%dT%H:%M')
        except ValueError:
            return text

    editing_receipt = None
    edit_id = request.args.get('edit_id', '').strip()
    if edit_id:
        editing_receipt = db_mdb.get_triage_receipt_by_id(edit_id, unit=unit, sector=TRIAGE_SECTOR)
        if not editing_receipt:
            flash('Recebimento nao encontrado para edicao.', 'danger')
            return redirect(return_to)

    form_order_id_value = str(editing_receipt.get('order_id', '')).strip() if editing_receipt else ''
    form_customer_code_value = str(editing_receipt.get('customer_code', '')).strip() if editing_receipt else ''
    form_customer_name_value = str(editing_receipt.get('customer_name', '')).strip() if editing_receipt else ''
    form_quantity_value = int(editing_receipt.get('quantity') or 1) if editing_receipt else 1
    form_received_at_value = format_datetime_local(editing_receipt.get('received_at')) if editing_receipt else ''
    form_notes_value = str(editing_receipt.get('notes', '')).strip() if editing_receipt else ''

    if request.method == 'POST':
        form_action = request.form.get('form_action', 'add').strip().lower()
        post_return_to = request.form.get('return_to', '').strip() or return_to

        if form_action == 'import':
            file_obj = request.files.get('import_file')
            if not file_obj or not file_obj.filename:
                flash('Selecione uma planilha para importar.', 'warning')
                return redirect(post_return_to)

            if not file_obj.filename.lower().endswith('.xlsx'):
                flash('Formato invalido. Envie arquivo .xlsx.', 'danger')
                return redirect(post_return_to)

            rows, parse_errors = parse_triage_excel_rows(file_obj)
            if parse_errors and not rows:
                flash(parse_errors[0], 'danger')
                return redirect(post_return_to)

            inserted = 0
            updated = 0
            for item in rows:
                result = db_mdb.upsert_triage_receipt(
                    order_id=item['order_id'],
                    customer_code=item['customer_code'],
                    customer_name=item['customer_name'],
                    service_name=item['service_name'],
                    quantity=item['quantity'],
                    received_at=item['received_at'],
                    received_by=current_user,
                    notes=item.get('notes', ''),
                    status='received',
                    unit=unit,
                    sector=TRIAGE_SECTOR,
                )
                if result.get('is_new'):
                    inserted += 1
                else:
                    updated += 1

            details = f'Importacao triagem: {inserted} inseridos, {updated} atualizados'
            if parse_errors:
                details += f', {len(parse_errors)} com erro'
            db_mdb.add_movement(
                username=current_user,
                action='triage_receipt_import',
                details=details,
                timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                unit=unit,
                sector=TRIAGE_SECTOR
            )

            flash(f'Importacao concluida. Inseridos: {inserted}. Atualizados: {updated}.', 'success')
            if parse_errors:
                flash(f'Linhas ignoradas por erro: {len(parse_errors)}.', 'warning')
            return redirect(post_return_to)

        if form_action == 'delete':
            receipt_id = request.form.get('receipt_id', '').strip()
            if not receipt_id:
                flash('Selecione um recebimento para excluir.', 'warning')
                return redirect(post_return_to)

            existing_receipt = db_mdb.get_triage_receipt_by_id(receipt_id, unit=unit, sector=TRIAGE_SECTOR)
            if not existing_receipt:
                flash('Recebimento nao encontrado para exclusao.', 'danger')
                return redirect(post_return_to)

            delete_result = db_mdb.delete_triage_receipt_by_id(receipt_id, unit=unit, sector=TRIAGE_SECTOR)
            if delete_result.get('deleted'):
                db_mdb.add_movement(
                    username=current_user,
                    action='triage_receipt_delete',
                    order_id=str(existing_receipt.get('order_id', '')).strip(),
                    details=f"Exclusao de recebimento triagem | Cliente {existing_receipt.get('customer_code', '')}",
                    timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    unit=unit,
                    sector=TRIAGE_SECTOR,
                )
                flash('Recebimento excluido com sucesso.', 'success')
            else:
                flash('Nao foi possivel excluir o recebimento.', 'danger')
            return redirect(post_return_to)

        if form_action == 'edit':
            receipt_id = request.form.get('receipt_id', '').strip()
            if not receipt_id:
                flash('Selecione um recebimento para editar.', 'warning')
                return redirect(post_return_to)

            existing_receipt = db_mdb.get_triage_receipt_by_id(receipt_id, unit=unit, sector=TRIAGE_SECTOR)
            if not existing_receipt:
                flash('Recebimento nao encontrado para edicao.', 'danger')
                return redirect(post_return_to)

            order_id = request.form.get('order_id', '').strip().upper()
            customer_code = request.form.get('customer_code', '').strip().upper()
            customer_name = request.form.get('customer_name', '').strip()
            quantity = parse_int(request.form.get('quantity', '').strip(), default=1)
            received_at = request.form.get('received_at', '').strip()
            notes = request.form.get('notes', '').strip()

            if not customer_code or not received_at:
                flash('Preencha todos os campos obrigatorios da triagem.', 'danger')
                return redirect(post_return_to)

            customer_name_db = db_mdb.get_triage_customer_name_by_code(
                customer_code=customer_code,
                unit=unit,
                sector=TRIAGE_SECTOR,
            )
            customer_name = customer_name_db or customer_name

            if not customer_name:
                flash('Codigo do cliente nao encontrado na base. Cadastre/importe o cliente primeiro.', 'danger')
                return redirect(post_return_to)

            result = db_mdb.update_triage_receipt_by_id(
                receipt_id=receipt_id,
                order_id=order_id or existing_receipt.get('order_id', ''),
                customer_code=customer_code,
                customer_name=customer_name,
                service_name=existing_receipt.get('service_name', ''),
                quantity=quantity,
                received_at=received_at,
                received_by=existing_receipt.get('received_by') or current_user,
                notes=notes,
                status='received',
                unit=unit,
                sector=TRIAGE_SECTOR,
            )

            if result.get('updated'):
                db_mdb.add_movement(
                    username=current_user,
                    action='triage_receipt_edit',
                    order_id=str(order_id or existing_receipt.get('order_id', '')).strip(),
                    details=f'Recebimento triagem editado | Cliente {customer_code}',
                    timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    unit=unit,
                    sector=TRIAGE_SECTOR,
                )
                flash('Recebimento de triagem atualizado com sucesso.', 'success')
            else:
                flash('Nao foi possivel atualizar o recebimento.', 'danger')

            return redirect(post_return_to)

        order_id = request.form.get('order_id', '').strip().upper()
        customer_code = request.form.get('customer_code', '').strip().upper()
        customer_name = request.form.get('customer_name', '').strip()
        quantity = parse_int(request.form.get('quantity', '').strip(), default=1)
        received_at = request.form.get('received_at', '').strip()
        notes = request.form.get('notes', '').strip()

        if not customer_code or not received_at:
            flash('Preencha todos os campos obrigatorios da triagem.', 'danger')
            return redirect(post_return_to)

        # Pedido sempre automatico e sequencial (nao depende do front-end).
        order_id_num = db_mdb.get_next_triage_order_id(unit=unit, sector=TRIAGE_SECTOR)
        order_id = str(order_id_num).zfill(2)

        # Nome do cliente sempre vem do banco com base no codigo informado.
        customer_name_db = db_mdb.get_triage_customer_name_by_code(
            customer_code=customer_code,
            unit=unit,
            sector=TRIAGE_SECTOR,
        )
        customer_name = customer_name_db or customer_name

        if not customer_name:
            flash('Codigo do cliente nao encontrado na base. Cadastre/importe o cliente primeiro.', 'danger')
            return redirect(post_return_to)

        active_client_orders = db_mdb.get_active_orders_by_client_number(
            client_number=customer_code,
            unit=unit,
            sector=TRIAGE_SECTOR,
            limit=8,
        )

        result = db_mdb.upsert_triage_receipt(
            order_id=order_id,
            customer_code=customer_code,
            customer_name=customer_name,
            service_name='',
            quantity=quantity,
            received_at=received_at,
            received_by=current_user,
            notes=notes,
            status='received',
            unit=unit,
            sector=TRIAGE_SECTOR,
        )

        action_name = 'triage_receipt_add' if result.get('is_new') else 'triage_receipt_update'
        db_mdb.add_movement(
            username=current_user,
            action=action_name,
            order_id=order_id,
            details=f'Recebimento triagem | Cliente {customer_code}',
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            unit=unit,
            sector=TRIAGE_SECTOR
        )

        if active_client_orders:
            preview = []
            for row in active_client_orders[:4]:
                order_ref = str(row.get('order_id', '')).strip() or 'SEM_OS'
                position_ref = str(row.get('position', '')).strip() or 'SEM_POSICAO'
                preview.append(f'{order_ref} em {position_ref}')
            resumo = '; '.join(preview)
            extra = ''
            if len(active_client_orders) > 4:
                extra = f' e mais {len(active_client_orders) - 4}'
            flash(
                f'Aviso: cliente {customer_code} ja possui servico(s) enderecado(s) na triagem: {resumo}{extra}.',
                'warning'
            )

        flash(f'Recebimento de triagem salvo com sucesso. Pedido: {order_id}.', 'success')
        return redirect(post_return_to)

    query = request.args.get('q', '').strip()
    filters = {
        'order_id': request.args.get('order_id', '').strip(),
        'customer_code': request.args.get('customer_code', '').strip(),
        'customer_name': request.args.get('customer_name', '').strip(),
        'service_name': request.args.get('service_name', '').strip(),
        'received_by': request.args.get('received_by', '').strip(),
        'notes': request.args.get('notes', '').strip(),
        'date_from': request.args.get('date_from', '').strip(),
        'date_to': request.args.get('date_to', '').strip(),
    }

    has_filters = any([query, *filters.values()])
    if has_filters:
        receipts = db_mdb.search_triage_receipts(
            query=query,
            unit=unit,
            sector=TRIAGE_SECTOR,
            order_id=filters['order_id'],
            customer_code=filters['customer_code'],
            customer_name=filters['customer_name'],
            service_name=filters['service_name'],
            received_by=filters['received_by'],
            notes=filters['notes'],
            date_from=filters['date_from'],
            date_to=filters['date_to'],
        )
    else:
        receipts = db_mdb.get_recent_triage_receipts(limit=150, unit=unit, sector=TRIAGE_SECTOR)

    now_iso = datetime.now().strftime('%Y-%m-%dT%H:%M')
    next_order_id = str(db_mdb.get_next_triage_order_id(unit=unit, sector=TRIAGE_SECTOR)).zfill(2)
    customer_code_suggestions = db_mdb.get_top_triage_customer_codes(limit=8, unit=unit, sector=TRIAGE_SECTOR)
    return render_template(
        'triage_receiving.html',
        receipts=receipts,
        query=query,
        filters=filters,
        customer_code_suggestions=customer_code_suggestions,
        now_iso=now_iso,
        next_order_id=next_order_id,
        return_to=return_to,
        editing_receipt=editing_receipt,
        form_order_id_value=form_order_id_value or next_order_id,
        form_customer_code_value=form_customer_code_value,
        form_customer_name_value=form_customer_name_value,
        form_quantity_value=form_quantity_value,
        form_received_at_value=form_received_at_value or now_iso,
        form_notes_value=form_notes_value,
    )


@app.route('/api/triagem/customer-name', methods=['GET'])
@login_required
def triage_customer_name_lookup():
    """Resolve nome do cliente a partir do codigo para a tela de triagem."""
    access_denied = require_triage_access()
    if access_denied:
        return jsonify({'ok': False, 'message': 'Acesso negado.'}), 403

    code = request.args.get('code', '').strip().upper()
    if not code:
        return jsonify({'ok': False, 'customer_name': ''}), 400

    name = db_mdb.get_triage_customer_name_by_code(
        customer_code=code,
        unit=get_current_unit(),
        sector=TRIAGE_SECTOR,
    )
    return jsonify({'ok': bool(name), 'customer_name': name})


@app.route('/api/triagem/client-active-services', methods=['GET'])
@login_required
def triage_client_active_services_lookup():
    """Retorna servicos ativos da TRIAGEM para o codigo do cliente."""
    access_denied = require_triage_access()
    if access_denied:
        return jsonify({'ok': False, 'message': 'Acesso negado.'}), 403

    code = request.args.get('code', '').strip().upper()
    if not code:
        return jsonify({'ok': False, 'count': 0, 'items': []}), 400

    rows = db_mdb.get_active_orders_by_client_number(
        client_number=code,
        unit=get_current_unit(),
        sector=TRIAGE_SECTOR,
        limit=None,
    )

    items = []
    for row in rows:
        items.append({
            'order_id': str(row.get('order_id', '')).strip(),
            'position': str(row.get('position', '')).strip(),
        })

    return jsonify({'ok': True, 'count': len(items), 'items': items})


@app.route('/api/triagem/next-order-id', methods=['GET'])
@login_required
def triage_next_order_id_lookup():
    """Retorna o proximo pedido sequencial para a tela de triagem."""
    access_denied = require_triage_access()
    if access_denied:
        return jsonify({'ok': False, 'message': 'Acesso negado.'}), 403

    next_order_id = db_mdb.get_next_triage_order_id(
        unit=get_current_unit(),
        sector=TRIAGE_SECTOR,
    )
    return jsonify({'ok': True, 'next_order_id': str(next_order_id).zfill(2)})

@app.route('/order/checkout', methods=['GET', 'POST'])
@login_required
def checkout_order():
    """Dar saída em pedido pelo ID"""
    access_denied = require_feature_access('checkout', 'Acesso restrito à saída de pedidos.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    sector = get_current_sector()
    is_triage_sector = str(sector or '').strip().upper() == TRIAGE_SECTOR
    if request.method == 'POST':
        order_id = request.form.get('order_id', '').strip()
        
        if not order_id:
            flash('Digite o ID do pedido', 'danger')
            return redirect(url_for('checkout_order'))

        if not is_valid_triage_os(order_id):
            flash('A OS/ID do pedido deve conter exatamente 8 digitos numericos', 'danger')
            return redirect(url_for('checkout_order'))

        # Bloquear saída no modo "Todos os setores" — setor ambíguo
        if sector is None:
            flash('Selecione um setor específico para dar saída em pedidos. Use "Trocar Setor" no menu superior.', 'warning')
            return redirect(url_for('checkout_order'))

        order = db_mdb.get_order_by_id(order_id, unit=unit, sector=sector)
        
        if not order or order.get('status', 'add') != 'add':
            flash(f'Pedido {order_id} não encontrado no setor {sector} ou já foi removido', 'warning')
            return redirect(url_for('checkout_order'))
        
        # Atualizar status para removido
        position = order.get('position', '')
        box = order.get('box', '')
        
        db_mdb.update_order_status(
            order_id=order_id,
            status='removed',
            removed_at=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            removed_by=session.get('user', 'Sistema'),
            unit=unit,
            sector=sector
        )
        
        # Registrar movimento
        db_mdb.add_movement(
            username=session.get('user'),
            action='order_checkout',
            position=position,
            order_id=order_id,
            box=box,
            details=f'Saída do sistema - estava em {position}',
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            unit=unit,
            sector=get_current_sector() or DEFAULT_SECTOR
        )
        
        flash(f'✅ Pedido {order_id} retirado com sucesso da posição {position}!', 'success')
        wms_logger.info(f'ORDER CHECKOUT | pedido={order_id} pos={position} user={session.get("user")} unit={unit}')
        return redirect(url_for('checkout_order'))
    
    return render_template('checkout.html', is_triage_sector=is_triage_sector)

@app.route('/order/remove', methods=['POST'])
@login_required
def remove_order():
    """Remove um pedido (muda status para removed)"""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    position = request.form.get('position', '').strip().upper()
    order_id = request.form.get('order_id', '').strip()
    order_sector = get_current_sector()  # pode ser None em modo ALL

    order = db_mdb.get_order_by_id(order_id, unit=unit, sector=order_sector)

    if order and order.get('position') == position:
        actual_sector = order.get('sector') or order_sector or DEFAULT_SECTOR
        db_mdb.update_order_status(order_id, 'removed', unit=unit, sector=actual_sector)
        db_mdb.add_movement(
            username=session.get('user'),
            action='order_remove',
            position=position,
            order_id=order_id,
            box=order.get('box', ''),
            details='Pedido removido',
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            unit=unit,
            sector=actual_sector
        )
        flash(f'Pedido {order_id} removido', 'success')
        wms_logger.info(f'ORDER REMOVE | pedido={order_id} pos={position} user={session.get("user")} unit={unit}')
    else:
        flash('Pedido não encontrado', 'warning')
    
    return redirect(url_for('position_detail', code=position))

@app.route('/order/move', methods=['POST'])
@login_required
def move_order():
    """Move um pedido para outra posição (apenas admin)"""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    # Verificar se o usuário é admin
    current_user = session.get('user', '')
    if current_user.lower() != 'admin':
        flash('Apenas o usuário admin pode mover pedidos!', 'danger')
        position = request.form.get('position', '').strip().upper()
        return redirect(url_for('position_detail', code=position))
    
    position = request.form.get('position', '').strip().upper()
    order_id = request.form.get('order_id', '').strip()
    destination = request.form.get('destination', '').strip().upper()
    
    # Validar se destino foi selecionado
    if not destination:
        flash('Selecione um local para mover o pedido', 'warning')
        return redirect(url_for('position_detail', code=position))
    
    if position == destination:
        flash('Origem e destino são iguais', 'warning')
        return redirect(url_for('position_detail', code=position))
    
    # Buscar pedido — escopar pelo setor atual (None em ALL: retorna qualquer setor)
    move_sector = get_current_sector()
    order = db_mdb.get_order_by_id(order_id, unit=unit, sector=move_sector)
    
    if not order or order.get('position') != position:
        flash('Pedido não encontrado', 'danger')
        return redirect(url_for('position_detail', code=position))
    
    actual_sector = order.get('sector') or move_sector or DEFAULT_SECTOR
    dest_count = count_orders_at_position(destination, unit=unit, sector=actual_sector)
    shelves = db_mdb.get_all_shelves(unit=unit, sector=actual_sector)
    
    dest_capacity = 7  # default
    for shelf in shelves:
        positions = get_shelf_positions(shelf.get('zone'), shelf.get('module'),
                                       shelf.get('levels', 1), shelf.get('columns', 1))
        if destination in positions:
            dest_capacity = shelf.get('slots', 7)
            break
    
    if dest_count >= dest_capacity:
        flash(f'Posição {destination} está cheia', 'danger')
        return redirect(url_for('position_detail', code=position))
    
    # Mover pedido usando UPDATE
    # Para isso, precisamos apenas atualizar a posição do pedido.
    try:
        db_mdb.update_order_position(order_id, destination, unit=unit, sector=actual_sector)
        
        # Registrar movimento
        db_mdb.add_movement(
            username=session.get('user'),
            action='order_move',
            position=f'{position}→{destination}',
            order_id=order_id,
            box=order.get('box', ''),
            details=f'Pedido movido de {position} para {destination}',
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            unit=unit,
            sector=actual_sector
        )
        
        flash(f'Pedido {order_id} movido de {position} para {destination}', 'success')
        wms_logger.info(f'ORDER MOVE | pedido={order_id} {position}->{destination} user={current_user} unit={unit}')
    except Exception as e:
        flash(f'Erro ao mover pedido: {str(e)}', 'danger')
        return redirect(url_for('position_detail', code=position))
    
    return redirect(url_for('position_detail', code=destination))


@app.route('/api/positions/all')
@login_required
def api_positions_all():
    """Retorna todas as posições com ocupação atual (para dropdown de mover)."""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    sector = get_current_sector()
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    position_counts = db_mdb.count_all_orders_in_positions(unit=unit, sector=sector)
    result = []
    for shelf in shelves:
        zone = shelf.get('zone', '')
        module = shelf.get('module', '')
        levels = shelf.get('levels', 1)
        columns = shelf.get('columns', 1)
        slots = shelf.get('slots', 7)
        for pos in get_shelf_positions(zone, module, levels, columns):
            result.append({'position': pos, 'count': position_counts.get(pos, 0), 'capacity': slots})
    result.sort(key=lambda x: x['position'])
    return jsonify({'positions': result})


@app.route('/api/order/remove', methods=['POST'])
@login_required
def api_remove_order():
    """Remove um pedido via AJAX (retorna JSON)."""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    position = request.form.get('position', '').strip().upper()
    order_id = request.form.get('order_id', '').strip()
    if not position or not order_id:
        return jsonify({'ok': False, 'message': 'Dados insuficientes'}), 400
    order = db_mdb.get_order_by_id(order_id, unit=unit, sector=get_current_sector())
    if not order or order.get('position') != position:
        return jsonify({'ok': False, 'message': 'Pedido não encontrado'}), 404
    api_remove_sector = order.get('sector') or get_current_sector() or DEFAULT_SECTOR
    db_mdb.update_order_status(order_id, 'removed', unit=unit, sector=api_remove_sector)
    db_mdb.add_movement(
        username=session.get('user'),
        action='order_remove',
        position=position,
        order_id=order_id,
        box=order.get('box', ''),
        details='Pedido removido via painel visual',
        timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        unit=unit,
        sector=api_remove_sector
    )
    wms_logger.info(f'ORDER REMOVE (api) | pedido={order_id} pos={position} user={session.get("user")} unit={unit}')
    return jsonify({'ok': True, 'message': f'Pedido {order_id} removido'})


@app.route('/api/order/move', methods=['POST'])
@login_required
def api_move_order():
    """Move um pedido via AJAX (retorna JSON). Apenas admin."""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.')
    if access_denied:
        return access_denied
    current_user = session.get('user', '')
    if current_user.lower() != 'admin':
        return jsonify({'ok': False, 'message': 'Apenas admin pode mover pedidos'}), 403
    unit = get_current_unit()
    sector = get_current_sector()
    position = request.form.get('position', '').strip().upper()
    order_id = request.form.get('order_id', '').strip()
    destination = request.form.get('destination', '').strip().upper()
    if not position or not order_id or not destination:
        return jsonify({'ok': False, 'message': 'Dados insuficientes'}), 400
    if position == destination:
        return jsonify({'ok': False, 'message': 'Origem e destino são iguais'})
    order = db_mdb.get_order_by_id(order_id, unit=unit, sector=sector)
    if not order or order.get('position') != position:
        return jsonify({'ok': False, 'message': 'Pedido não encontrado'}), 404
    api_move_sector = order.get('sector') or sector or DEFAULT_SECTOR
    dest_count = count_orders_at_position(destination, unit=unit, sector=api_move_sector)
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    dest_capacity = 7
    for shelf in shelves:
        positions = get_shelf_positions(shelf.get('zone'), shelf.get('module'),
                                        shelf.get('levels', 1), shelf.get('columns', 1))
        if destination in positions:
            dest_capacity = shelf.get('slots', 7)
            break
    if dest_count >= dest_capacity:
        return jsonify({'ok': False, 'message': f'Posição {destination} está cheia ({dest_count}/{dest_capacity})'})
    try:
        db_mdb.update_order_position(order_id, destination, unit=unit, sector=api_move_sector)
        db_mdb.add_movement(
            username=session.get('user'),
            action='order_move',
            position=f'{position}→{destination}',
            order_id=order_id,
            box=order.get('box', ''),
            details=f'Pedido movido de {position} para {destination} via painel visual',
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            unit=unit,
            sector=api_move_sector
        )
        wms_logger.info(f'ORDER MOVE (api) | pedido={order_id} {position}->{destination} user={current_user} unit={unit}')
        return jsonify({'ok': True, 'message': f'Pedido {order_id} movido para {destination}'})
    except Exception as e:
        return jsonify({'ok': False, 'message': f'Erro: {str(e)}'}), 500


# ============================================================================
# ROTAS DE VISUALIZAÇÃO
# ============================================================================

@app.route('/movements')
@login_required
def view_movements():
    """Visualiza histórico de movimentos"""
    access_denied = require_feature_access('movements', 'Acesso restrito ao histórico de movimentos.')
    if access_denied:
        return access_denied
    movement_filters = {
        'date_from': request.args.get('date_from', '').strip(),
        'date_to': request.args.get('date_to', '').strip(),
        'username': request.args.get('username', '').strip(),
        'action': request.args.get('action', '').strip(),
        'order_id': request.args.get('order_id', '').strip(),
        'box': request.args.get('box', '').strip(),
        'position': request.args.get('position', '').strip(),
    }
    has_filters = any(movement_filters.values())
    movements = db_mdb.get_all_movements(
        unit=get_current_unit(),
        sector=get_current_sector(),
        filters=movement_filters if has_filters else None,
    )
    movement_username_suggestions = db_mdb.get_top_movements_suggestions(
        'username',
        limit=8,
        unit=get_current_unit(),
        sector=get_current_sector(),
    )
    movement_action_suggestions = db_mdb.get_top_movements_suggestions(
        'action',
        limit=8,
        unit=get_current_unit(),
        sector=get_current_sector(),
    )
    return render_template(
        'movements.html',
        movements=movements,
        filters=movement_filters,
        has_filters=has_filters,
        movement_username_suggestions=movement_username_suggestions,
        movement_action_suggestions=movement_action_suggestions,
    )

@app.route('/api/level/<zone>/<module>/<int:level>')
@login_required
def api_level_detail(zone, module, level):
    """API que retorna pedidos ativos de um andar em JSON"""
    access_denied = require_feature_access('dashboard', 'Acesso restrito ao dashboard/prateleiras.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    sector = get_current_sector()
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)
    
    # Encontrar a prateleira
    shelf = None
    for s in shelves:
        if s.get('zone') == zone and s.get('module') == module:
            shelf = s
            break
    
    if not shelf:
        return jsonify({'error': 'Prateleira não encontrada'}), 404
    
    # Gerar posições do andar
    columns = shelf.get('columns', 1)
    positions_in_level = []
    if columns == 1:
        positions_in_level.append(f"{zone}-{module}-{level:02d}")
    else:
        for col in range(1, columns + 1):
            positions_in_level.append(f"{zone}-{module}-{level:02d}-{col:02d}")
    
    # Obter pedidos ativos deste andar
    all_orders = db_mdb.get_all_orders(status_filter='add', unit=unit, sector=sector)
    level_orders = [o for o in all_orders if o.get('position', '').strip().upper() in positions_in_level]
    
    return jsonify({
        'level': level,
        'zone': zone,
        'module': module,
        'shelf': shelf,
        'orders': level_orders,
        'total': len(level_orders)
    })

@app.route('/search', methods=['GET'])
@login_required
def search_orders():
    """Página de busca de pedidos"""
    access_denied = require_feature_access('search', 'Acesso restrito à busca de pedidos.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    sector = get_search_sector_scope()
    sort_order = request.args.get('sort', 'DESC').strip().upper()
    if sort_order not in {'ASC', 'DESC'}:
        sort_order = 'DESC'
    all_orders = db_mdb.get_all_orders(status_filter='add', unit=unit, sector=sector, sort_order=sort_order)
    
    # Busca por query string
    query = request.args.get('q', '').strip().upper()
    
    triage_matches = []
    if query:
        filtered_orders = db_mdb.search_orders(query, unit=unit, sector=sector)
        triage_matches = db_mdb.search_triage_receipts(query, unit=unit, sector=TRIAGE_SECTOR)

        triage_order_ids = [item.get('order_id') for item in triage_matches if item.get('order_id')]
        triage_orders = db_mdb.get_active_orders_by_order_ids(
            triage_order_ids,
            unit=unit,
            sector=TRIAGE_SECTOR,
        )

        seen_order_ids = set()
        all_orders = []
        for order in filtered_orders + triage_orders:
            order_id = order.get('order_id')
            if order_id and order_id not in seen_order_ids:
                all_orders.append(order)
                seen_order_ids.add(order_id)

    order_ids = [o.get('order_id') for o in all_orders if o.get('order_id')]
    triage_map = {}
    triage_for_orders = db_mdb.get_triage_receipts_by_order_ids(order_ids, unit=unit, sector=TRIAGE_SECTOR)
    for item in triage_for_orders:
        key = item.get('order_id')
        if key and key not in triage_map:
            triage_map[key] = item

    for order in all_orders:
        triage_info = triage_map.get(order.get('order_id'))
        sector_code = str(order.get('sector', '') or '').strip().upper()
        is_triage_order = sector_code == TRIAGE_SECTOR

        if is_triage_order:
            order['display_customer_code'] = str(order.get('box', '') or '').strip()
            order['display_box'] = str(order.get('os_opto', '') or '').strip()
        else:
            order['display_customer_code'] = ''
            order['display_box'] = str(order.get('box', '') or '').strip()

        if triage_info:
            order['triage_received'] = True
            order['triage_customer_code'] = triage_info.get('customer_code', '')
            order['triage_received_at'] = triage_info.get('received_at', '')
            if not is_triage_order:
                order['display_customer_code'] = order['triage_customer_code']
        else:
            order['triage_received'] = False
            order['triage_customer_code'] = ''
            order['triage_received_at'] = ''

    return render_template('search.html', 
                         orders=all_orders, 
                         total=len(all_orders),
                         query=query,
                         sort_order=sort_order,
                         triage_matches=triage_matches,
                         triage_total=len(triage_matches))

@app.route('/api/search/autocomplete')
@login_required
def search_autocomplete():
    """API de autocomplete para busca de pedidos (apenas ativos)"""
    access_denied = require_feature_access('search', 'Acesso restrito à busca de pedidos.')
    if access_denied:
        return access_denied
    query = request.args.get('q', '').strip().upper()
    
    if not query or len(query) < 2:
        return jsonify({'suggestions': []})
    
    sector = get_search_sector_scope()
    orders = db_mdb.get_all_orders(status_filter='add', unit=get_current_unit(), sector=sector)
    
    suggestions = []
    seen = set()
    cross_sector_search = sector is None
    
    for order in orders:
        order_id = order.get('order_id', '')
        position = order.get('position', '')
        box = order.get('box', '')
        order_sector = str(order.get('sector', '') or '').strip().upper()
        
        order_id_key = ('order_id', order_id, order_sector)
        if order_id.upper().startswith(query) and order_id_key not in seen:
            label = f"📦 Pedido: {order_id}"
            if cross_sector_search and order_sector:
                label += f" • Setor: {order_sector}"
            suggestions.append({
                'value': order_id,
                'label': label,
                'type': 'order_id'
            })
            seen.add(order_id_key)
        
        position_key = ('position', position, order_sector)
        if position.upper().startswith(query) and position_key not in seen:
            label = f"📍 Posição: {position}"
            if cross_sector_search and order_sector:
                label += f" • Setor: {order_sector}"
            suggestions.append({
                'value': position,
                'label': label,
                'type': 'position'
            })
            seen.add(position_key)
        
        box_key = ('box', box, order_sector)
        if box and box.upper().startswith(query) and box_key not in seen:
            label = f"📦 Caixa: {box}"
            if cross_sector_search and order_sector:
                label += f" • Setor: {order_sector}"
            suggestions.append({
                'value': box,
                'label': label,
                'type': 'box'
            })
            seen.add(box_key)
    
    suggestions = suggestions[:10]
    
    return jsonify({'suggestions': suggestions})

@app.route('/users')
@login_required
def list_users():
    """Lista todos os usuários"""
    access_denied = require_feature_access('users', 'Acesso restrito ao gerenciamento de usuários.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    users = db_mdb.get_all_users(unit=unit)
    current_user = session.get('user')
    
    return render_template('users.html', users=users, current_user=current_user, current_unit=unit)

@app.route('/user/reset-password', methods=['POST'])
@login_required
def reset_user_password():
    """Reseta a senha de um usuário (apenas admin)"""
    access_denied = require_feature_access('users', 'Acesso restrito ao gerenciamento de usuários.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    target_username = request.form.get('username', '').strip()
    new_password = request.form.get('new_password', '').strip()
    master_pass = request.form.get('master_password', '').strip()
    
    # Validar senha mestre
    if not is_master(master_pass):
        flash('Senha mestre incorreta!', 'danger')
        return redirect(url_for('list_users'))
    
    # Validar nova senha
    valid, msg = validate_password(new_password)
    if not valid:
        flash(msg, 'danger')
        return redirect(url_for('list_users'))
    
    # Verificar se usuário existe
    user = find_user(target_username, unit=unit)
    if not user:
        flash(f'Usuário "{target_username}" não encontrado!', 'danger')
        return redirect(url_for('list_users'))
    
    # Atualizar senha
    try:
        db_mdb.update_user(target_username, unit=unit, password=new_password)
        wms_logger.info(f'USER RESET-SENHA | alvo={target_username} unit={unit} por={session.get("user")}')
        flash(f'Senha de "{target_username}" alterada com sucesso!', 'success')
        
        # Registrar auditoria
        db_mdb.add_movement(
            username=session.get('user'),
            action='reset_password',
            details=f'Senha do usuário "{target_username}" foi resetada',
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            unit=unit,
            sector=get_current_sector() or DEFAULT_SECTOR
        )
    except Exception as e:
        flash(f'Erro ao resetar senha: {str(e)}', 'danger')
    
    return redirect(url_for('list_users'))

@app.route('/user/toggle-status', methods=['POST'])
@login_required
def toggle_user_status():
    """Ativa ou desativa um usuário (apenas admin)"""
    access_denied = require_feature_access('users', 'Acesso restrito ao gerenciamento de usuários.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    target_username = request.form.get('username', '').strip()
    master_pass = request.form.get('master_password', '').strip()
    
    # Validar senha mestre
    if not is_master(master_pass):
        flash('Senha mestre incorreta!', 'danger')
        return redirect(url_for('list_users'))
    
    # Não permitir desativar o próprio usuário
    if target_username == session.get('user'):
        flash('Você não pode desativar sua própria conta!', 'warning')
        return redirect(url_for('list_users'))
    
    # Verificar se usuário existe
    user = find_user(target_username, unit=unit)
    if not user:
        flash(f'Usuário "{target_username}" não encontrado!', 'danger')
        return redirect(url_for('list_users'))
    
    # Alternar status
    try:
        new_status = not bool(user.get('active', True))
        db_mdb.update_user(target_username, unit=unit, active=new_status)
        
        status_text = 'ativado' if new_status else 'desativado'
        wms_logger.info(f'USER TOGGLE | alvo={target_username} status={status_text} unit={unit} por={session.get("user")}')
        flash(f'Usuário "{target_username}" {status_text} com sucesso!', 'success')
        
        # Registrar auditoria
        db_mdb.add_movement(
            username=session.get('user'),
            action='toggle_user_status',
            details=f'Usuário "{target_username}" {status_text}',
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            unit=unit,
            sector=get_current_sector() or DEFAULT_SECTOR
        )
    except Exception as e:
        flash(f'Erro ao alterar status: {str(e)}', 'danger')
    
    return redirect(url_for('list_users'))

@app.route('/user/delete', methods=['POST'])
@login_required
def delete_user():
    """Deleta um usuário do sistema (apenas admin)"""
    access_denied = require_feature_access('users', 'Acesso restrito ao gerenciamento de usuários.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    target_username = request.form.get('username', '').strip()
    master_pass = request.form.get('master_password', '').strip()
    
    # Validar senha mestre
    if not is_master(master_pass):
        flash('Senha mestre incorreta!', 'danger')
        return redirect(url_for('list_users'))
    
    # Não permitir deletar o próprio usuário
    if target_username == session.get('user'):
        flash('Você não pode deletar sua própria conta!', 'warning')
        return redirect(url_for('list_users'))
    
    # Verificar se usuário existe
    user = find_user(target_username, unit=unit)
    if not user:
        flash(f'Usuário "{target_username}" não encontrado!', 'danger')
        return redirect(url_for('list_users'))
    
    # Deletar usuário
    try:
        db_mdb.delete_user(target_username, unit=unit)
        wms_logger.warning(f'USER DELETE | alvo={target_username} unit={unit} por={session.get("user")}')
        flash(f'Usuário "{target_username}" deletado com sucesso!', 'success')
        
        # Registrar auditoria
        db_mdb.add_movement(
            username=session.get('user'),
            action='delete_user',
            details=f'Usuário "{target_username}" foi deletado do sistema',
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            unit=unit,
            sector=get_current_sector() or DEFAULT_SECTOR
        )
    except Exception as e:
        flash(f'Erro ao deletar usuário: {str(e)}', 'danger')
    
    return redirect(url_for('list_users'))

@app.route('/user/edit-sector', methods=['POST'])
@login_required
def edit_user_sector():
    """Edita o setor de um usuário (apenas admin)"""
    access_denied = require_feature_access('users', 'Acesso restrito ao gerenciamento de usuários.')
    if access_denied:
        return access_denied
    unit = get_current_unit()
    target_username = request.form.get('username', '').strip()
    new_sector = request.form.get('sector', '').strip()
    master_pass = request.form.get('master_password', '').strip()
    
    # Validar senha mestre
    if not is_master(master_pass):
        flash('Senha mestre incorreta!', 'danger')
        return redirect(url_for('list_users'))
    
    # Verificar se usuário existe
    user = find_user(target_username, unit=unit)
    if not user:
        flash(f'Usuário "{target_username}" não encontrado!', 'danger')
        return redirect(url_for('list_users'))
    
    # Atualizar setor
    try:
        db_mdb.update_user(target_username, unit=unit, sector=new_sector)
        wms_logger.info(f'USER EDIT-SETOR | alvo={target_username} setor_novo="{new_sector or "Geral"}" unit={unit} por={session.get("user")}')
        flash(f'Setor de "{target_username}" atualizado para "{new_sector or "Geral"}"!', 'success')
        
        # Registrar auditoria
        db_mdb.add_movement(
            username=session.get('user'),
            action='edit_user_sector',
            details=f'Setor do usuário "{target_username}" alterado para "{new_sector or "Geral"}"',
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            unit=unit,
            sector=get_current_sector() or DEFAULT_SECTOR
        )
    except Exception as e:
        flash(f'Erro ao editar setor: {str(e)}', 'danger')
    
    return redirect(url_for('list_users'))

@app.route('/about')
def about():
    """Página de informações"""
    return render_template('about.html')


@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    """Página de configurações — acesso restrito a admin."""
    access_denied = require_feature_access('settings', 'Acesso restrito às configurações.', redirect_endpoint=None)
    if access_denied:
        return access_denied
    thresholds = load_time_thresholds()
    tg_cfg = load_telegram_config()
    if request.method == 'POST':
        form_type = request.form.get('form_type', 'thresholds')

        if form_type == 'telegram' and is_admin_user():
            tg_cfg['notify_status_alerts'] = 'notify_status_alerts' in request.form
            tg_cfg['notify_daily_report']  = 'notify_daily_report'  in request.form
            raw_tiers = request.form.getlist('notify_tiers')
            tg_cfg['notify_tiers'] = [t for t in raw_tiers if t in ('attention', 'urgent', 'critical')]
            try:
                tg_cfg['daily_report_hour'] = max(0, min(23, int(request.form.get('daily_report_hour', 8))))
            except (ValueError, TypeError):
                tg_cfg['daily_report_hour'] = 8
            # Relatório de período
            tg_cfg['scheduled_report_enabled']   = 'scheduled_report_enabled' in request.form
            tg_cfg['scheduled_report_mode']       = request.form.get('scheduled_report_mode', 'month_to_date')
            if tg_cfg['scheduled_report_mode'] not in ('month_to_date', 'full_month', 'custom_days'):
                tg_cfg['scheduled_report_mode'] = 'month_to_date'
            try:
                tg_cfg['scheduled_report_hour']      = max(0, min(23, int(request.form.get('scheduled_report_hour', 8))))
                tg_cfg['scheduled_report_start_day'] = max(1, min(28, int(request.form.get('scheduled_report_start_day', 1))))
                tg_cfg['scheduled_report_end_day']   = max(0, min(31, int(request.form.get('scheduled_report_end_day', 0))))
            except (ValueError, TypeError):
                pass
            save_telegram_config(tg_cfg)
            flash('Configurações do Telegram salvas!', 'success')
        else:
            try:
                green_days  = int(request.form.get('green_days',  thresholds['green_days']))
                yellow_days = int(request.form.get('yellow_days', thresholds['yellow_days']))
                red_days    = int(request.form.get('red_days',    thresholds['red_days']))
                if not (0 < green_days < yellow_days < red_days):
                    flash('Os limiares devem ser crescentes e maiores que zero.', 'danger')
                else:
                    save_time_thresholds(green_days, yellow_days, red_days)
                    flash('Configurações salvas com sucesso!', 'success')
                    thresholds = load_time_thresholds()
            except (ValueError, TypeError):
                flash('Valores inválidos. Informe números inteiros.', 'danger')

    import telegram_notifier as tg
    db_mode = load_db_mode()
    active_users = get_active_users() if is_admin_user() else []
    opto_cfg = load_opto_scheduler_config()
    return render_template('settings.html',
                           thresholds=thresholds,
                           backup_log=get_backup_log_tail(),
                           backup_dir=BACKUP_DIR,
                           tg_cfg=tg_cfg,
                           tg_configured=tg.is_configured(),
                           db_mode=db_mode,
                           db_path_prod=db_mdb.DB_PATH_PROD,
                           db_path_test=db_mdb.DB_PATH_TEST,
                           db_path_active=db_mdb.get_db_path(),
                           active_users=active_users,
                           ip_acl=get_ip_acl(),
                           maintenance_state=get_maintenance_state(),
                           opto_cfg=opto_cfg)


@app.route('/admin/set-db-mode', methods=['POST'])
@login_required
def admin_set_db_mode():
    """Troca o banco de dados ativo (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'})
    mode = request.form.get('mode', 'production')
    if mode not in ('production', 'test'):
        return jsonify({'ok': False, 'message': 'Modo inválido.'})
    try:
        apply_db_mode(mode)
        save_db_mode(mode)
        label = 'Produção' if mode == 'production' else 'Teste'
        wms_logger.warning(f'DB | Banco alterado para {label} por {session.get("user")}')
        return jsonify({'ok': True, 'message': f'Banco de {label} ativado.', 'path': db_mdb.get_db_path()})
    except Exception as exc:
        return jsonify({'ok': False, 'message': str(exc)})


@app.route('/admin/ip/add', methods=['POST'])
@login_required
def admin_ip_add():
    """Adiciona IP à blacklist ou whitelist (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'})
    ip   = (request.form.get('ip', '') or '').strip()
    lst  = request.form.get('list', '')  # 'blacklist' | 'whitelist'
    if not ip or not _valid_ipv4(ip):
        return jsonify({'ok': False, 'message': 'IP inválido.'})
    if lst not in ('blacklist', 'whitelist'):
        return jsonify({'ok': False, 'message': 'Lista inválida.'})
    with _ip_acl_lock:
        acl = _load_ip_acl_file()
        other = 'whitelist' if lst == 'blacklist' else 'blacklist'
        # Remove da lista oposta para evitar conflito
        if ip in acl[other]:
            acl[other].remove(ip)
        if ip not in acl[lst]:
            acl[lst].append(ip)
        _save_ip_acl_file(acl)
    wms_logger.warning(f'IP_ACL | {ip} adicionado à {lst} por {session.get("user")}')
    return jsonify({'ok': True, 'message': f'IP {ip} adicionado à {lst}.', 'acl': get_ip_acl()})


@app.route('/admin/ip/remove', methods=['POST'])
@login_required
def admin_ip_remove():
    """Remove IP da blacklist ou whitelist (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'})
    ip  = (request.form.get('ip', '') or '').strip()
    lst = request.form.get('list', '')
    if not ip or lst not in ('blacklist', 'whitelist'):
        return jsonify({'ok': False, 'message': 'Parâmetros inválidos.'})
    with _ip_acl_lock:
        acl = _load_ip_acl_file()
        if ip in acl[lst]:
            acl[lst].remove(ip)
            _save_ip_acl_file(acl)
    wms_logger.info(f'IP_ACL | {ip} removido da {lst} por {session.get("user")}')
    return jsonify({'ok': True, 'message': f'IP {ip} removido da {lst}.', 'acl': get_ip_acl()})


@app.route('/admin/ip/whitelist-mode', methods=['POST'])
@login_required
def admin_ip_whitelist_mode():
    """Ativa/desativa modo whitelist (somente IPs na whitelist têm acesso)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'})
    enabled = request.form.get('enabled', 'false').lower() == 'true'
    with _ip_acl_lock:
        acl = _load_ip_acl_file()
        acl['whitelist_mode'] = enabled
        _save_ip_acl_file(acl)
    label = 'ativado' if enabled else 'desativado'
    wms_logger.warning(f'IP_ACL | Modo whitelist {label} por {session.get("user")}')
    return jsonify({'ok': True, 'message': f'Modo whitelist {label}.', 'acl': get_ip_acl()})


@app.route('/admin/maintenance', methods=['POST'])
@login_required
def admin_maintenance():
    """Ativa/desativa modo manutenção e kica todos os usuários não-admin."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'})
    action = request.form.get('action', '')  # 'start' | 'stop'
    if action == 'start':
        try:
            minutes = max(1, min(480, int(request.form.get('minutes', 30))))
        except (ValueError, TypeError):
            minutes = 30
        message = (request.form.get('message', '') or 'Sistema em manutenção. Aguarde.').strip()
        set_maintenance_state(True, minutes=minutes, message=message)
        # Invalida sessões ativas de não-admins na próxima requisição
        with _active_sessions_lock:
            for k, v in list(_active_sessions.items()):
                if v.get('user', '').lower() != 'admin':
                    del _active_sessions[k]
        state = get_maintenance_state()
        until_str = state['until'].strftime('%H:%M:%S') if state['until'] else '?'
        wms_logger.warning(f'MANUT | Modo manutenção ativado por {minutes}min por {session.get("user")}')
        return jsonify({'ok': True, 'message': f'Manutenção ativa até {until_str}.', 'state': {
            'active': True, 'until': until_str, 'message': message
        }})
    elif action == 'stop':
        set_maintenance_state(False)
        wms_logger.warning(f'MANUT | Modo manutenção desativado por {session.get("user")}')
        return jsonify({'ok': True, 'message': 'Manutenção encerrada.', 'state': {'active': False}})
    return jsonify({'ok': False, 'message': 'Ação inválida. Use "start" ou "stop".'})


@app.route('/maintenance')
def maintenance_page():
    """Página pública exibida durante modo manutenção."""
    state = get_maintenance_state()
    until_iso = state['until'].isoformat() if state.get('until') else ''
    message = state.get('message', 'Sistema em manutenção. Aguarde.')
    return render_template('maintenance.html', message=message, until_iso=until_iso)


@app.route('/api/maintenance-status')
def maintenance_status():
    """Retorna o estado atual do modo manutenção sem redirecionar."""
    state = get_maintenance_state()
    payload = {
        'active': bool(state.get('active')),
        'until': state['until'].isoformat() if state.get('until') else None,
        'message': state.get('message', 'Sistema em manutenção. Aguarde.'),
    }
    response = jsonify(payload)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    return response


@app.route('/admin/log-tail')
@login_required
def admin_log_tail():
    """Retorna as últimas N linhas do WMS.log (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'lines': []}), 403
    n = min(int(request.args.get('n', 80)), 300)
    lines = []
    try:
        if os.path.isfile(_WMS_LOG_PATH):
            with open(_WMS_LOG_PATH, 'r', encoding='utf-8', errors='replace') as f:
                all_lines = f.readlines()
            lines = [l.rstrip() for l in all_lines[-n:]]
    except Exception:
        pass
    return jsonify({'ok': True, 'lines': lines})


@app.route('/admin/backup', methods=['POST'])
@login_required
def admin_backup():
    """Backup manual do banco de dados (somente admin)."""
    if session.get('user', '').lower() != 'admin':
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('settings'))
    user = session.get('user', 'admin')
    ok, msg = perform_backup(triggered_by=f'{user} (manual)')
    if ok:
        flash(f'✓ {msg}', 'success')
    else:
        flash(f'Erro no backup: {msg}', 'danger')
    return redirect(url_for('settings'))


@app.route('/admin/opto/schedule', methods=['POST'])
@login_required
def admin_opto_schedule():
    """Salva configuração do agendador OPTO (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'}), 403
    try:
        enabled = request.form.get('enabled') == '1'
        hour   = max(0, min(23, int(request.form.get('hour',   16))))
        minute = max(0, min(59, int(request.form.get('minute', 30))))
        companies = [c.strip().upper() for c in request.form.getlist('companies') if c.strip()]
        if not companies:
            companies = list(_OPTO_SCHEDULER_DEFAULTS['companies'])
        cfg = {'enabled': enabled, 'hour': hour, 'minute': minute, 'companies': companies}
        save_opto_scheduler_config(cfg)
        status_lbl = 'ativado' if enabled else 'desativado'
        wms_logger.warning(
            f'OPTO SCHEDULER | Config alterada por {session.get("user")}: '
            f'ativo={enabled}, horário={hour:02d}:{minute:02d}'
        )
        return jsonify({
            'ok': True,
            'message': f'Configurado para {hour:02d}:{minute:02d} ({status_lbl}).',
            'cfg': cfg,
        })
    except (ValueError, TypeError) as exc:
        return jsonify({'ok': False, 'message': f'Valores inválidos: {exc}'}), 400


@app.route('/admin/opto/generate-now', methods=['POST'])
@login_required
def admin_opto_generate_now():
    """Gera as planilhas OPTO imediatamente para hoje (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'}), 403
    try:
        integrador_opto = _import_integrador_opto()
        cfg = load_opto_scheduler_config()
        companies = cfg.get('companies') or None
        date_str = datetime.now().strftime('%d/%m/%Y')
        integrador_opto.init_database()
        res = integrador_opto.generate_scheduled_export(companies=companies, date_str=date_str)
        files_map = res.get('files', {}) if isinstance(res, dict) else {}
        errors = res.get('errors', []) if isinstance(res, dict) else []
        files = [os.path.basename(p) for p in files_map]
        n_ok = sum(files_map.values()) if files_map else 0
        parts = []
        if n_ok:
            parts.append(f'{n_ok} pedido(s) em {len(files)} planilha(s)')
        if errors:
            parts.append(f'{len(errors)} com erro (ver ERROS_INTEGRACAO)')
        msg = ' | '.join(parts) if parts else 'Nenhum pedido encontrado para hoje.'
        wms_logger.info(
            f'OPTO SCHEDULER | Geração manual por {session.get("user")}: '
            f'{n_ok} ok, {len(errors)} erro(s)'
        )
        return jsonify({
            'ok': True, 'message': msg, 'files': files,
            'errors': errors, 'generated': files_map,
        })
    except Exception as exc:
        wms_logger.error(f'OPTO SCHEDULER | Erro na geração manual: {exc}')
        return jsonify({'ok': False, 'message': str(exc)}), 500


@app.route('/audit', methods=['GET', 'POST'])
@login_required
def audit_select():
    """Seleção de endereço para entrar no modo de conferência."""
    access_denied = require_feature_access('audit', 'Acesso restrito à conferência de endereço.')
    if access_denied:
        return access_denied
    if request.method == 'POST':
        address = request.form.get('address', '').strip().upper()
        if not address:
            flash('Informe um endereço para conferir.', 'danger')
        else:
            return redirect(url_for('audit_conference', address=address))

    prefill = request.args.get('address', '').strip().upper()
    unit = get_current_unit()
    sector = get_current_sector()
    shelves = db_mdb.get_all_shelves(unit=unit, sector=sector)

    shelf_map = {}
    for s in sorted(shelves, key=lambda x: (x.get('zone', ''), x.get('module', ''))):
        zm = f"{s.get('zone', '')}-{s.get('module', '')}"
        if zm not in shelf_map:
            shelf_map[zm] = int(s.get('levels', 1) or 1)

    zone_module_data = [
        {'address': zm, 'levels': list(range(levels, 0, -1))}
        for zm, levels in shelf_map.items()
    ]

    return render_template('audit.html',
                           conference_mode=False,
                           zone_module_data=zone_module_data,
                           prefill=prefill)


@app.route('/audit/history', methods=['GET'])
@login_required
def audit_history_list():
    """Retorna as ultimas conferencias (escopo global por unidade/setor)."""
    unit = get_current_unit()
    sector = get_current_sector()
    items = _audit_history_list(unit=unit, sector=sector, limit=AUDIT_HISTORY_LIMIT)

    response_items = []
    for item in items:
        response_items.append({
            'draft_id': item.get('draft_id'),
            'address': item.get('address'),
            'username': item.get('username'),
            'status': item.get('status', 'draft'),
            'scan_count': int(item.get('scan_count', 0) or 0),
            'updated_at': item.get('updated_at', ''),
            'updated_at_br': _to_br_datetime(item.get('updated_at', '')),
            'unit': item.get('unit', ''),
            'sector': item.get('sector', ''),
        })

    return jsonify({'ok': True, 'items': response_items})


@app.route('/audit/history/<draft_id>', methods=['GET'])
@login_required
def audit_history_get(draft_id):
    """Retorna um rascunho de conferencia para retomada."""
    unit = get_current_unit()
    sector = get_current_sector()
    item = _audit_history_get(draft_id=draft_id, unit=unit, sector=sector)
    if not item:
        return jsonify({'ok': False, 'message': 'Rascunho não encontrado.'}), 404

    return jsonify({'ok': True, 'item': item})


@app.route('/audit/history/save', methods=['POST'])
@login_required
def audit_history_save():
    """Salva/atualiza rascunho de conferencia durante a bipagem."""
    payload = request.get_json(silent=True) if request.is_json else request.form.to_dict(flat=True)
    payload = payload or {}

    address = str(payload.get('address', '')).strip().upper()
    if not address:
        return jsonify({'ok': False, 'message': 'Endereço é obrigatório.'}), 400

    scanned_raw = payload.get('scanned_ids', [])
    scanned_list = _coerce_scanned_ids(scanned_raw)
    reason = str(payload.get('reason', 'autosave')).strip().lower()
    draft_id = str(payload.get('draft_id', '')).strip()

    # Proteção contra perda de bipagens: autosave vazio não pode sobrescrever
    # um rascunho existente, exceto quando o usuário usa a ação explícita de limpar.
    if not scanned_list and reason != 'clear':
        unit = get_current_unit()
        sector = get_current_sector()
        existing = _audit_history_get(draft_id=draft_id, unit=unit, sector=sector) if draft_id else None
        if existing and _coerce_scanned_ids(existing.get('scanned_ids', [])):
            return jsonify({
                'ok': True,
                'draft_id': existing.get('draft_id'),
                'scan_count': int(existing.get('scan_count', 0) or 0),
                'updated_at': existing.get('updated_at', ''),
                'updated_at_br': _to_br_datetime(existing.get('updated_at', '')),
                'preserved': True,
            })

        return jsonify({
            'ok': True,
            'draft_id': draft_id,
            'scan_count': 0,
            'updated_at': '',
            'updated_at_br': '',
            'skipped': True,
        })

    item = _audit_history_upsert(
        draft_id=draft_id,
        address=address,
        scanned_ids=scanned_list,
        username=session.get('user', ''),
        unit=get_current_unit(),
        sector=get_current_sector(),
        status='draft',
    )

    return jsonify({
        'ok': True,
        'draft_id': item.get('draft_id'),
        'scan_count': int(item.get('scan_count', 0) or 0),
        'updated_at': item.get('updated_at', ''),
        'updated_at_br': _to_br_datetime(item.get('updated_at', '')),
    })


@app.route('/audit/history/complete', methods=['POST'])
@login_required
def audit_history_complete():
    """Marca um rascunho de conferencia como concluido."""
    payload = request.get_json(silent=True) if request.is_json else request.form.to_dict(flat=True)
    payload = payload or {}

    address = str(payload.get('address', '')).strip().upper()
    if not address:
        return jsonify({'ok': False, 'message': 'Endereço é obrigatório.'}), 400

    scanned_raw = payload.get('scanned_ids', [])
    draft_id = str(payload.get('draft_id', '')).strip()
    result_summary = payload.get('result_summary') if isinstance(payload.get('result_summary'), dict) else None

    item = _audit_history_upsert(
        draft_id=draft_id,
        address=address,
        scanned_ids=scanned_raw,
        username=session.get('user', ''),
        unit=get_current_unit(),
        sector=get_current_sector(),
        status='completed',
        result_summary=result_summary,
    )

    return jsonify({'ok': True, 'draft_id': item.get('draft_id')})


@app.route('/audit/<path:address>', methods=['GET', 'POST'])
@login_required
def audit_conference(address):
    """Modo de conferência double-checking para um endereço."""
    access_denied = require_feature_access('audit', 'Acesso restrito à conferência de endereço.')
    if access_denied:
        return access_denied
    address = address.strip().upper()
    unit = get_current_unit()
    sector = get_current_sector()

    positions = get_positions_for_address(address, unit, sector)
    if not positions:
        flash(f'Endereço "{address}" não encontrado no cadastro de prateleiras.', 'warning')
        return redirect(url_for('audit_select'))

    # Pedidos esperados: ativos nas posições do endereço
    expected_orders = {}
    for pos in positions:
        for o in db_mdb.get_orders_by_position(pos, unit=unit, sector=sector):
            if o.get('status') == 'add':
                expected_orders[o['order_id']] = o

    result = None
    scanned_raw = ''

    resume_draft_id = request.args.get('resume_draft', '').strip()

    if request.method == 'POST':
        scanned_raw = request.form.get('scanned_ids', '')
        draft_id = request.form.get('draft_id', '').strip()
        scanned = [line.strip() for line in scanned_raw.splitlines() if line.strip()]
        scanned_set = set(scanned)
        expected_set = set(expected_orders.keys())

        ok = [oid for oid in scanned if oid in expected_set]
        missing = [expected_orders[oid] for oid in expected_set if oid not in scanned_set]

        wrong_location = []
        not_found = []
        for oid in scanned:
            if oid not in expected_set:
                order = db_mdb.get_order_by_id(oid, unit=unit)
                if order and order.get('status') == 'add':
                    wrong_location.append({'scanned_id': oid, 'order': order})
                else:
                    not_found.append(oid)

        result = {
            'ok': ok,
            'missing': missing,
            'wrong_location': wrong_location,
            'not_found': not_found,
            'total_scanned': len(scanned),
            'total_expected': len(expected_orders),
        }

        _audit_history_upsert(
            draft_id=draft_id,
            address=address,
            scanned_ids=scanned,
            username=session.get('user', ''),
            unit=unit,
            sector=sector,
            status='completed',
            result_summary={
                'ok': len(ok),
                'missing': len(missing),
                'wrong_location': len(wrong_location),
                'not_found': len(not_found),
                'total_scanned': len(scanned),
                'total_expected': len(expected_orders),
            },
        )

        status_str = (
            f'{len(ok)} OK, {len(missing)} faltando, '
            f'{len(wrong_location)} endereço errado, {len(not_found)} não cadastrado'
        )
        db_mdb.add_movement(
            username=session.get('user'),
            action='audit',
            position=address,
            order_id='',
            box='',
            details=f'Conferência {address}: {status_str}',
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            unit=unit,
            sector=sector or DEFAULT_SECTOR
        )

    return render_template('audit.html',
                           conference_mode=True,
                           address=address,
                           positions=positions,
                           expected_orders=list(expected_orders.values()),
                           scanned_raw=scanned_raw,
                           result=result,
                           resume_draft_id=resume_draft_id,
                           history_enabled=True,
                           history_limit=AUDIT_HISTORY_LIMIT)


# ============================================================================
# TELEGRAM — ROTAS
# ============================================================================

@app.route('/telegram/send-conference-csv', methods=['POST'])
@login_required
def telegram_send_conference_csv():
    """Envia o CSV de conferência para o Telegram (acessível a todos os usuários)."""
    import telegram_notifier as tg

    if not tg.is_configured():
        return jsonify({'ok': False, 'message': 'Telegram não configurado no servidor.'})

    address = request.form.get('address', '').strip().upper()
    scanned_ids_raw = request.form.get('scanned_ids', '')

    if not address:
        return jsonify({'ok': False, 'message': 'Endereço não informado.'})

    unit = get_current_unit()
    sector = get_current_sector()

    scanned = [line.strip() for line in scanned_ids_raw.splitlines() if line.strip()]
    scanned_set = set(scanned)

    positions = get_positions_for_address(address, unit, sector)
    expected_orders = {}
    for pos in positions:
        for o in db_mdb.get_orders_by_position(pos, unit=unit, sector=sector):
            if o.get('status') == 'add':
                expected_orders[o['order_id']] = o

    expected_set = set(expected_orders.keys())
    ok_list = [oid for oid in scanned if oid in expected_set]
    missing = [expected_orders[oid] for oid in expected_set if oid not in scanned_set]

    wrong_location = []
    not_found = []
    for oid in scanned:
        if oid not in expected_set:
            order = db_mdb.get_order_by_id(oid, unit=unit)
            if order and order.get('status') == 'add':
                wrong_location.append({'scanned_id': oid, 'order': order})
            else:
                not_found.append(oid)

    result = {
        'ok': ok_list,
        'missing': missing,
        'wrong_location': wrong_location,
        'not_found': not_found,
    }

    csv_bytes = tg.build_conference_csv(address, result, scanned_by=session.get('user', ''))
    filename = f'conferencia_{address}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    caption = (
        f'📋 Conferência <b>{address}</b>\n'
        f'✅ OK: {len(ok_list)} | ❌ Faltando: {len(missing)} | '
        f'⚠️ End. errado: {len(wrong_location)} | ❓ Não cadastrado: {len(not_found)}\n'
        f'👤 Por: {session.get("user", "")} | {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )

    send_ok, msg = tg.send_document(csv_bytes, filename, caption=caption)
    if send_ok:
        wms_logger.info(
            f'TELEGRAM | CSV conferência {address} enviado por {session.get("user")}'
        )
        return jsonify({'ok': True, 'message': 'Relatório enviado para o Telegram!'})
    return jsonify({'ok': False, 'message': f'Erro ao enviar: {msg}'})


@app.route('/telegram/test', methods=['POST'])
@login_required
def telegram_test():
    """Testa a conexão com o Telegram (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'})

    import telegram_notifier as tg

    if not tg.is_configured():
        return jsonify({
            'ok': False,
            'message': 'Credenciais ausentes no .env (TELEGRAM_BOT_TOKEN e TELEGRAM_CHAT_ID).'
        })

    test_msg = (
        f'✅ <b>WMS — Teste de Conexão</b>\n'
        f'Integração com o Telegram funcionando!\n'
        f'🕒 {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
    )
    send_ok, err = tg.send_message(test_msg)
    if send_ok:
        wms_logger.info(f'TELEGRAM | Teste de conexão bem-sucedido por {session.get("user")}')
        return jsonify({'ok': True, 'message': 'Mensagem de teste enviada com sucesso!'})
    return jsonify({'ok': False, 'message': f'Erro: {err}'})


@app.route('/telegram/send-daily-report', methods=['POST'])
@login_required
def telegram_send_daily_report_manual():
    """Envia o relatório diário manualmente (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'})

    import telegram_notifier as tg

    if not tg.is_configured():
        return jsonify({'ok': False, 'message': 'Telegram não configurado.'})

    try:
        try:
            all_orders = db_mdb.get_all_orders(status_filter='add')
        except TypeError:
            all_orders = db_mdb.get_all_orders()
        thresholds = load_time_thresholds()
        msg = tg.build_daily_report_message(
            all_orders, thresholds,
            unit=get_current_unit(),
            sector=session.get('sector', ''),
        )
        send_ok, err = tg.send_message(msg)
        if send_ok:
            wms_logger.info(f'TELEGRAM | Relatório diário manual por {session.get("user")}')
            return jsonify({'ok': True, 'message': 'Relatório enviado!'})
        return jsonify({'ok': False, 'message': f'Erro: {err}'})
    except Exception as exc:
        return jsonify({'ok': False, 'message': str(exc)})


@app.route('/telegram/send-period-report', methods=['POST'])
@login_required
def telegram_send_period_report():
    """Gera e envia o relatório de período (CSV) para o Telegram (somente admin)."""
    if not is_admin_user():
        return jsonify({'ok': False, 'message': 'Acesso restrito a administradores.'})

    import telegram_notifier as tg
    from datetime import datetime as _dt

    if not tg.is_configured():
        return jsonify({'ok': False, 'message': 'Telegram não configurado.'})

    # Datas vindas do formulário (YYYY-MM-DD) ou fallback para período configurado
    raw_from = request.form.get('date_from', '').strip()
    raw_to   = request.form.get('date_to',   '').strip()

    try:
        if raw_from and raw_to:
            dt_from = _dt.strptime(raw_from, '%Y-%m-%d')
            dt_to   = _dt.strptime(raw_to,   '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        else:
            cfg = load_telegram_config()
            dt_from, dt_to = tg.resolve_report_period(cfg)
    except Exception as exc:
        return jsonify({'ok': False, 'message': f'Data inválida: {exc}'})

    try:
        try:
            all_orders = db_mdb.get_all_orders()
        except Exception:
            all_orders = []

        added, removed, active = tg.filter_orders_by_period(all_orders, dt_from, dt_to)
        unit = get_current_unit()
        csv_bytes = tg.build_period_report_csv(added, removed, active, dt_from, dt_to, unit=unit)
        caption   = tg.build_period_report_message(added, removed, active, dt_from, dt_to, unit=unit)
        fname     = f'relatorio_{dt_from.strftime("%Y%m%d")}_{dt_to.strftime("%Y%m%d")}.csv'

        send_ok, err = tg.send_document(csv_bytes, fname, caption=caption)
        if send_ok:
            wms_logger.info(
                f'TELEGRAM | Relatório de período {dt_from.strftime("%d/%m/%Y")}→{dt_to.strftime("%d/%m/%Y")}'
                f' enviado por {session.get("user")}'
            )
            return jsonify({'ok': True, 'message': f'Relatório enviado! ({len(added)} cadastrados, {len(removed)} retirados, {len(active)} ativos)'})
        return jsonify({'ok': False, 'message': f'Erro ao enviar: {err}'})
    except Exception as exc:
        return jsonify({'ok': False, 'message': str(exc)})


# ============================================================================
# MANIPULAÇÃO DE ERROS
# ============================================================================

@app.errorhandler(404)
def not_found(error):
    """Página não encontrada"""
    return render_template('error.html', 
                         title='404 - Página não encontrada',
                         message='A página que você procura não existe.'), 404

@app.errorhandler(500)
def internal_error(error):
    """Erro interno do servidor"""
    return render_template('error.html',
                         title='500 - Erro interno',
                         message='Ocorreu um erro interno no servidor.'), 500


