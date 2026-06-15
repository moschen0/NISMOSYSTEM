from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import NamedTuple, Union
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from parser_siou import PATH_OS_SIOU

# ---------------------------------------------------------------------------
# Integração com o banco WMS (para buscar a "OS OPTO" pelo ID do serviço)
# ---------------------------------------------------------------------------
# O módulo db_mdb vive na pasta WMS_SISTEMA (irmã de OPTO_INTEGRATIONS).
_WMS_SISTEMA_DIR = Path(__file__).parent.parent / "WMS_SISTEMA"
if _WMS_SISTEMA_DIR.exists() and str(_WMS_SISTEMA_DIR) not in sys.path:
    sys.path.insert(0, str(_WMS_SISTEMA_DIR))

try:
    import db_mdb  # type: ignore
except Exception as _db_exc:  # pragma: no cover - ambiente sem WMS/pyodbc
    db_mdb = None
    _DB_IMPORT_ERROR = _db_exc
else:
    _DB_IMPORT_ERROR = None

# ---------------------------------------------------------------------------
# Lookup de-para via planilhas de referência
# ---------------------------------------------------------------------------

# CODIGO TESTE 2BA-27352


class LookupSpec(NamedTuple):
    """Especifica um de-para: pega o valor no campo `field` do .txt,
    busca esse código em `table` (col A) e retorna o valor da `result_col`.
    Se `required=True` e o código não for encontrado, lança ValueError."""
    field: int        # índice no .txt (ex: 35 = codigo_produto_od)
    table: str        # "BD_PROD" ou "DB_TRAT"
    result_col: str   # letra da coluna de retorno: "C", "D" ou "E"
    required: bool = False


class MapSpec(NamedTuple):
    """Mapeamento inline: pega o valor no campo `field` do .txt e aplica
    o dicionário `mapping`. Valor não encontrado retorna `default`."""
    field: int
    mapping: dict
    default: str = ""


class MultiLookupSpec(NamedTuple):
    """Varre `fields` em ordem, busca cada código em `table` (col A) e
    retorna o valor da `result_col` do primeiro que tiver correspondência.
    Se `required=True` e nenhum campo produzir resultado, lança ValueError."""
    fields: tuple        # ex: (66, 67, 68, 69, 70)
    table: str           # "BD_PROD" ou "DB_TRAT"
    result_col: str      # letra da coluna de retorno: "C", "D" ou "E"
    required: bool = False


_SCRIPT_DIR = Path(__file__).parent
BD_PROD_PATH = _SCRIPT_DIR / "BD_PROD.xlsx"
DB_TRAT_PATH = _SCRIPT_DIR / "DB_TRAT.xlsx"

# ---------------------------------------------------------------------------
# Destino das planilhas por empresa (prefixo da "OS OPTO")
# ---------------------------------------------------------------------------
# prefixo → (código_empresa, caminho_base_de_rede)
COMPANY_MAP: dict[str, tuple[str, Path]] = {
    "2BA": ("MASTER", Path(r"\\192.168.1.210\apps master\Planilhas OPTO\2BA - MASTER")),
    "6VA": ("MATRIX", Path(r"\\192.168.1.210\apps master\Planilhas OPTO\6VA - MATRIX")),
    "9MA": ("AMX", Path(r"\\192.168.1.210\apps master\Planilhas OPTO\9MA - AMX")),
}

# Prefixo/pasta de fallback para OS sem "OS OPTO" cadastrada no WMS.
SEM_OPTO_PREFIX = "SEM_OPTO"
SEM_OPTO_BASE = _SCRIPT_DIR / "SEM_OPTO"

# Nomes dos meses por extenso (índice 1 = Janeiro)
MESES_PT = [
    "",
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def _load_bd_prod() -> dict[str, dict[str, str]]:
    """Carrega BD_PROD.xlsx → {codigo: {"C": tipo_lente, "D": foto, "E": material}}."""
    if not BD_PROD_PATH.exists():
        raise FileNotFoundError(f"BD_PROD.xlsx não encontrado em {BD_PROD_PATH}")
    wb = openpyxl.load_workbook(BD_PROD_PATH, read_only=True, data_only=True)
    ws = wb.active
    result: dict[str, dict[str, str]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # pula cabeçalho
            continue
        if row[0] is None:
            continue
        key = str(row[0]).strip()
        result[key] = {
            "C": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
            "D": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
            "E": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
        }
    wb.close()
    return result


def _load_db_trat() -> dict[str, str]:
    """Carrega DB_TRAT.xlsx → {codigo: valor_col_C (TRAT OPTO)}."""
    if not DB_TRAT_PATH.exists():
        raise FileNotFoundError(f"DB_TRAT.xlsx não encontrado em {DB_TRAT_PATH}")
    wb = openpyxl.load_workbook(DB_TRAT_PATH, read_only=True, data_only=True)
    ws = wb.active
    result: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # pula cabeçalho
            continue
        if row[0] is None:
            continue
        key = str(row[0]).strip()
        result[key] = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
    wb.close()
    return result


BD_PROD: dict[str, dict[str, str]] = _load_bd_prod()
DB_TRAT: dict[str, str] = _load_db_trat()

# ---------------------------------------------------------------------------
# Cabeçalhos da planilha de importação do OPTO (ordem fixa)
# ---------------------------------------------------------------------------
OPTO_HEADERS = [
    "Código da OS",
    "Tratamento",
    "Quantidade (0.5 ou 1.0)",
    "Fabricante",
    "Tipo de Lente",
    "Observação Lente/Serviço",
    "Fotossensibilidade",
    "Inteira/Recortada",
    "Material",
    "OS do Cliente",
    "Esférico R",
    "Cilíndrico R",
    "Eixo R",
    "Adição R",
    "Esférico L",
    "Cilíndrico L",
    "Eixo L",
    "Adição L",
    "Observações Gerais",
]

# ---------------------------------------------------------------------------
# Especificação de colunas:
#   int      → buscar pelo índice no arquivo .txt do SIOU
#   str      → valor literal fixo (não busca no .txt)
#   MapSpec  → busca índice no .txt e aplica mapeamento inline
# ---------------------------------------------------------------------------
_QTD_MAP = {"1": "1", "2": "0.5", "3": "0.5 Esquerdo"}

COLUMN_SPEC: list[Union[int, str, LookupSpec, MapSpec, MultiLookupSpec]] = [
    "",                              # Código da OS  → preenchido com a OS OPTO buscada no WMS (field 0 → order_id)
    MultiLookupSpec((66, 67, 68, 69, 70), "DB_TRAT", "C", required=True),  # Tratamento → primeiro cod_trat que existir no DB_TRAT
    MapSpec(1, _QTD_MAP, default=""),      # Quantidade            → 1→"1" | 2→"0.5" | 3→"0.5 Esquerdo"
    "Outros",                        # Fabricante            → literal fixo
    LookupSpec(35, "BD_PROD", "C"),  # Tipo de Lente         → codigo_produto_od → col C BD_PROD
    "Sem Observações",               # Observação Lente/Serv → observacoes
    LookupSpec(35, "BD_PROD", "D"),  # Fotossensibilidade    → codigo_produto_od → col D BD_PROD
    "Inteira",                       # Inteira/Recortada     → tipo_armacao
    LookupSpec(35, "BD_PROD", "E"),  # Material              → codigo_produto_od → col E BD_PROD
    0,                               # OS do Cliente         → os_laboratorio
    3,                               # Esférico R            → esf_od
    4,                               # Cilíndrico R          → cil_od
    5,                               # Eixo R                → eixo_od
    6,                               # Adição R              → adicao_od
    7,                               # Esférico L            → esf_oe
    8,                               # Cilíndrico L          → cil_oe
    9,                               # Eixo L                → eixo_oe
    10,                              # Adição L              → adicao_oe
    "POSITION",                     # Observações Gerais    → será preenchido com o endereço (posição) do WMS
]

assert len(COLUMN_SPEC) == len(OPTO_HEADERS), (
    "COLUMN_SPEC e OPTO_HEADERS estão com tamanhos diferentes."
)

# ---------------------------------------------------------------------------
# Funções principais
# ---------------------------------------------------------------------------

def init_database() -> None:
    """Valida que o módulo de banco do WMS está disponível para o integrador.

    O banco ativo deve ser definido pelo processo chamador antes do uso desta
    função. Em especial, o servidor de teste já faz `db_mdb.switch_database()`
    no bootstrap. Evitar troca implícita aqui impede que uma chamada de geração
    OPTO altere o modo global da aplicação em produção.
    """
    if db_mdb is None:
        raise RuntimeError(
            "Módulo db_mdb do WMS não pôde ser importado "
            f"({_DB_IMPORT_ERROR}). Verifique a pasta WMS_SISTEMA e o pyodbc."
        )


def fetch_os_opto(order_id: str) -> str:
    """Busca no banco WMS a "OS OPTO" cadastrada para o serviço `order_id`.
    Retorna string vazia se o pedido não existir ou não tiver os_opto."""
    if not order_id or db_mdb is None:
        return ""
    order = db_mdb.get_order_by_id(order_id)
    if not order:
        return ""
    return str(order.get("os_opto") or "").strip().upper()


def extract_prefix(os_code: str) -> str:
    """Extrai o prefixo da empresa da "OS OPTO" (ex: '2BA-27352' → '2BA').
    Retorna string vazia se não houver prefixo."""
    code = (os_code or "").strip()
    if "-" not in code:
        return ""
    return code.split("-", 1)[0].strip().upper()


def build_output_path(prefix: str) -> Path:
    """Monta o caminho final da planilha para a empresa do `prefix`, criando a
    estrutura de pastas ANO / Mês (por extenso) / Dia. Retorna o caminho do
    arquivo `{prefixo}_{empresa}_{HH-MM-SS}.xlsx`."""
    if prefix in COMPANY_MAP:
        company_code, base_path = COMPANY_MAP[prefix]
        label = f"{prefix}_{company_code}"
    else:  # SEM_OPTO
        base_path = SEM_OPTO_BASE
        label = prefix
    now = datetime.now()
    day_dir = base_path / str(now.year) / MESES_PT[now.month] / f"{now.day:02d}"
    day_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{label}_{now:%H-%M-%S}.xlsx"
    return day_dir / filename


def find_txt(os_id: str) -> Path:
    """Localiza {os_id}.txt em PATH_OS_SIOU (busca recursiva)."""
    base = Path(PATH_OS_SIOU)
    matches = list(base.rglob(f"{os_id}.txt"))
    if not matches:
        raise FileNotFoundError(
            f"Arquivo '{os_id}.txt' não encontrado em {PATH_OS_SIOU}"
        )
    return matches[0]


def parse_txt(path: Path) -> dict[int, str]:
    """Lê a linha única do .txt SIOU e retorna {índice: valor}."""
    raw = path.read_text(encoding="utf-8", errors="replace").strip()
    fields = raw.split(",")
    return {i: v.strip() for i, v in enumerate(fields)}


def build_row(fields: dict[int, str]) -> list[str]:
    """
    Monta uma linha da planilha aplicando COLUMN_SPEC.
    - int:      busca pelo índice no .txt; ausente → "0"
    - str:      valor literal fixo
    - LookupSpec: busca o código no .txt e faz de-para na tabela de referência; ausente → ""
    - MapSpec:  busca pelo índice no .txt e aplica mapeamento inline; ausente → default
    """
    row = []
    for spec in COLUMN_SPEC:
        if isinstance(spec, MultiLookupSpec):
            value = ""
            tried_codes: list[str] = []
            for f in spec.fields:
                code = fields.get(f, "").strip()
                if not code:
                    continue
                tried_codes.append(code)
                if spec.table == "BD_PROD":
                    candidate = BD_PROD.get(code, {}).get(spec.result_col, "")
                else:  # "DB_TRAT"
                    candidate = DB_TRAT.get(code, "")
                if candidate:
                    value = candidate
                    break
            if spec.required and not value:
                raise ValueError(
                    f"Tratamento não encontrado. Códigos testados nos campos "
                    f"{list(spec.fields)}: {tried_codes}. Verifique o DB_TRAT.xlsx."
                )
            row.append(value)
        elif isinstance(spec, LookupSpec):
            code = fields.get(spec.field, "").strip()
            if spec.table == "BD_PROD":
                row.append(BD_PROD.get(code, {}).get(spec.result_col, ""))
            else:  # "DB_TRAT"
                value = DB_TRAT.get(code, "")
                if spec.required and not value:
                    raise ValueError(
                        f"Tratamento não encontrado para o código '{code}' (campo {spec.field}). "
                        "Verifique o DB_TRAT.xlsx."
                    )
                row.append(value)
        elif isinstance(spec, MapSpec):
            raw = fields.get(spec.field, "").strip()
            row.append(spec.mapping.get(raw, spec.default))
        elif isinstance(spec, int):
            row.append(fields.get(spec, "0"))
        else:
            row.append(spec)
    return row


def export_excel(rows: list[list[str]], output_path: Path) -> Path:
    """
    Exporta as linhas acumuladas para `output_path`.
    Retorna o caminho do arquivo gerado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importação OS"

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(OPTO_HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append(row)

    # Ajusta largura das colunas automaticamente
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    wb.save(output_path)
    return output_path


def _write_error_report(order_id: str, os_opto: str, reason: str, detail: str) -> Path:
    """Grava .txt de erro de integracao para a OS informada.

    Pasta:   OPTO_INTEGRATIONS/ERROS_INTEGRACAO/<ano>/<mes>/<dia>/
    Arquivo: <order_id>_erro_integracao.txt
    """
    now = datetime.now()
    err_dir = (
        _SCRIPT_DIR
        / "ERROS_INTEGRACAO"
        / str(now.year)
        / MESES_PT[now.month]
        / f"{now.day:02d}"
    )
    err_dir.mkdir(parents=True, exist_ok=True)
    err_path = err_dir / f"{order_id}_erro_integracao.txt"

    lines = [
        f"ERRO DE INTEGRACAO OPTO -- {now:%d/%m/%Y %H:%M:%S}",
        "=" * 55,
        f"OS WMS (order_id) : {order_id}",
        f"OS OPTO           : {os_opto or '(nao cadastrada)'}",
        f"Motivo            : {reason}",
        "",
        "Detalhes:",
        f"  {detail}",
        "",
        "Possiveis causas e acoes:",
    ]

    if reason == "TXT_NAO_ENCONTRADO":
        lines += [
            "  1. O arquivo .txt do SIOU ainda nao foi gerado para esta OS.",
            "     -> Verifique se a OS foi processada e impressa no SIOU.",
            f"  2. O arquivo '{order_id}.txt' pode estar em subpasta diferente.",
            f"     -> Caminho pesquisado: {PATH_OS_SIOU}",
            "  3. Permissao de leitura no compartilhamento de rede negada.",
            "     -> Confirme acesso a " + str(PATH_OS_SIOU),
        ]
    elif reason == "TRATAMENTO_NAO_ENCONTRADO":
        lines += [
            "  1. O codigo de tratamento da lente nao esta cadastrado no DB_TRAT.xlsx.",
            f"     -> Detalhe: {detail}",
            "  2. Abra DB_TRAT.xlsx e adicione a linha com o codigo e o tratamento OPTO.",
            "  3. Se o produto nao usa tratamento, verifique se o campo de tratamento",
            "     esta em branco no SIOU (campos 66-70).",
        ]
    elif reason == "OS_OPTO_AUSENTE":
        lines += [
            "  1. O campo 'OS OPTO' nao foi preenchido no WMS ao enderecar este servico.",
            "     -> Acesse o WMS, localize o pedido e adicione o codigo OS OPTO.",
            "  2. Confirme que o setor AR tem permissao para o campo OS OPTO.",
        ]
    else:
        lines += [
            "  1. Erro inesperado durante o processamento.",
            f"     -> {detail}",
            "  2. Verifique os logs do WMS para mais detalhes.",
        ]

    err_path.write_text("\n".join(lines), encoding="utf-8")
    return err_path


def generate_scheduled_export(companies: list[str] = None, date_str: str = None) -> dict:
    """Gera exportacao programada para as empresas (prefixos) informadas.

    - Consulta db_mdb.get_orders_by_date(date_str) para pedidos ativos do dia.
    - Pedidos com todas as informacoes necessarias sao exportados para planilha OPTO.
    - Pedidos exportados com sucesso recebem registro 'opto_auto_export' (username=AUTO)
      no historico de movimentacoes do WMS.
    - Pedidos com falha (txt ausente, tratamento nao cadastrado, OS OPTO vazia)
      geram <order_id>_erro_integracao.txt em ERROS_INTEGRACAO/<ano>/<mes>/<dia>/.
    - 'companies': lista de prefixos a incluir (ex: ['2BA','9MA']); None = todos.

    Retorna {'files': {caminho: n_linhas}, 'errors': [{order_id, os_opto, reason}]}.
    """
    init_database()
    if db_mdb is None:
        raise RuntimeError("db_mdb nao disponivel")

    if not date_str:
        date_str = datetime.now().strftime("%d/%m/%Y")

    try:
        orders = db_mdb.get_orders_by_date(date_str)
    except Exception:
        orders = []

    session_groups: dict[str, list[tuple]] = {}
    failed_orders: list[dict] = []

    for o in orders:
        order_id = str(o.get("order_id") or "").strip()
        os_opto  = str(o.get("os_opto")  or "").strip().upper()
        position = str(o.get("position") or "").strip()
        unit     = str(o.get("unit")     or "").strip()

        if not order_id:
            continue

        # Validacao 1: OS OPTO cadastrada no WMS
        if not os_opto:
            failed_orders.append({
                "order_id": order_id, "os_opto": os_opto,
                "reason": "OS_OPTO_AUSENTE",
                "detail": f"Pedido {order_id} ({position}) nao possui OS OPTO registrada no WMS.",
            })
            continue

        prefix = extract_prefix(os_opto)

        # Filtro por empresa solicitada
        if companies and prefix not in companies:
            continue

        # Validacao 2: arquivo .txt do SIOU
        try:
            txt_path = find_txt(order_id)
        except FileNotFoundError as exc:
            failed_orders.append({
                "order_id": order_id, "os_opto": os_opto,
                "reason": "TXT_NAO_ENCONTRADO",
                "detail": str(exc),
                "position": position, "unit": unit,
            })
            continue

        # Validacao 3: build_row (inclui lookup de tratamento no DB_TRAT)
        try:
            fields = parse_txt(txt_path)
            row = build_row(fields)
            row[0] = os_opto
            # Substitui o marcador 'POSITION' pela posição/endereço vindo do WMS
            try:
                row_index = len(row) - 1
                row[row_index] = position
            except Exception:
                pass
        except ValueError as exc:
            failed_orders.append({
                "order_id": order_id, "os_opto": os_opto,
                "reason": "TRATAMENTO_NAO_ENCONTRADO",
                "detail": str(exc),
                "position": position, "unit": unit,
            })
            continue
        except Exception as exc:
            failed_orders.append({
                "order_id": order_id, "os_opto": os_opto,
                "reason": "ERRO_INESPERADO",
                "detail": str(exc),
                "position": position, "unit": unit,
            })
            continue

        group_key = prefix if prefix in COMPANY_MAP else SEM_OPTO_PREFIX
        session_groups.setdefault(group_key, []).append((o, row))

    # Exporta planilhas por empresa
    result: dict = {}
    exported_orders: list[tuple] = []

    for prefix, items in session_groups.items():
        try:
            rows = [item[1] for item in items]
            out_path = build_output_path(prefix)
            export_excel(rows, out_path)
            result[str(out_path)] = len(rows)
            for o, _ in items:
                exported_orders.append((o, out_path))
        except Exception as exc:
            for o, _ in items:
                failed_orders.append({
                    "order_id": str(o.get("order_id") or ""),
                    "os_opto": str(o.get("os_opto") or ""),
                    "reason": "ERRO_INESPERADO",
                    "detail": f"Falha ao salvar planilha: {exc}",
                    "position": str(o.get("position") or ""),
                    "unit": str(o.get("unit") or ""),
                })

    # Marca pedidos exportados no historico do WMS (username=AUTO)
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    for o, out_path in exported_orders:
        try:
            db_mdb.add_movement(
                username="AUTO",
                action="opto_auto_export",
                position=str(o.get("position") or ""),
                order_id=str(o.get("order_id") or ""),
                box=str(o.get("box") or ""),
                details=f"AUTO | planilha: {Path(out_path).name}",
                timestamp=now_str,
                unit=str(o.get("unit") or db_mdb.DEFAULT_UNIT),
                sector=str(o.get("sector") or db_mdb.DEFAULT_SECTOR),
            )
        except Exception:
            pass

    # Grava arquivos de erro de integracao
    error_summary: list[dict] = []
    for err in failed_orders:
        error_summary.append({
            "order_id": err.get("order_id", ""),
            "os_opto": err.get("os_opto", ""),
            "reason": err.get("reason", ""),
        })
        try:
            _write_error_report(
                order_id=err["order_id"],
                os_opto=err.get("os_opto", ""),
                reason=err["reason"],
                detail=err["detail"],
            )
        except Exception:
            pass

    return {"files": result, "errors": error_summary}


# ---------------------------------------------------------------------------
# Loop de terminal
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 55)
    print("  Integrador SIOU → OPTO")
    print("  Digite o ID da OS para adicionar à planilha.")
    print("  Digite 'sair' para exportar e encerrar.")
    print("=" * 55)

    try:
        init_database()
    except Exception as e:
        print(f"  [ERRO] Não foi possível conectar ao banco WMS: {e}")
        return
    modo = "TESTE" if os.environ.get("WMS_MDB_PATH_TEST", "").strip() else "PRODUÇÃO"
    print(f"  Banco WMS: {modo} → {db_mdb.get_db_path()}")

    # Agrupa as linhas por prefixo de empresa: {prefixo: [row, ...]}
    session_groups: dict[str, list[list[str]]] = {}
    session_ids: list[str] = []

    while True:
        os_id = input("\nID da OS: ").strip()

        if os_id.lower() == "sair":
            total = sum(len(rows) for rows in session_groups.values())
            if total == 0:
                print("Nenhuma OS adicionada. Encerrando sem exportar.")
            else:
                print(f"\nExportando {total} OS(s) em {len(session_groups)} planilha(s):")
                for prefix, rows in session_groups.items():
                    output_path = export_excel(rows, build_output_path(prefix))
                    label = COMPANY_MAP[prefix][0] if prefix in COMPANY_MAP else prefix
                    print(f"  [{prefix} - {label}] {len(rows)} OS(s) → {output_path}")
            break

        if not os_id:
            continue

        if os_id in session_ids:
            print(f"  [AVISO] OS '{os_id}' já foi adicionada nesta sessão.")
            continue

        try:
            txt_path = find_txt(os_id)
            fields = parse_txt(txt_path)
            # Busca a OS OPTO no WMS pelo field 0 (os_laboratorio = order_id)
            order_key = fields.get(0, "").strip() or os_id
            os_opto = fetch_os_opto(order_key)
            prefix = extract_prefix(os_opto)
            if prefix not in COMPANY_MAP:
                prefix = SEM_OPTO_PREFIX
            row = build_row(fields)
            row[0] = os_opto          # Coluna A = OS OPTO cadastrada no WMS
            # Preenche a última coluna com a posição/endereço do WMS (se disponível)
            try:
                if db_mdb is not None:
                    order = db_mdb.get_order_by_id(order_key)
                    pos = str(order.get("position") or "").strip()
                else:
                    pos = ""
                row[-1] = pos
            except Exception:
                pass
            session_groups.setdefault(prefix, []).append(row)
            session_ids.append(os_id)
            total = sum(len(rows) for rows in session_groups.values())
            if prefix == SEM_OPTO_PREFIX:
                print(
                    f"  [AVISO] OS '{os_id}' sem OS OPTO no WMS "
                    f"(order_id '{order_key}'). Adicionada em '{SEM_OPTO_PREFIX}'. "
                    f"Total na sessão: {total}"
                )
            else:
                print(
                    f"  [OK] OS '{os_id}' adicionada → {os_opto} "
                    f"({prefix} - {COMPANY_MAP[prefix][0]}). Total na sessão: {total}"
                )
        except FileNotFoundError as e:
            print(f"  [ERRO] {e}")
        except Exception as e:
            print(f"  [ERRO] Falha ao processar OS '{os_id}': {e}")


if __name__ == "__main__":
    main()
